#!/usr/bin/env python3
# === WHM Email Management & Monitoring Script ===

import sys
import os
import socket
import re
import random
from datetime import datetime
import time

# استيراد الدوال المشتركة
from common_functions import *

# === دوال إدارة الإيميل الأساسية ===

def create_email_account(server, cpanel_user, email, password, quota=250):
    """إنشاء حساب إيميل جديد"""
    try:
        params = {
            "email": email,
            "password": password,
            "quota": quota
        }
        
        print(f"      🔍 Creating email account: {email}")
        print(f"      📋 Using cPanel API: Email::add_pop")
        print(f"      📊 Parameters: {params}")
        
        result = cpanel_api_call(server, cpanel_user, "Email", "add_pop", params)
        print(f"      📤 cPanel API response: {result}")
        
        if not result:
            return {"success": False, "error": "cPanel API returned no response"}
        
        if "error" in result:
            return {"success": False, "error": result["error"]}
            
        if "result" in result:
            if result["result"].get("status") == 1:
                return {"success": True, "message": "Email account created successfully"}
            else:
                error_msg = result["result"].get("errors", ["Unknown error"])[0]
                return {"success": False, "error": error_msg}
        
        # إذا لم تكن النتيجة في الشكل المتوقع، اعيد رسالة خطأ واضحة
        return {"success": False, "error": f"Unexpected response format: {result}"}
        
    except Exception as e:
        logging.error(f"Error creating email account: {str(e)}")
        return {"success": False, "error": str(e)}

def delete_email_account(server, cpanel_user, email):
    """حذف حساب إيميل"""
    try:
        params = {
            "email": email
        }
        
        result = cpanel_api_call(server, cpanel_user, "Email", "delete_pop", params)
        
        if "error" in result:
            return {"success": False, "error": result["error"]}
            
        if "result" in result:
            if result["result"].get("status") == 1:
                return {"success": True, "message": "Email account deleted successfully"}
            else:
                error_msg = result["result"].get("errors", ["Unknown error"])[0]
                return {"success": False, "error": error_msg}
        
        return {"success": False, "error": "Unexpected response format"}
        
    except Exception as e:
        logging.error(f"Error deleting email account: {str(e)}")
        return {"success": False, "error": str(e)}

def change_email_password(server, cpanel_user, email, new_password):
    """تغيير كلمة مرور الإيميل"""
    try:
        # استخراج domain من الإيميل
        email_domain = email.split('@')[1] if '@' in email else None
        
        # البحث عن الحساب الصحيح للدومين
        if email_domain and email_domain != cpanel_user:
            # البحث عن الحساب الذي يحتوي على هذا الدومين
            try:
                # محاولة استخدام get_domain_info للعثور على الحساب الصحيح
                domain_info_result = whm_api_call(server, "get_domain_info", {"domain": email_domain})
                if domain_info_result and "data" in domain_info_result:
                    domain_data = domain_info_result["data"]
                    if "domains" in domain_data:
                        for domain_info in domain_data["domains"]:
                            if domain_info.get("domain") == email_domain:
                                correct_user = domain_info.get("user")
                                if correct_user and correct_user != cpanel_user:
                                    print(f"   🔍 Using correct cPanel user: {correct_user} for {email_domain}")
                                    cpanel_user = correct_user
                                break
            except Exception as e:
                print(f"   ⚠️  Could not determine correct user for {email_domain}, using: {cpanel_user}")
        
        params = {
            "email": email,
            "password": new_password
        }
        
        # البحث عن الحساب الصحيح للدومين
        correct_user = cpanel_user
        email_domain = email.split('@')[1] if '@' in email else None
        
        # إذا كان الإيميل ينتهي بـ diyaraltameer.com، استخدم egypetoo
        if email_domain == "diyaraltameer.com":
            print(f"      🔍 diyaraltameer.com is addon domain, using egypetoo user")
            correct_user = "egypetoo"
        elif email_domain and email_domain != cpanel_user:
            try:
                # البحث عن الحساب الذي يحتوي على هذا الدومين
                domain_info_result = whm_api_call(server, "get_domain_info", {"domain": email_domain})
                if domain_info_result and "data" in domain_info_result:
                    domain_data = domain_info_result["data"]
                    if "domains" in domain_data:
                        for domain_info in domain_data["domains"]:
                            if domain_info.get("domain") == email_domain:
                                found_user = domain_info.get("user")
                                if found_user and found_user != cpanel_user:
                                    print(f"      🔍 Found correct user: {found_user} for {email_domain}")
                                    correct_user = found_user
                                break
            except Exception as e:
                print(f"      ⚠️  Could not determine correct user for {email_domain}, using: {cpanel_user}")
        
        # استخدام cPanel API مباشرة مع apiversion 2 (مثل delete_email_account)
        result = cpanel_api_call(server, cpanel_user, "Email", "passwd_pop", {
            "email": email,
            "password": new_password,
            "apiversion": "2"
        })
        
        if "error" in result:
            return {"success": False, "error": result["error"]}
            
        # محاولة معالجة الاستجابة بطرق مختلفة
        if "cpanelresult" in result:
            cpanel_result = result["cpanelresult"]
            if cpanel_result.get("data", {}).get("result") == "1":
                return {"success": True, "message": "Email password changed successfully"}
            else:
                error_msg = cpanel_result.get("data", {}).get("reason", "Unknown error")
                return {"success": False, "error": error_msg}
        
        # محاولة معالجة الاستجابة المباشرة
        if "result" in result:
            if result["result"].get("status") == 1:
                return {"success": True, "message": "Email password changed successfully"}
            else:
                error_msg = result["result"].get("errors", ["Unknown error"])[0]
                return {"success": False, "error": error_msg}
        
        # محاولة معالجة الاستجابة البسيطة
        if "status" in result and result["status"] == 1:
            return {"success": True, "message": "Email password changed successfully"}
        
        # إذا فشلت جميع المحاولات، طباعة الاستجابة الكاملة
        print(f"      ❌ All parsing attempts failed. Full response: {result}")
        return {"success": False, "error": f"Unexpected response format: {result}"}
        
    except Exception as e:
        logging.error(f"Error changing email password: {str(e)}")
        return {"success": False, "error": str(e)}

def get_email_usage(server, cpanel_user, email):
    """جلب معلومات استخدام الإيميل بطريقة محسنة"""
    try:
        # محاولة جلب البيانات من API مختلف أكثر دقة
        params = {
            "account": email
        }
        
        # تجربة APIs مختلفة للحصول على البيانات الصحيحة
        apis_to_try = [
            ("Email", "get_pop_quota"),
            ("Email", "get_disk_usage"), 
            ("Fileman", "get_disk_usage"),
            ("Quota", "getquotas"),
            ("Email", "list_pops_with_disk"),
            ("Email", "get_pop_statistics")
        ]
        
        for module, function in apis_to_try:
            try:
                if function == "list_pops_with_disk":
                    # استخدام list_pops مع معلومات الديسك
                    result = cpanel_api_call(server, cpanel_user, "Email", "list_pops", {"include_disk_usage": 1})
                elif function == "getquotas":
                    result = cpanel_api_call(server, cpanel_user, module, function, {})
                elif function == "get_disk_usage" and module == "Fileman":
                    # استخدام Fileman للحصول على استخدام الديسك
                    email_user = email.split('@')[0]
                    result = cpanel_api_call(server, cpanel_user, module, function, {"dir": f"mail/{email.split('@')[1]}/{email_user}"})
                elif function == "get_pop_statistics":
                    # احصائيات الإيميل
                    result = cpanel_api_call(server, cpanel_user, module, function, {"account": email})
                else:
                    result = cpanel_api_call(server, cpanel_user, module, function, params)
                
                if "error" not in result and "result" in result:
                    if function == "list_pops_with_disk":
                        # البحث عن الإيميل المحدد في النتائج
                        if "data" in result["result"]:
                            for email_data in result["result"]["data"]:
                                if email_data.get("email") == email:
                                    quota = email_data.get("diskquota", 0)
                                    used = email_data.get("diskused", 0)
                                    
                                    # معالجة البيانات
                                    quota_mb = round(float(quota) / (1024 * 1024), 1) if quota > 0 else 0
                                    used_mb = round(float(used) / (1024 * 1024), 1) if used > 0 else 0
                                    
                                    # إذا كانت الحصة 0، استخدم "Unlimited" أو قيمة افتراضية
                                    if quota_mb == 0:
                                        quota_display = "Unlimited"
                                        usage_percent = "N/A"
                                    else:
                                        usage_percent = f"{round((used_mb / quota_mb) * 100, 2)}%"
                                        quota_display = f"{quota_mb}MB"
                                    
                                    return {
                                        "quota": quota_display,
                                        "used": f"{used_mb}MB",
                                        "usage_percent": usage_percent
                                    }
                    elif function == "get_disk_usage" and module == "Fileman":
                        # معالجة Fileman disk usage
                        if "data" in result["result"]:
                            data = result["result"]["data"]
                            used = data.get("diskused", data.get("used", 0))
                            
                            if used:
                                used_mb = round(float(used) / (1024 * 1024), 1) if used > 0 else 0
                                
                                return {
                                    "quota": "Unlimited",
                                    "used": f"{used_mb}MB",
                                    "usage_percent": "N/A"
                                }
                    
                    elif function == "get_pop_statistics":
                        # معالجة احصائيات الإيميل
                        if "data" in result["result"]:
                            data = result["result"]["data"]
                            used = data.get("diskused", data.get("used", 0))
                            quota = data.get("quota", data.get("diskquota", 0))
                            
                            if used is not None:
                                used_mb = round(float(used) / (1024 * 1024), 1) if used > 0 else 0
                                quota_mb = round(float(quota) / (1024 * 1024), 1) if quota > 0 else 0
                                
                                if quota_mb == 0:
                                    quota_display = "Unlimited"
                                    usage_percent = "N/A"
                                else:
                                    usage_percent = f"{round((used_mb / quota_mb) * 100, 2)}%"
                                    quota_display = f"{quota_mb}MB"
                                
                                return {
                                    "quota": quota_display,
                                    "used": f"{used_mb}MB",
                                    "usage_percent": usage_percent
                                }
                    
                    else:
                        # معالجة APIs أخرى
                        data = result["result"].get("data", {})
                        if isinstance(data, list) and len(data) > 0:
                            data = data[0]
                        
                        quota = data.get("quota", data.get("diskquota", 0))
                        used = data.get("diskused", data.get("used", 0))
                        
                        if quota is not None or used is not None:
                            quota_mb = round(float(quota) / (1024 * 1024), 1) if quota and quota > 0 else 0
                            used_mb = round(float(used) / (1024 * 1024), 1) if used and used > 0 else 0
                            
                            if quota_mb == 0:
                                quota_display = "Unlimited"
                                usage_percent = "N/A"
                            else:
                                usage_percent = f"{round((used_mb / quota_mb) * 100, 2)}%"
                                quota_display = f"{quota_mb}MB"
                            
                            return {
                                "quota": quota_display,
                                "used": f"{used_mb}MB",
                                "usage_percent": usage_percent
                            }
            except Exception as api_error:
                logging.warning(f"API {module}::{function} failed: {str(api_error)}")
                continue
        
        # إذا فشلت جميع المحاولات، استخدم البيانات الأساسية
        emails = list_email_accounts(server, cpanel_user)
        for email_info in emails:
            if email_info.get("email") == email:
                quota = email_info.get("diskquota", 0)
                used = email_info.get("diskused", 0)
                
                quota_mb = round(quota / (1024 * 1024), 1) if quota > 0 else 0
                used_mb = round(used / (1024 * 1024), 1) if used > 0 else 0
                
                if quota_mb == 0:
                    quota_display = "Unlimited"
                    usage_percent = "N/A"
                else:
                    usage_percent = f"{round((used_mb / quota_mb) * 100, 2)}%"
                    quota_display = f"{quota_mb}MB"
                
                return {
                    "quota": quota_display,
                    "used": f"{used_mb}MB",
                    "usage_percent": usage_percent
                }
        
        return {"quota": "Unknown", "used": "Unknown", "usage_percent": "N/A"}
        
    except Exception as e:
        logging.error(f"Error getting email usage: {str(e)}")
        return {"quota": "Unknown", "used": "Unknown", "usage_percent": "N/A"}

# === دوال إدارة الإيميل المتقدمة ===
def create_single_email(servers):
    """إنشاء إيميل واحد"""
    print("\n📧 Create Single Email Account")
    print("=" * 50)
    
    # خيارات البحث
    print("🔍 Search Options:")
    print("1. 🌐 Search by domain")
    print("2. 📧 Search by email address")
    
    search_type = input("Choose search type (1-2, default 1): ").strip()
    
    if search_type == "2":
        # البحث بالإيميل مباشرة
        email_address = input("📧 Enter full email address: ").strip()
        if not email_address or "@" not in email_address:
            print("❌ Invalid email address format")
            return
            
        domain = email_address.split("@")[1]
        print(f"🌐 Extracted domain: {domain}")
        
        # البحث عن السيرفر
        print(f"\n🔍 Searching for email: {email_address}...")
        server, acct, server_name = find_server_by_email(email_address, servers)
        
        if not server:
            print("❌ Email not found on any server!")
            return
            
        cpanel_user = acct["user"]
        print(f"\n✅ Email found on Server {server_name}")
        print(f"📋 Domain: {domain}")
        print(f"👤 cPanel User: {cpanel_user}")
        
        # معلومات الإيميل الجديد
        email_prefix = email_address.split("@")[0]
        print(f"📧 Email prefix: {email_prefix}")
        
    else:
        # البحث بالدومين (الطريقة القديمة)
        domain = input("🌐 Enter domain: ").strip()
        if not domain:
            print("❌ Domain cannot be empty")
            return
    
    # خيارات البحث
    print("\n🔍 Search Options:")
    print("1. 🚀 Fast search (main domains only)")
    print("2. 🧠 Smart search (main domains first, then subdomains if needed)")
    print("3. 🔍 Full search (all domains + subdomains)")
    
    search_choice = input("Choose search type (1-3, default 2): ").strip()
    
    if search_choice == "1":
        search_mode = "fast"
    elif search_choice == "3":
        search_mode = "full"
    else:
        search_mode = "smart"
    
    print(f"\n🔍 Searching for domain: {domain}...")
    
    # اختيار دالة البحث المناسبة بناءً على نوع البحث
    if search_mode == "fast":
        print("🚀 Using fast search (main domains only)...")
        server, acct, server_name = find_server_by_domain_fast(domain, servers)
    elif search_mode == "full":
        print("🧠 Using full search (all domains + subdomains)...")
        server, acct, server_name = find_server_by_domain_full(domain, servers)
    else:  # smart mode
        print("🧠 Using smart search (main domains first, then subdomains if needed)...")
        server, acct, server_name = find_server_by_domain_smart(domain, servers)
    
    if not server:
        print("❌ Domain not found on any server!")
        return
        
    cpanel_user = acct["user"]
    print(f"\n✅ Domain found on Server {server_name}")
    print(f"📋 Domain: {domain}")
    print(f"👤 cPanel User: {cpanel_user}")
    
    # معلومات الإيميل الجديد
    email_prefix = input("\n📧 Email prefix (before @): ").strip()
    if not email_prefix:
        print("❌ Email prefix cannot be empty")
        return
        
    email_address = f"{email_prefix}@{domain}"
    
    # خيار كلمة المرور
    print(f"\n🔑 Password Options:")
    print("1. Enter password manually")
    print("2. Generate random strong password")
    password_choice = input("Choose option [1/2]: ").strip()
    
    if password_choice == "2":
        password = generate_strong_password(16)
        print(f"\n✅ Generated strong password: {password}")
    else:
        while True:
            password = get_secure_password()
            if not password:
                print("❌ Password cannot be empty!")
                continue
            
            if len(password) < 8:
                print("❌ Password must be at least 8 characters!")
                continue
                
            confirm_password = get_secure_password()
            if password != confirm_password:
                print("❌ Passwords do not match!")
                continue
            
            break
    
    quota = input("💾 Quota in MB (default: 250): ").strip() or "250"
    
    try:
        quota = int(quota)
    except ValueError:
        print("❌ Invalid quota, using default 250MB")
        quota = 250
    
    print(f"\n📋 Email Details:")
    print(f"   Email: {email_address}")
    print(f"   Password Length: {len(password)} characters")
    print(f"   Quota: {quota}MB")
    print(f"   Server: {server_name}")
    
    if confirm_action(f"Create email account {email_address}?"):
        result = create_email_account(server, cpanel_user, email_address, password, quota)
        
        if result["success"]:
            print("✅ Email account created successfully!")
            print("=" * 50)
            print(f"🌐 Domain: {domain}")
            print(f"📧 Email Address: {email_address}")
            print(f"🔑 Password: {password}")
            print(f"💻 Webmail URL: https://webmail.{domain}")
            print(f"🖥️  Server: {server_name} ({server['ip']})")
            print("=" * 50)
            logging.info(f"Email created: {email_address} on {domain}")
        else:
            print(f"❌ Failed to create email: {result['error']}")

def bulk_create_emails(servers):
    """إنشاء إيميلات متعددة"""
    print("\n📧 Bulk Create Email Accounts")
    print("=" * 50)
    
    domain = input("🌐 Enter domain: ").strip()
    if not domain:
        print("❌ Domain cannot be empty")
        return
    
    # خيارات البحث
    print("\n🔍 Search Options:")
    print("1. 🚀 Fast search (main domains only)")
    print("2. 🧠 Smart search (main domains first, then subdomains if needed)")
    print("3. 🔍 Full search (all domains + subdomains)")
    
    search_choice = input("Choose search type (1-3, default 2): ").strip()
    
    if search_choice == "1":
        search_mode = "fast"
    elif search_choice == "3":
        search_mode = "full"
    else:
        search_mode = "smart"
    
    print(f"\n🔍 Searching for domain: {domain}...")
    
    # اختيار دالة البحث المناسبة بناءً على نوع البحث
    if search_mode == "fast":
        print("🚀 Using fast search (main domains only)...")
        server, acct, server_name = find_server_by_domain_fast(domain, servers)
    elif search_mode == "full":
        print("🧠 Using full search (all domains + subdomains)...")
        server, acct, server_name = find_server_by_domain_full(domain, servers)
    else:  # smart mode
        print("🧠 Using smart search (main domains first, then subdomains if needed)...")
        server, acct, server_name = find_server_by_domain_smart(domain, servers)
    
    if not server:
        print("❌ Domain not found on any server!")
        return
        
    cpanel_user = acct["user"]
    print(f"\n✅ Domain found on Server {server_name}")
    print(f"📋 Domain: {domain}")
    print(f"👤 cPanel User: {cpanel_user}")
    
    # خيارات كلمات المرور
    print(f"\n🔑 Password Options:")
    print("1. Generate random passwords automatically")
    print("2. Enter passwords manually")
    
    password_choice = input("Choose password option (1-2): ").strip()
    auto_generate = password_choice == "1"
    
    if auto_generate:
        print("🎲 Random passwords will be generated automatically")
    else:
        print("✏️  You will enter passwords manually")
    
    # جمع معلومات الإيميلات
    print(f"\n📝 Enter email accounts to create:")
    if auto_generate:
        print("Format: username,quota(optional)")
        print("Example: info,500")
        print("Note: Domain and passwords will be added automatically")
    else:
        print("Format: username,password,quota(optional)")
        print("Example: info,mypassword,500")
        print("Note: Domain will be added automatically")
    print("Enter empty line to finish")
    
    emails_to_create = []
    while True:
        email_info = input("Email info: ").strip()
        if not email_info:
            break
            
        parts = email_info.split(",")
        if auto_generate:
            # توليد كلمة مرور تلقائياً
            if len(parts) < 1:
                print("❌ Invalid format. Use: username,quota(optional)")
                continue
                
            username = parts[0].strip()
            quota = int(parts[1].strip()) if len(parts) > 1 and parts[1].strip().isdigit() else 250
            password = generate_password(12)  # توليد كلمة مرور قوية
            
            print(f"   🎲 Generated password: {password}")
        else:
            # إدخال كلمة مرور يدوياً
            if len(parts) < 2:
                print("❌ Invalid format. Use: username,password,quota(optional)")
                continue
                
            username = parts[0].strip()
            password = parts[1].strip()
            quota = int(parts[2].strip()) if len(parts) > 2 and parts[2].strip().isdigit() else 250
        
        # إزالة @ من username إذا كان موجوداً
        if '@' in username:
            username = username.split('@')[0]
            print(f"   ⚠️  Removed @ from username: {username}")
        
        # إنشاء الإيميل الكامل
        email_address = f"{username}@{domain}"
        emails_to_create.append({
            "email": email_address,
            "password": password,
            "quota": quota
        })
        
        print(f"   ✓ Added: {email_address} (Quota: {quota}MB)")
    
    if not emails_to_create:
        print("❌ No emails to create")
        return
    
    print(f"\n📊 Summary: {len(emails_to_create)} emails to create")
    
    if confirm_action("Create all email accounts?"):
        successful = 0
        failed = 0
        
        print(f"\n🔄 Creating email accounts...")
        for email_info in emails_to_create:
            print(f"📧 Creating {email_info['email']}...")
            
            result = create_email_account(server, cpanel_user, email_info['email'], 
                                        email_info['password'], email_info['quota'])
            
            if result["success"]:
                print(f"✅ Created: {email_info['email']}")
                successful += 1
                logging.info(f"Email created: {email_info['email']} on {domain}")
            else:
                print(f"❌ Failed: {email_info['email']} - {result['error']}")
                failed += 1
            
            time.sleep(1)  # تأخير قصير
        
        print(f"\n📊 Bulk Email Creation Results:")
        print(f"✅ Successful: {successful}")
        print(f"❌ Failed: {failed}")
        print(f"📈 Success Rate: {(successful/(successful+failed))*100:.1f}%")
        
        # عرض تفاصيل الإيميلات المُنشأة
        if successful > 0:
            print(f"\n📋 Created Email Accounts:")
            print("=" * 60)
            for email_info in emails_to_create:
                print(f"📧 {email_info['email']}")
                print(f"   🔑 Password: {email_info['password']}")
                print(f"   💾 Quota: {email_info['quota']}MB")
                print(f"   💻 Webmail: https://webmail.{domain}")
                print("-" * 40)

def change_email_passwords(servers):
    """تغيير كلمات مرور الإيميلات"""
    print("\n🔑 Change Email Password(s)")
    print("=" * 50)
    
    # خيارات البحث
    print("🔍 Search Options:")
    print("1. 🌐 Search by domain")
    print("2. 📧 Search by email address")
    
    search_type = input("Choose search type (1-2, default 1): ").strip()
    
    if search_type == "2":
        # البحث بالإيميل مباشرة
        email_address = input("📧 Enter full email address: ").strip()
        if not email_address or "@" not in email_address:
            print("❌ Invalid email address format")
            return
            
        domain = email_address.split("@")[1]
        print(f"🌐 Extracted domain: {domain}")
        
        # البحث عن السيرفر
        print(f"\n🔍 Searching for email: {email_address}...")
        server, acct, server_name = find_server_by_email(email_address, servers)
        
        if not server:
            print("❌ Email not found on any server!")
            return
            
        cpanel_user = acct["user"]
        print(f"\n✅ Email found on Server {server_name}")
        print(f"📋 Domain: {domain}")
        print(f"👤 cPanel User: {cpanel_user}")
        
        # البحث عن الإيميل في قائمة الإيميلات
        print(f"📧 Loading email accounts...")
        emails = list_email_accounts(server, cpanel_user, domain)
        
        if not emails:
            print("❌ No email accounts found or error loading accounts")
            return
        
        # البحث عن الإيميل المحدد
        target_email = None
        for email in emails:
            if email.get("email", "").lower() == email_address.lower():
                target_email = email
                break
        
        if not target_email:
            print(f"❌ Email {email_address} not found in the account!")
            return
        
        # عرض معلومات الإيميل
        print(f"\n📋 Email Account Found:")
        quota_info = get_email_usage(server, cpanel_user, email_address)
        print(f"📧 Email: {email_address}")
        print(f"💾 Quota: {quota_info['quota']}")
        print(f"📊 Used: {quota_info['used']}")
        print(f"📈 Usage: {quota_info['usage_percent']}")
        
        # خيارات تغيير الباسورد
        print(f"\n🔑 Password Change Options:")
        print("1. Change password for this email")
        print("2. Generate random password")
        print("0. Back")
        
        password_choice = input("Choose option: ").strip()
        
        if password_choice == "0":
            return
        elif password_choice == "1":
            # تغيير الباسورد يدوياً
            new_password = get_secure_password()
            if not new_password:
                print("❌ Password cannot be empty!")
                return
                
            if len(new_password) < 8:
                print("❌ Password must be at least 8 characters!")
                return
                
            confirm_password = get_secure_password()
            if new_password != confirm_password:
                print("❌ Passwords do not match!")
                return
            
            # تغيير الباسورد
            if confirm_action(f"Change password for {email_address}?"):
                result = change_email_password(server, cpanel_user, email_address, new_password)
                if result["success"]:
                    print(f"✅ Password changed successfully for {email_address}")
                    print(f"🔑 New password: {new_password}")
                else:
                    print(f"❌ Failed to change password: {result['error']}")
                    
        elif password_choice == "2":
            # توليد باسورد عشوائي
            new_password = generate_strong_password(16)
            print(f"🎲 Generated password: {new_password}")
            
            if confirm_action(f"Change password for {email_address} to the generated password?"):
                result = change_email_password(server, cpanel_user, email_address, new_password)
                if result["success"]:
                    print(f"✅ Password changed successfully for {email_address}")
                    print(f"🔑 New password: {new_password}")
                else:
                    print(f"❌ Failed to change password: {result['error']}")
        
        return
        
    else:
        # البحث بالدومين (الطريقة القديمة)
        domain = input("🌐 Enter domain: ").strip()
        if not domain:
            print("❌ Domain cannot be empty")
            return
            
        print(f"\n🔍 Search Options:")
        print("1. Fast search (main domains only)")
        print("2. Smart search (main domains + subdomains)")
        print("3. Full search (all domains + subdomains + addon domains)")
        
        search_choice = input("Choose search type (1-3, default 2): ").strip()
        if not search_choice:
            search_choice = "2"
        
        print(f"\n🔍 Searching for domain: {domain}...")
        
        # اختيار نوع البحث
        if search_choice == "1":
            print("🚀 Using fast search (main domains only)...")
            server, acct, server_name = find_server_by_domain_fast(domain, servers)
        elif search_choice == "2":
            print("🧠 Using smart search (main domains + subdomains)...")
            server, acct, server_name = find_server_by_domain_smart(domain, servers)
        elif search_choice == "3":
            print("🔍 Using full search (all domains + subdomains + addon domains)...")
            server, acct, server_name = find_server_by_domain_full(domain, servers)
        else:
            print("❌ Invalid choice, using smart search...")
            server, acct, server_name = find_server_by_domain_smart(domain, servers)
        
        if not server:
            print("❌ Domain not found on any server!")
            return
            
        cpanel_user = acct["user"]
        print(f"\n✅ Domain found on Server {server_name}")
        
        # جلب قائمة الإيميلات الموجودة
        print(f"📧 Loading email accounts...")
        emails = list_email_accounts(server, cpanel_user, domain)
        
        if not emails:
            print("❌ No email accounts found or error loading accounts")
            return
        
        print(f"\n📋 Available Email Accounts ({len(emails)} found):")
        
        # إضافة خيار البحث
        print("\n🔍 Search Options:")
        print("1. Search for specific email")
        print("2. Show all emails")
        print("0. Back")
        
        search_choice = input("Choose option: ").strip()
        
        if search_choice == "0":
            return
        elif search_choice == "1":
            # البحث عن إيميل محدد
            search_email = input("🔍 Enter email address to search: ").strip().lower()
            if not search_email:
                print("❌ Email address cannot be empty")
                return
                
            # البحث في قائمة الإيميلات
            found_emails = []
            for email in emails:
                if search_email in email.get("email", "").lower():
                    found_emails.append(email)
            
            if not found_emails:
                print(f"❌ No emails found matching '{search_email}'")
                return
                
            print(f"\n✅ Found {len(found_emails)} matching email(s):")
            for i, email in enumerate(found_emails, 1):
                email_address = email.get("email", "Unknown")
                quota_info = get_email_usage(server, cpanel_user, email_address)
                print(f"{i}. {email_address} (Quota: {quota_info['quota']}, Used: {quota_info['used']}, Usage: {quota_info['usage_percent']})")
            
            # استخدام الإيميلات الموجودة بدلاً من كل الإيميلات
            emails_to_use = found_emails
            print(f"\n🔑 Password Change Options for {len(emails_to_use)} found email(s):")
            
        elif search_choice == "2":
            # عرض كل الإيميلات
            for i, email in enumerate(emails, 1):
                email_address = email.get("email", "Unknown")
                quota_info = get_email_usage(server, cpanel_user, email_address)
                print(f"{i}. {email_address} (Quota: {quota_info['quota']}, Used: {quota_info['used']}, Usage: {quota_info['usage_percent']})")
            
            emails_to_use = emails
            print(f"\n🔑 Password Change Options:")
        else:
            print("❌ Invalid option")
            return
    
    # عرض عدد الإيميلات المحددة
    if 'search_choice' in locals() and search_choice == "1":
        print(f"\n📊 Showing {len(emails_to_use)} of {len(emails)} total emails")
    else:
        print(f"\n📊 Showing all {len(emails_to_use)} emails")
    
    print("1. Change password for specific email")
    print("2. Change passwords for multiple emails")
    print("3. Generate random passwords for all emails")
    print("0. Back")
    
    choice = input("Choose option: ").strip()
    
    if choice == "1":
        email_index = input(f"Enter email number (1-{len(emails_to_use)}): ").strip()
        try:
            index = int(email_index) - 1
            if 0 <= index < len(emails_to_use):
                email_address = emails_to_use[index].get("email", "")
                domain = email_address.split('@')[1]
                
                # عرض خيارات الباسورد
                print(f"\n📧 Email Address: {email_address}")
                print(f"🌐 Domain: {domain}")
                print("\n🔑 Password Options:")
                print("1. Generate random password")
                print("2. Enter password manually")
                
                pass_choice = input("\nChoose option (1-2): ").strip()
                
                if pass_choice == "1":
                    new_password = generate_password(12)
                    print(f"🎲 Generated Password: {new_password}")
                elif pass_choice == "2":
                    new_password = input("Enter password: ")
                else:
                    print("❌ Invalid option")
                    return
                
                if new_password.strip():
                    if confirm_action(f"Change password for {email_address}?"):
                        result = change_email_password(server, cpanel_user, email_address, new_password)
                        
                        if result["success"]:
                            print("\n" + "=" * 60)
                            print("✅ PASSWORD CHANGED SUCCESSFULLY!")
                            print("=" * 60)
                            print(f"📧 Email Address: {email_address}")
                            print(f"🌐 Domain: {domain}")
                            print(f"🔑 New Password: {new_password}")
                            print(f"💻 Webmail URL: https://webmail.{domain}")
                            print("=" * 60)
                        else:
                            print(f"\n❌ Failed to change password: {result['error']}")
                else:
                    print("\n❌ Password cannot be empty")
            else:
                print("❌ Invalid email number")
        except ValueError:
            print("❌ Invalid input")
    
    elif choice == "2":
        print(f"Enter email numbers separated by commas (e.g., 1,3,5):")
        indices_input = input("Email numbers: ").strip()
        
        try:
            indices = [int(x.strip()) - 1 for x in indices_input.split(",")]
            valid_indices = [i for i in indices if 0 <= i < len(emails_to_use)]
            
            if valid_indices:
                print(f"\n📧 Selected emails:")
                for i in valid_indices:
                    print(f"   {emails_to_use[i].get('email', '')}")
                
                # عرض خيارات الباسورد
                print("\n🔑 Password Options:")
                print("1. Generate random password")
                print("2. Enter password manually")
                
                pass_choice = input("\nChoose option (1-2): ").strip()
                
                if pass_choice == "1":
                    new_password = generate_password(12)
                    print(f"🎲 Generated Password: {new_password}")
                elif pass_choice == "2":
                    new_password = input("Enter new password for all selected emails: ")
                else:
                    print("❌ Invalid option")
                    return
                
                if new_password.strip():
                    if confirm_action(f"Change passwords for {len(valid_indices)} emails?"):
                        successful = 0
                        failed = 0
                        
                        print("\n🔄 Changing passwords...")
                        for i in valid_indices:
                            email_address = emails_to_use[i].get("email", "")
                            email_domain = email_address.split('@')[1] if '@' in email_address else domain
                            
                            # توليد باسورد مختلف لكل إيميل
                            individual_password = generate_password(12)
                            
                            result = change_email_password(server, cpanel_user, email_address, individual_password)
                            
                            if result["success"]:
                                print("=" * 50)
                                print(f"📧 Email Address: {email_address}")
                                print(f"🌐 Domain: {email_domain}")
                                print(f"🔑 New Password: {individual_password}")
                                print(f"💻 Webmail URL: https://webmail.{email_domain}")
                                print("=" * 50)
                                successful += 1
                            else:
                                print(f"❌ {email_address}: Failed - {result['error']}")
                                failed += 1
            
            else:
                print("❌ No valid email numbers provided")
        except ValueError:
            print("❌ Invalid input format")
    
    elif choice == "3":
        if confirm_action(f"Generate random passwords for ALL {len(emails_to_use)} emails?"):
            successful = 0
            failed = 0
            passwords_data = []  # تخزين بيانات كلمات المرور للتصدير
            
            print(f"\n🔄 Changing passwords for all emails...")
            for email in emails_to_use:
                email_address = email.get("email", "")
                new_password = generate_password(12)
                
                result = change_email_password(server, cpanel_user, email_address, new_password)
                
                if result["success"]:
                    print("=" * 50)
                    print(f"📧 Email Address: {email_address}")
                    print(f"🌐 Domain: {domain}")
                    print(f"🔑 New Password: {new_password}")
                    print(f"💻 Webmail URL: https://webmail.{domain}")
                    print("=" * 50)
                    
                    # إضافة البيانات للتصدير
                    passwords_data.append({
                        "Email": email_address,
                        "Domain": domain,
                        "New Password": new_password,
                        "Webmail URL": f"https://webmail.{domain}",
                        "Status": "Success"
                    })
                    
                    successful += 1
                else:
                    print(f"❌ {email_address}: Failed - {result['error']}")
                    
                    # إضافة البيانات للتصدير (حتى لو فشل)
                    passwords_data.append({
                        "Email": email_address,
                        "Domain": domain,
                        "New Password": "FAILED",
                        "Webmail URL": f"https://webmail.{domain}",
                        "Status": f"Failed: {result['error']}"
                    })
                    
                    failed += 1
                
                time.sleep(1)
            
            print(f"\n📊 Bulk Password Change Results:")
            print(f"✅ Successful: {successful}")
            print(f"❌ Failed: {failed}")
            print(f"📈 Success Rate: {(successful/(successful+failed))*100:.1f}%")
            
            # خيار التصدير
            if successful > 0:
                print(f"\n💾 Export Options:")
                print("1. Export to Excel (.xlsx)")
                print("2. Export to CSV (.csv)")
                print("0. Skip export")
                
                export_choice = input("Choose export option (1-2, default 0): ").strip()
                
                if export_choice == "1":
                    # تصدير إلى Excel
                    filename = f"passwords_{domain}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                    export_passwords_to_excel(passwords_data, filename)
                elif export_choice == "2":
                    # تصدير إلى CSV
                    filename = f"passwords_{domain}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
                    export_passwords_to_csv(passwords_data, filename)

def delete_email_accounts(servers):
    """حذف حسابات الإيميل"""
    print("\n🗑️ Delete Email Account(s)")
    print("=" * 50)
    
    domain = input("🌐 Enter domain: ").strip()
    if not domain:
        print("❌ Domain cannot be empty")
        return
        
    # خيارات البحث
    print("\n🔍 Search Options:")
    print("1. 🚀 Fast search (main domains only)")
    print("2. 🧠 Smart search (main domains first, then subdomains if needed)")
    print("3. 🔍 Full search (all domains + subdomains)")
    
    search_choice = input("Choose search type (1-3, default 2): ").strip()
    
    if search_choice == "1":
        search_mode = "fast"
    elif search_choice == "3":
        search_mode = "full"
    else:
        search_mode = "smart"
    
    print(f"\n🔍 Searching for domain: {domain}...")
    
    # اختيار دالة البحث المناسبة بناءً على نوع البحث
    if search_mode == "fast":
        print("🚀 Using fast search (main domains only)...")
        server, acct, server_name = find_server_by_domain_fast(domain, servers)
    elif search_mode == "full":
        print("🧠 Using full search (all domains + subdomains)...")
        server, acct, server_name = find_server_by_domain_full(domain, servers)
    else:  # smart mode
        print("🧠 Using smart search (main domains first, then subdomains if needed)...")
        server, acct, server_name = find_server_by_domain_smart(domain, servers)
    
    if not server:
        print("❌ Domain not found on any server!")
        return
        
    cpanel_user = acct["user"]
    print(f"\n✅ Domain found on Server {server_name}")
    
    # جلب قائمة الإيميلات
    print(f"📧 Loading email accounts...")
    emails = list_email_accounts(server, cpanel_user, domain)
    
    if not emails:
        print("❌ No email accounts found")
        return
    
    print(f"\n📋 Email Accounts ({len(emails)} found):")
    
    # إضافة خيار البحث
    print("\n🔍 Search Options:")
    print("1. Search for specific email")
    print("2. Show all emails")
    print("0. Back")
    
    search_choice = input("Choose option: ").strip()
    
    if search_choice == "0":
        return
    elif search_choice == "1":
        # البحث عن إيميل محدد
        search_email = input("🔍 Enter email address to search: ").strip().lower()
        if not search_email:
            print("❌ Email address cannot be empty")
            return
            
        # البحث في قائمة الإيميلات
        found_emails = []
        for email in emails:
            if search_email in email.get("email", "").lower():
                found_emails.append(email)
        
        if not found_emails:
            print(f"❌ No emails found matching '{search_email}'")
            return
            
        print(f"\n✅ Found {len(found_emails)} matching email(s):")
        for i, email in enumerate(found_emails, 1):
            email_address = email.get("email", "Unknown")
            quota_info = get_email_usage(server, cpanel_user, email_address)
            print(f"{i}. {email_address} (Quota: {quota_info['quota']}, Used: {quota_info['used']}, Usage: {quota_info['usage_percent']})")
        
        # استخدام الإيميلات الموجودة بدلاً من كل الإيميلات
        emails_to_use = found_emails
        print(f"\n🗑️ Deletion Options for {len(emails_to_use)} found email(s):")
        
    elif search_choice == "2":
        # عرض كل الإيميلات
        for i, email in enumerate(emails, 1):
            email_address = email.get("email", "Unknown")
            quota_info = get_email_usage(server, cpanel_user, email_address)
            print(f"{i}. {email_address} (Quota: {quota_info['quota']}, Used: {quota_info['used']}, Usage: {quota_info['usage_percent']})")
        
        emails_to_use = emails
        print(f"\n🗑️ Deletion Options:")
    else:
        print("❌ Invalid option")
        return
    
    # عرض عدد الإيميلات المحددة
    if search_choice == "1":
        print(f"\n📊 Showing {len(emails_to_use)} of {len(emails)} total emails")
    else:
        print(f"\n📊 Showing all {len(emails_to_use)} emails")
    
    print("1. Delete specific email")
    print("2. Delete multiple emails")
    print("0. Back")
    
    choice = input("Choose option: ").strip()
    
    if choice == "1":
        email_index = input(f"Enter email number (1-{len(emails_to_use)}): ").strip()
        try:
            index = int(email_index) - 1
            if 0 <= index < len(emails_to_use):
                email_address = emails_to_use[index].get("email", "")
                
                print(f"⚠️ WARNING: This will permanently delete {email_address} and ALL its data!")
                if confirm_action(f"Delete {email_address}?"):
                    result = delete_email_account(server, cpanel_user, email_address)
                    
                    if result["success"]:
                        print(f"✅ Email {email_address} deleted successfully")
                        logging.info(f"Email deleted: {email_address} from {domain}")
                    else:
                        print(f"❌ Failed to delete email: {result['error']}")
            else:
                print("❌ Invalid email number")
        except ValueError:
            print("❌ Invalid input")
    
    elif choice == "2":
        print(f"Enter email numbers separated by commas (e.g., 1,3,5):")
        indices_input = input("Email numbers: ").strip()
        
        try:
            indices = [int(x.strip()) - 1 for x in indices_input.split(",")]
            valid_indices = [i for i in indices if 0 <= i < len(emails_to_use)]
            
            if valid_indices:
                print(f"\n📧 Emails to delete:")
                for i in valid_indices:
                    print(f"   {emails_to_use[i].get('email', '')}")
                
                print(f"\n⚠️ WARNING: This will permanently delete {len(valid_indices)} emails and ALL their data!")
                if confirm_action(f"Delete {len(valid_indices)} emails?"):
                    successful = 0
                    failed = 0
                    
                    for i in valid_indices:
                        email_address = emails_to_use[i].get("email", "")
                        
                        print(f"🗑️ Deleting {email_address}...")
                        result = delete_email_account(server, cpanel_user, email_address)
                        
                        if result["success"]:
                            print(f"✅ Deleted: {email_address}")
                            successful += 1
                            logging.info(f"Email deleted: {email_address} from {domain}")
                        else:
                            print(f"❌ Failed: {email_address} - {result['error']}")
                            failed += 1
            
            else:
                print("❌ No valid email numbers provided")
        except ValueError:
            print("❌ Invalid input format")

def list_and_export_emails(servers):
    """عرض وتصدير قائمة الإيميلات"""
    print("\n📋 List & Export Email Accounts")
    print("=" * 50)
    
    domain = input("🌐 Enter domain: ").strip()
    if not domain:
        print("❌ Domain cannot be empty")
        return
        
    # خيارات البحث
    print("\n🔍 Search Options:")
    print("1. 🚀 Fast search (main domains only)")
    print("2. 🧠 Smart search (main domains first, then subdomains if needed)")
    print("3. 🔍 Full search (all domains + subdomains)")
    
    search_choice = input("Choose search type (1-3, default 2): ").strip()
    
    if search_choice == "1":
        search_mode = "fast"
    elif search_choice == "3":
        search_mode = "full"
    else:
        search_mode = "smart"
    
    print(f"\n🔍 Searching for domain: {domain}...")
    
    # اختيار دالة البحث المناسبة بناءً على نوع البحث
    if search_mode == "fast":
        print("🚀 Using fast search (main domains only)...")
        server, acct, server_name = find_server_by_domain_fast(domain, servers)
    elif search_mode == "full":
        print("🧠 Using full search (all domains + subdomains)...")
        server, acct, server_name = find_server_by_domain_full(domain, servers)
    else:  # smart mode
        print("🧠 Using smart search (main domains first, then subdomains if needed)...")
        server, acct, server_name = find_server_by_domain_smart(domain, servers)
    
    if not server:
        print("❌ Domain not found on any server!")
        return
        
    cpanel_user = acct["user"]
    print(f"\n✅ Domain found on Server {server_name}")
    
    # جلب قائمة الإيميلات
    print(f"📧 Loading email accounts...")
    emails = list_email_accounts(server, cpanel_user, domain)
    
    if not emails:
        print("❌ No email accounts found")
        return
    
    print(f"\n📊 Email Accounts Report for {domain}")
    print("=" * 85)
    print(f"Server: {server_name} ({server['ip']})")
    print(f"cPanel User: {cpanel_user}")
    print(f"Total Emails: {len(emails)}")
    print("=" * 85)
    
    # إضافة خيارات متقدمة
    print("\n🔍 Options:")
    print("1. Search for specific email")
    print("2. Show all emails")
    print("3. Export directly without display")
    print("0. Back")
    
    search_choice = input("Choose option: ").strip()
    
    if search_choice == "0":
        return
    elif search_choice == "1":
        # البحث عن إيميل محدد
        search_email = input("🔍 Enter email address to search: ").strip().lower()
        if not search_email:
            print("❌ Email address cannot be empty")
            return
            
        # البحث في قائمة الإيميلات
        found_emails = []
        for email in emails:
            if search_email in email.get("email", "").lower():
                found_emails.append(email)
        
        if not found_emails:
            print(f"❌ No emails found matching '{search_email}'")
            return
            
        print(f"\n✅ Found {len(found_emails)} matching email(s):")
        emails_to_use = found_emails
        display_mode = "search"
        
    elif search_choice == "2":
        # عرض كل الإيميلات
        emails_to_use = emails
        display_mode = "show_all"
    elif search_choice == "3":
        # تصدير مباشر بدون عرض
        emails_to_use = emails
        display_mode = "export_only"
        print(f"\n📊 Preparing export for {len(emails_to_use)} emails...")
    else:
        print("❌ Invalid option")
        return
    
    # عرض عدد الإيميلات المحددة (فقط إذا لم يكن تصدير مباشر)
    if display_mode != "export_only":
        if search_choice == "1":
            print(f"\n📊 Showing {len(emails_to_use)} of {len(emails)} total emails")
        else:
            print(f"\n📊 Showing all {len(emails_to_use)} emails")
        
        # عرض تفصيلي للإيميلات
        print(f"{'#':<3} {'Email Address':<35} {'Quota':<12} {'Used':<12} {'Usage %':<10} {'Forward':<25}")
        print("-" * 110)
    
    # تبسيط - لا نحاول جلب Forward Rules
    all_forwards = {}
    
    # عرض الإيميلات فقط
    email_details = []
    
    for i, email in enumerate(emails_to_use, 1):
        email_address = email.get("email", "Unknown")
        
        # استخدام البيانات الأساسية
        used_bytes = email.get("diskused", 0)
        used_mb = round(used_bytes / (1024 * 1024), 2) if used_bytes > 0 else 0
        used_display = f"{used_mb}MB"
        usage_percent = "0.0%" if used_mb == 0 else f"{round((used_mb / 1024) * 100, 2)}%"
        quota_display = "Unlimited"
        
        # جلب معلومات Forward Rules من القائمة المحملة مسبقاً
        forward_info = "No Forward"
        if email_address in all_forwards:
            forward_to = all_forwards[email_address]
            forward_info = f"→ {forward_to}"
        
        # عرض الإيميل
        if display_mode != "export_only":
            print(f"{i:<3} {email_address:<35} {quota_display:<12} {used_display:<12} {usage_percent:<10} {forward_info:<25}")
        
        email_details.append({
            "email": email_address,
            "quota": quota_display,
            "used": used_display,
            "usage_percent": usage_percent,
            "forward": forward_info,
            "domain": domain,
            "server": server_name,
            "cpanel_user": cpanel_user
        })
        
        # جلب معلومات Forward Rules من القائمة المحملة مسبقاً
        forward_info = "No Forward"
        if email_address in all_forwards:
            forward_to = all_forwards[email_address]
            forward_info = f"→ {forward_to}"
        
        # عرض الإيميل فقط إذا لم يكن تصدير مباشر
        if display_mode != "export_only":
            print(f"{i:<3} {email_address:<35} {quota_display:<12} {used_display:<12} {usage_percent:<10} {forward_info:<25}")
        
        email_details.append({
            "email": email_address,
            "quota": quota_display,
            "used": used_display,
            "usage_percent": usage_percent,
            "forward": forward_info,
            "domain": domain,
            "server": server_name,
            "cpanel_user": cpanel_user
        })
        
        # تأخير أقصر للتصدير المباشر
        if display_mode == "export_only":
            time.sleep(0.1)  # تأخير سريع للتصدير
        else:
            time.sleep(0.3)  # تأخير أطول للحصول على بيانات دقيقة
    
    # إحصائيات إجمالية (فقط إذا لم يكن تصدير مباشر)
    if display_mode != "export_only":
        print("-" * 110)
        if total_quota > 0 and total_used > 0:
            overall_usage = (total_used / total_quota) * 100
            print(f"📊 Summary Statistics:")
            print(f"   Total Quota: {total_quota:.1f}MB")
            print(f"   Total Used: {total_used:.1f}MB")
            print(f"   Overall Usage: {overall_usage:.1f}%")
            print(f"   Available Space: {total_quota - total_used:.1f}MB")
        else:
            print(f"📊 Summary: {len(emails)} email accounts found")
    
    # خيارات التصدير
    if display_mode == "export_only":
        # للتصدير المباشر، استخدم خيار افتراضي
        print(f"\n📤 Exporting {len(emails_to_use)} emails...")
        export_choice = "3"  # تصدير بصيغتي Excel و CSV
    else:
        print(f"\n📤 Export Options:")
        print("1. Export to Excel")
        print("2. Export to CSV")
        print("3. Export both formats")
        print("0. Skip export")
        
        export_choice = input("Choose export option: ").strip()
    
    if export_choice in ["1", "3"]:
        headers = ["Email Address", "Quota (MB)", "Used (MB)", "Usage %", "Forward Rules", "Domain", 
                  "Server", "cPanel User", "Export Date"]
        data_rows = []
        
        for email in email_details:
            data_rows.append([
                email['email'],
                email['quota'],
                email['used'],
                email['usage_percent'],
                email['forward'],
                email['domain'],
                email['server'],
                email['cpanel_user'],
                datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ])
        
        export_to_excel(data_rows, headers, f"emails_{domain.replace('.', '_')}", f"Emails - {domain}")
    
    if export_choice in ["2", "3"]:
        headers = ["Email Address", "Quota (MB)", "Used (MB)", "Usage %", "Forward Rules", "Domain", 
                  "Server", "cPanel User", "Export Date"]
        data_rows = []
        
        for email in email_details:
            data_rows.append([
                email['email'],
                email['quota'],
                email['used'],
                email['usage_percent'],
                email['forward'],
                email['domain'],
                email['server'],
                email['cpanel_user'],
                datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ])
        
        export_to_csv(data_rows, headers, f"emails_{domain.replace('.', '_')}")
    
    # رسائل التصدير
    if export_choice != "0":
        if display_mode == "export_only":
            print(f"\n✅ Direct export completed successfully!")
            print(f"📁 Files saved in current directory")
            print(f"📊 Total emails exported: {len(email_details)}")
            print(f"⏱️  Export completed without displaying emails")
        else:
            print(f"\n✅ Export completed successfully!")
            print(f"📁 Files saved in current directory")
            print(f"📊 Total emails exported: {len(email_details)}")
    else:
        print(f"\nℹ️  Export skipped")

# === دوال مراقبة الإيميل ===
def get_failed_emails_report(server, days=7):
    """جلب تقرير الإيميلات الفاشلة من السيرفر"""
    try:
        # محاولة جلب إحصائيات Exim
        result = whm_api_call(server, "get_mailserver_stats")
        
        if "error" in result:
            # طريقة بديلة - تقدير بناءً على حالة الحسابات
            return create_basic_failed_report(server, days)
        
        # معالجة البيانات إذا كانت متاحة
        if "data" in result:
            stats = result["data"]
            
            # استخراج الإحصائيات
            failed_report = {
                "success": True,
                "total_failures": stats.get("failed_deliveries", 0),
                "bounces": stats.get("bounced", 0),
                "rejects": stats.get("rejected", 0),
                "failed_emails": [],
                "period_days": days
            }
            
            # إضافة تفاصيل الفشل إذا كانت متاحة
            if "failures" in stats:
                failed_report["failed_emails"] = stats["failures"][-50:]  # آخر 50 فشل
            
            return failed_report
        
        return {"success": False, "error": "No data available"}
        
    except Exception as e:
        logging.error(f"Error getting failed emails report: {str(e)}")
        return {"success": False, "error": str(e)}

def analyze_failed_emails_by_accounts(server, days=7):
    """تحليل الإيميلات الفاشلة حسب الحسابات المحددة"""
    try:
        print(f"\n🔍 Analyzing failed emails by accounts...")
        
        # جلب جميع الحسابات
        accounts = list_accounts(server)
        if not accounts:
            return {
                "success": False,
                "error": "No accounts found",
                "problematic_accounts": []
            }
        
        problematic_accounts = []
        total_failures = 0
        
        for account in accounts:
            account_failures = 0
            risk_factors = []
            
            # حساب نقاط الخطر
            risk_score = 0
            
            # 1. الحساب معلق
            if account.get('suspended', 0) == 1:
                risk_score += 5
                risk_factors.append("Account suspended")
                account_failures += 10  # تقدير 10 فشل للحساب المعلق
            
            # 2. استخدام قرص عالي
            try:
                disk_used = float(account.get('diskused', 0))
                if disk_used > 1000:  # أكثر من 1GB
                    risk_score += 3
                    risk_factors.append(f"High disk usage ({disk_used:.0f}MB)")
                    account_failures += 5
                elif disk_used > 500:
                    risk_score += 2
                    risk_factors.append(f"Moderate disk usage ({disk_used:.0f}MB)")
                    account_failures += 3
            except:
                pass
            
            # 3. تاريخ إنشاء حديث
            try:
                creation_date = datetime.fromtimestamp(int(account.get('unix_startdate', 0)))
                days_old = (datetime.now() - creation_date).days
                if days_old < 7:  # أقل من أسبوع
                    risk_score += 2
                    risk_factors.append(f"Recently created ({days_old} days ago)")
                    account_failures += 2
            except:
                pass
            
            # 4. أنماط الدومين المشبوهة
            domain = account.get('domain', '').lower()
            suspicious_patterns = ['temp', 'test', 'spam', 'bulk', 'mail', 'send', 'newsletter']
            if any(pattern in domain for pattern in suspicious_patterns):
                risk_score += 4
                risk_factors.append(f"Suspicious domain pattern: {domain}")
                account_failures += 8
            
            # 5. عدد حسابات الإيميل
            try:
                cpanel_user = account.get('user', '')
                emails = list_email_accounts(server, cpanel_user)
                email_count = len(emails) if emails else 0
                
                if email_count > 50:
                    risk_score += 5
                    risk_factors.append(f"High email count ({email_count})")
                    account_failures += 15
                elif email_count > 20:
                    risk_score += 3
                    risk_factors.append(f"Moderate email count ({email_count})")
                    account_failures += 8
                elif email_count > 10:
                    risk_score += 1
                    risk_factors.append(f"Elevated email count ({email_count})")
                    account_failures += 3
                
            except Exception as email_error:
                logging.warning(f"Error checking emails for {account.get('user', '')}: {str(email_error)}")
                email_count = 0
            
            # 6. إضافة جميع الحسابات مع نقاط الخطر
            # حساب تقديري للإيميلات الفاشلة بناءً على نقاط الخطر
            estimated_failures = int(risk_score * 1.2) if risk_score > 0 else 0
            account_failures = max(account_failures, estimated_failures)
            
            # إضافة الحساب إذا كان لديه نقاط خطر أو إيميلات فاشلة
            if risk_score > 0 or account_failures > 0:
                problematic_accounts.append({
                    'domain': account.get('domain', 'Unknown'),
                    'user': account.get('user', 'Unknown'),
                    'risk_score': risk_score,
                    'risk_factors': risk_factors,
                    'email_accounts': email_count,
                    'estimated_failures': account_failures,
                    'suspended': account.get('suspended', 0) == 1,
                    'disk_used': account.get('diskused', 0),
                    'creation_date': datetime.fromtimestamp(int(account.get('unix_startdate', 0))).strftime('%Y-%m-%d') if account.get('unix_startdate') else 'Unknown'
                })
                
                total_failures += account_failures
        
        # ترتيب الحسابات حسب نقاط الخطر
        problematic_accounts.sort(key=lambda x: x['risk_score'], reverse=True)
        
        return {
            "success": True,
            "total_failures": total_failures,
            "problematic_accounts": problematic_accounts,
            "total_accounts": len(accounts),
            "problematic_count": len(problematic_accounts),
            "period_days": days,
            "method": "account_analysis",
            "note": "Analysis based on account characteristics and risk factors"
        }
        
    except Exception as e:
        logging.error(f"Error analyzing failed emails by accounts: {str(e)}")
        return {
            "success": False,
            "error": str(e),
            "problematic_accounts": []
        }

def create_basic_failed_report(server, days):
    """إنشاء تقرير أساسي للإيميلات الفاشلة"""
    try:
        # محاولة الحصول على معلومات أساسية من الحسابات
        accounts = list_accounts(server)
        
        if not accounts:
            return {
                "success": False,
                "error": "No accounts found"
            }
        
        total_failures = 0
        suspended_accounts = 0
        high_disk_accounts = 0
        high_email_accounts = 0
        
        for account in accounts:
            account_failures = 0
            
            # 1. الحسابات المعلقة
            if account.get('suspended', 0) == 1:
                suspended_accounts += 1
                account_failures += 10
            
            # 2. استخدام قرص عالي
            try:
                disk_used = float(account.get('diskused', 0))
                if disk_used > 1000:
                    high_disk_accounts += 1
                    account_failures += 5
                elif disk_used > 500:
                    account_failures += 3
            except:
                pass
            
            # 3. عدد حسابات الإيميل
            try:
                cpanel_user = account.get('user', '')
                emails = list_email_accounts(server, cpanel_user)
                email_count = len(emails) if emails else 0
                
                if email_count > 50:
                    high_email_accounts += 1
                    account_failures += 15
                elif email_count > 20:
                    account_failures += 8
                elif email_count > 10:
                    account_failures += 3
                
            except Exception:
                pass
            
            total_failures += account_failures
        
        # إضافة تقدير إضافي بناءً على عدد الحسابات
        base_failures = len(accounts) * 0.5  # 0.5 فشل لكل حساب كأساس
        total_failures = max(total_failures, int(base_failures))
        
        return {
            "success": True,
            "total_failures": total_failures,
            "bounces": int(total_failures * 0.6),
            "rejects": int(total_failures * 0.4),
            "failed_emails": [],
            "period_days": days,
            "source": "enhanced_estimation",
            "note": f"Enhanced estimation: {suspended_accounts} suspended, {high_disk_accounts} high disk, {high_email_accounts} high email count",
            "details": {
                "suspended_accounts": suspended_accounts,
                "disks_accounts": high_disk_accounts,
                "high_email_accounts": high_email_accounts,
                "total_accounts": len(accounts)
            }
        }
        
    except Exception as e:
        return {
            "success": False,
            "error": f"Unable to generate report: {str(e)}"
        }

def check_blacklist_status(server, ip_address=None):
    """فحص حالة القوائم السوداء"""
    try:
        if not ip_address:
            # جلب IP السيرفر
            ip_address = server.get('ip', '')
        
        # قوائم سوداء شائعة للفحص
        blacklists = [
            "zen.spamhaus.org",
            "bl.spamcop.net", 
            "b.barracudacentral.org",
            "dnsbl.sorbs.net",
            "psbl.surriel.com",
            "ubl.unsubscore.com",
            "cbl.abuseat.org",
            "pbl.spamhaus.org"
        ]
        
        blacklist_results = {}
        
        for blacklist in blacklists:
            try:
                # عكس IP للفحص
                reversed_ip = '.'.join(ip_address.split('.')[::-1])
                query_host = f"{reversed_ip}.{blacklist}"
                
                # فحص DNS
                socket.gethostbyname(query_host)
                blacklist_results[blacklist] = "🔴 LISTED"
                
            except socket.gaierror:
                # عدم وجود في القائمة السوداء
                blacklist_results[blacklist] = "🟢 CLEAN"
            except Exception as e:
                blacklist_results[blacklist] = f"❓ ERROR: {str(e)}"
        
        return {
            "success": True,
            "ip_address": ip_address,
            "results": blacklist_results
        }
        
    except Exception as e:
        logging.error(f"Error checking blacklist status: {str(e)}")
        return {"error": str(e)}

def get_mail_queue_status(server):
    """جلب حالة طابور البريد باستخدام طرق بديلة"""
    try:
        # محاولة استخدام exim queue status
        result = whm_api_call(server, "exim_queue_status")
        
        if "error" not in result and "data" in result:
            queue_data = result["data"]
            return {
                "success": True,
                "queue_count": len(queue_data.get("messages", [])),
                "messages": queue_data.get("messages", []),
                "queue_size": queue_data.get("size", 0),
                "method": "exim_queue_status"
            }
        
        # محاولة استخدام mailq command عبر shell
        try:
            # استخدام shell command بدلاً من API
            import subprocess
            import re
            
            # محاولة الاتصال بالسيرفر عبر SSH (إذا كان متاحاً)
            # أو استخدام طريقة بديلة
            
            # إنشاء تقرير تقريبي بناءً على حالة الحسابات
            accounts = list_accounts(server)
            if accounts:
                # حساب تقريبي بناءً على الحسابات المعلقة
                suspended_accounts = sum(1 for acct in accounts if acct.get('suspended', 0) == 1)
                
                # تقدير حجم طابور البريد بناءً على حالة الحسابات
                estimated_queue = suspended_accounts * 3  # تقدير 3 رسائل لكل حساب معلق
                
                return {
                    "success": True,
                    "queue_count": estimated_queue,
                    "messages": [],
                    "queue_size": estimated_queue * 1024,  # تقدير 1KB لكل رسالة
                    "method": "estimated_from_accounts",
                    "note": "Queue size estimated based on account status - actual mail queue access unavailable"
                }
            
        except Exception as shell_error:
            logging.warning(f"Shell command failed: {str(shell_error)}")
        
        # محاولة استخدام cPanel API للحصول على معلومات البريد
        try:
            # البحث عن حساب واحد على الأقل لاستخدام cPanel API
            accounts = list_accounts(server)
            if accounts:
                test_user = accounts[0]["user"]
                
                # محاولة جلب معلومات البريد عبر cPanel API
                mail_info = cpanel_api_call(server, test_user, "Email", "list_pops")
                
                if "error" not in mail_info and "result" in mail_info:
                    # حساب تقريبي بناءً على عدد حسابات الإيميل
                    email_accounts = len(mail_info["result"].get("data", []))
                    
                    # تقدير حجم طابور البريد بناءً على عدد حسابات الإيميل
                    estimated_queue = email_accounts * 2  # تقدير 2 رسائل لكل حساب إيميل
                    
                    return {
                        "success": True,
                        "queue_count": estimated_queue,
                        "messages": [],
                        "queue_size": estimated_queue * 1024,
                        "method": "estimated_from_email_accounts",
                        "note": f"Queue size estimated from {email_accounts} email accounts - actual mail queue access unavailable"
                    }
        except Exception as cpanel_error:
            logging.warning(f"cPanel API attempt failed: {str(cpanel_error)}")
        
        # إذا فشلت جميع المحاولات
        return {
            "success": False,
            "error": "Mail queue information unavailable - WHM API doesn't support direct mail queue access",
            "queue_count": 0,
            "messages": [],
            "queue_size": 0,
            "method": "fallback"
        }
        
    except Exception as e:
        logging.error(f"Error getting mail queue: {str(e)}")
        return {
            "error": str(e),
            "queue_count": 0,
            "messages": [],
            "queue_size": 0,
            "method": "error"
        }

def get_mail_queue_status_advanced(server):
    """طريقة متقدمة لفحص طابور البريد باستخدام معلومات النظام"""
    try:
        # محاولة جلب معلومات النظام
        system_info = whm_api_call(server, "version")
        
        if "error" in system_info:
            return {
                "success": False,
                "error": "Cannot connect to server",
                "queue_count": 0,
                "method": "connection_failed"
            }
        
        # جلب معلومات الحسابات
        accounts = list_accounts(server)
        if not accounts:
            return {
                "success": True,
                "queue_count": 0,
                "messages": [],
                "queue_size": 0,
                "method": "no_accounts",
                "note": "No accounts found on server"
            }
        
        # حساب تقديري بناءً على عدة عوامل
        total_accounts = len(accounts)
        suspended_accounts = sum(1 for acct in accounts if acct.get('suspended', 0) == 1)
        active_accounts = total_accounts - suspended_accounts
        
        # حساب تقديري لطابور البريد
        base_queue = active_accounts * 0.5  # 0.5 رسالة لكل حساب نشط
        suspended_queue = suspended_accounts * 2  # 2 رسالة لكل حساب معلق
        
        estimated_queue = int(base_queue + suspended_queue)
        
        # تحديد حالة الطابور
        if estimated_queue == 0:
            queue_status = "Empty"
        elif estimated_queue < 10:
            queue_status = "Low"
        elif estimated_queue < 50:
            queue_status = "Normal"
        elif estimated_queue < 100:
            queue_status = "Moderate"
        else:
            queue_status = "High"
        
        return {
            "success": True,
            "queue_count": estimated_queue,
            "queue_status": queue_status,
            "messages": [],
            "queue_size": estimated_queue * 1024,
            "method": "advanced_estimation",
            "factors": {
                "total_accounts": total_accounts,
                "active_accounts": active_accounts,
                "suspended_accounts": suspended_accounts
            },
            "note": "Queue size estimated using advanced algorithm based on account status and activity"
        }
        
    except Exception as e:
        logging.error(f"Error in advanced mail queue check: {str(e)}")
        return {
            "success": False,
            "error": str(e),
            "queue_count": 0,
            "method": "advanced_error"
        }

def analyze_potential_spam_accounts(server):
    """تحليل الحسابات للبحث عن نشاط مشبوه"""
    try:
        accounts = list_accounts(server)
        
        if not accounts:
            return {"success": False, "error": "No accounts found"}
        
        suspicious_accounts = []
        
        for account in accounts:
            risk_score = 0
            risk_factors = []
            
            # عوامل الخطر المختلفة
            
            # 1. الحساب معلق
            if account.get('suspended', 0) == 1:
                risk_score += 3
                risk_factors.append("Account suspended")
            
            # 2. استخدام قرص عالي
            try:
                disk_used = float(account.get('diskused', 0))
                if disk_used > 1000:  # أكثر من 1GB
                    risk_score += 2
                    risk_factors.append("High disk usage")
            except:
                pass
            
            # 3. تاريخ إنشاء حديث
            try:
                creation_date = datetime.fromtimestamp(int(account.get('unix_startdate', 0)))
                days_old = (datetime.now() - creation_date).days
                if days_old < 7:  # أقل من أسبوع
                    risk_score += 1
                    risk_factors.append("Recently created")
            except:
                pass
            
            # 4. أنماط الدومين المشبوهة
            domain = account.get('domain', '').lower()
            suspicious_patterns = ['temp', 'test', 'spam', 'bulk', 'mail', 'send']
            if any(pattern in domain for pattern in suspicious_patterns):
                risk_score += 2
                risk_factors.append("Suspicious domain pattern")
            
            # 5. عدد كبير من حسابات الإيميل (تحليل تقريبي)
            try:
                # محاولة جلب عدد الإيميلات للحساب
                cpanel_user = account.get('user', '')
                emails = list_email_accounts(server, cpanel_user)
                email_count = len(emails) if emails else 0
                
                if email_count > 50:
                    risk_score += 3
                    risk_factors.append(f"High email count ({email_count})")
                elif email_count > 20:
                    risk_score += 1
                    risk_factors.append(f"Moderate email count ({email_count})")
                
            except:
                email_count = 0
            
            # إضافة للحسابات المشبوهة إذا كان النقاط أكثر من 3
            if risk_score >= 3:
                suspicious_accounts.append({
                    'domain': account.get('domain', 'Unknown'),
                    'user': account.get('user', 'Unknown'),
                    'risk_score': risk_score,
                    'risk_factors': risk_factors,
                    'email_accounts': email_count,
                    'suspended': account.get('suspended', 0) == 1,
                    'disk_used': account.get('diskused', 0)
                })
        
        # ترتيب حسب نقاط الخطر
        suspicious_accounts.sort(key=lambda x: x['risk_score'], reverse=True)
        
        return {
            "success": True,
            "total_accounts": len(accounts),
            "suspicious_count": len(suspicious_accounts),
            "suspicious_accounts": suspicious_accounts
        }
        
    except Exception as e:
        logging.error(f"Error analyzing potential spam accounts: {str(e)}")
        return {"success": False, "error": str(e)}

# === قوائم مراقبة الإيميل ===
def email_monitoring_dashboard(servers):
    """لوحة مراقبة الإيميل الشاملة"""
    print("\n📊 Email Monitoring Dashboard")
    print("=" * 50)
    
    while True:
        print(f"\n📈 Email Monitoring Options:")
        print("1. 🚫 Failed emails analysis") 
        print("2. ⚠️  Potential spam accounts analysis")
        print("3. 🔍 Blacklist status check")
        print("4. 📮 Mail queue status")
        print("5. 🎯 Quick health check (all servers)")
        print("6. 📋 Complete email audit report")
        print("0. 🚪 Back to main menu")
        
        dashboard_choice = input("\nChoose option: ").strip()
        
        if dashboard_choice == "1":
            failed_emails_analysis_menu(servers)
        elif dashboard_choice == "2":
            spam_analysis_menu(servers)
        elif dashboard_choice == "3":
            blacklist_check_menu(servers)
        elif dashboard_choice == "4":
            mail_queue_status_menu(servers)
        elif dashboard_choice == "5":
            quick_email_health_check_all_servers(servers)
        elif dashboard_choice == "6":
            complete_email_audit_menu(servers)
        elif dashboard_choice == "0":
            break
        else:
            print("❌ Invalid option")

def failed_emails_analysis_menu(servers):
    """قائمة تحليل الإيميلات الفاشلة"""
    print("\n🚫 Failed Emails Analysis")
    print("=" * 50)
    
    online_servers = get_online_servers(servers)
    if not online_servers:
        return
    
    print("1. 📊 Analyze specific server")
    print("2. 🔍 Quick check all servers")
    print("0. 🚪 Back")
    
    choice = input("\nChoose option: ").strip()
    
    if choice == "1":
        server_choice = input(f"\nChoose server ({'/'.join(online_servers.keys())}): ").strip()
        if server_choice in online_servers:
            failed_emails_report_for_server(online_servers[server_choice], server_choice)
        else:
            print("❌ Invalid server choice!")
    
    elif choice == "2":
        print("\n🔍 Quick Failed Emails Check - All Servers")
        print("=" * 60)
        
        for server_name, server in online_servers.items():
            print(f"\n🖥️  Checking {server_name}...")
            failed_report = get_failed_emails_report(server, 1)  # آخر يوم فقط
            
            if failed_report.get("success"):
                failures = failed_report['total_failures']
                if failures > 0:
                    print(f"   🔴 {failures} failed emails in last 24 hours")
                    if failures > 50:
                        print(f"   ⚠️  WARNING: High failure rate!")
                else:
                    print(f"   ✅ No failed emails")
            else:
                print(f"   ❌ Error checking failed emails")

def failed_emails_report_for_server(server, server_name):
    """تقرير الإيميلات الفاشلة لسيرفر محدد"""
    days = input("📅 Report period in days (default 7): ").strip() or "7"
    try:
        days = int(days)
    except ValueError:
        print("❌ Invalid number, using 7 days")
        days = 7
    
    print(f"\n🔍 Analyzing failed emails for last {days} days...")
    
    # استخدام التحليل الجديد للحسابات
    account_analysis = analyze_failed_emails_by_accounts(server, days)
    
    if account_analysis.get("success"):
        print(f"\n📊 Failed Emails Analysis - {server_name}")
        print("=" * 80)
        print(f"Period: Last {days} days")
        print(f"Total Accounts: {account_analysis['total_accounts']}")
        print(f"Problematic Accounts: {account_analysis['problematic_count']}")
        print(f"Estimated Total Failures: {account_analysis['total_failures']}")
        print(f"Daily Average: {account_analysis['total_failures']/days:.1f} failures/day")
        print(f"Method: {account_analysis['method']}")
        print(f"Note: {account_analysis['note']}")
        
        if account_analysis['problematic_accounts']:
            # فصل الحسابات الموقوفة والنشطة
            suspended_accounts = [acc for acc in account_analysis['problematic_accounts'] if acc['suspended']]
            active_risky_accounts = [acc for acc in account_analysis['problematic_accounts'] if not acc['suspended']]
            
            # جدول الحسابات الموقوفة
            if suspended_accounts:
                print(f"\n🔴 SUSPENDED ACCOUNTS (Already Stopped):")
                print("=" * 80)
                print(f"{'#':<3} {'Domain':<25} {'User':<15} {'Risk':<6} {'Emails':<8} {'Failures':<10} {'Disk (MB)':<12}")
                print("-" * 80)
                
                for i, account in enumerate(suspended_accounts, 1):
                    print(f"{i:<3} {account['domain']:<25} {account['user']:<15} "
                          f"{account['risk_score']:<6} {account['email_accounts']:<8} "
                          f"{account['estimated_failures']:<10} {account['disk_used']:<12}")
                
                print(f"📊 Total Suspended: {len(suspended_accounts)} accounts")
                total_suspended_failures = sum(acc['estimated_failures'] for acc in suspended_accounts)
                print(f"📈 Total Failures from Suspended: {total_suspended_failures}")
            
            # جدول الحسابات النشطة عالية الخطورة
            if active_risky_accounts:
                print(f"\n🟢 ACTIVE HIGH-RISK ACCOUNTS (Still Causing Problems):")
                print("=" * 80)
                print(f"{'#':<3} {'Domain':<25} {'User':<15} {'Risk':<6} {'Emails':<8} {'Failures':<10} {'Disk (MB)':<12}")
                print("-" * 80)
                
                for i, account in enumerate(active_risky_accounts, 1):
                    print(f"{i:<3} {account['domain']:<25} {account['user']:<15} "
                          f"{account['risk_score']:<6} {account['email_accounts']:<8} "
                          f"{account['estimated_failures']:<10} {account['disk_used']:<12}")
                
                print(f"📊 Total Active High-Risk: {len(active_risky_accounts)} accounts")
                total_active_failures = sum(acc['estimated_failures'] for acc in active_risky_accounts)
                print(f"📈 Total Failures from Active: {total_active_failures}")
                
                # تحذير للحسابات النشطة
                print(f"\n⚠️  WARNING: {len(active_risky_accounts)} active accounts are still causing problems!")
                print("💡 Recommendation: Consider suspending these accounts immediately.")
            
            # ملخص عام
            print(f"\n📊 OVERALL SUMMARY:")
            print("=" * 80)
            print(f"🔴 Suspended Accounts: {len(suspended_accounts)} (Failures: {total_suspended_failures})")
            print(f"🟢 Active High-Risk: {len(active_risky_accounts)} (Failures: {total_active_failures})")
            print(f"📈 Total Estimated Failures: {account_analysis['total_failures']}")
            
            # عرض تفاصيل أعلى 5 حسابات خطورة
            print(f"\n🔍 TOP 5 HIGH-RISK ACCOUNTS DETAILS:")
            print("=" * 80)
            
            for i, account in enumerate(account_analysis['problematic_accounts'][:5], 1):
                status_icon = "🔴" if account['suspended'] else "🟢"
                status_text = "SUSPENDED" if account['suspended'] else "ACTIVE"
                
                print(f"\n{i}. {status_icon} {account['domain']} ({status_text})")
                print(f"   👤 User: {account['user']}")
                print(f"   🚨 Risk Score: {account['risk_score']}/25")
                print(f"   📧 Email Accounts: {account['email_accounts']}")
                print(f"   📊 Estimated Failures: {account['estimated_failures']}")
                print(f"   💾 Disk Used: {account['disk_used']}MB")
                print(f"   📅 Created: {account['creation_date']}")
                print(f"   ⚠️  Risk Factors:")
                for factor in account['risk_factors']:
                    print(f"      • {factor}")
                print("-" * 40)
        
        # خيارات إضافية
        print(f"\n🔧 ACTION OPTIONS:")
        print("1. 📊 Export detailed report")
        print("2. 🚫 Suspend active high-risk accounts")
        print("3. 🔑 Change passwords for high-risk accounts")
        print("4. 📋 View all accounts on this server")
        print("5. 🔍 Focus on active problematic accounts")
        print("6. 🔍 Analyze specific domain in detail")
        print("7. 📮 Check mail queue for specific domain")
        print("8. 🔍 Compare analysis methods")
        print("0. 🚪 Back to main menu")
        
        action_choice = input("\nChoose action: ").strip()
        
        if action_choice == "1":
            # تصدير التقرير المفصل
            headers = ["Domain", "User", "Risk Score", "Email Accounts", "Estimated Failures", 
                      "Suspended", "Disk Used", "Creation Date", "Risk Factors", "Report Date"]
            data_rows = []
            
            for account in account_analysis['problematic_accounts']:
                data_rows.append([
                    account['domain'],
                    account['user'],
                    account['risk_score'],
                    account['email_accounts'],
                    account['estimated_failures'],
                    "Yes" if account['suspended'] else "No",
                    account['disk_used'],
                    account['creation_date'],
                    "; ".join(account['risk_factors']),
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ])
            
            export_to_excel(data_rows, headers, f"problematic_accounts_{server_name}", "Problematic Accounts Report")
            export_to_csv(data_rows, headers, f"problematic_accounts_{server_name}")
            
        elif action_choice == "2":
            # تعليق الحسابات المشبوهة
            if account_analysis['problematic_accounts']:
                print(f"\n🚫 SUSPEND PROBLEMATIC ACCOUNTS")
                print("=" * 50)
                print("Choose accounts to suspend (comma-separated numbers):")
                
                for i, account in enumerate(account_analysis['problematic_accounts'][:10], 1):
                    print(f"{i}. {account['domain']} (Risk: {account['risk_score']})")
                
                suspend_choice = input("\nEnter account numbers (e.g., 1,3,5): ").strip()
                if suspend_choice:
                    try:
                        indices = [int(x.strip()) - 1 for x in suspend_choice.split(",")]
                        valid_indices = [i for i in indices if 0 <= i < len(account_analysis['problematic_accounts'])]
                        
                        if valid_indices:
                            print(f"\n⚠️  WARNING: This will suspend {len(valid_indices)} accounts!")
                            if confirm_action("Continue with suspension?"):
                                for i in valid_indices:
                                    account = account_analysis['problematic_accounts'][i]
                                    domain = account['domain']
                                    print(f"🔄 Suspending {domain}...")
                                    
                                    # البحث عن الحساب في السيرفر
                                    accounts = list_accounts(server)
                                    acct = None
                                    for account in accounts:
                                        if account.get('domain', '').lower() == domain.lower():
                                            acct = account
                                            break
                                    
                                    if acct:
                                        params = {"user": acct["user"], "reason": "High email failure risk"}
                                        result = whm_api_call(server, "suspendacct", params)
                                        if "error" not in result:
                                            print(f"✅ {domain} suspended successfully")
                                        else:
                                            print(f"❌ Failed to suspend {domain}: {result['error']}")
                                    else:
                                        print(f"❌ Account {domain} not found")
                        else:
                            print("❌ No valid account numbers provided")
                    except ValueError:
                        print("❌ Invalid input format")
        
        elif action_choice == "3":
            # تغيير كلمات مرور الحسابات عالية الخطورة
            if account_analysis['problematic_accounts']:
                print(f"\n🔑 CHANGE PASSWORDS FOR HIGH-RISK ACCOUNTS")
                print("=" * 50)
                print("Choose accounts to change passwords (comma-separated numbers):")
                
                for i, account in enumerate(account_analysis['problematic_accounts'][:10], 1):
                    print(f"{i}. {account['domain']} (Risk: {account['risk_score']})")
                
                password_choice = input("\nEnter account numbers (e.g., 1,3,5): ").strip()
                if password_choice:
                    try:
                        indices = [int(x.strip()) - 1 for x in password_choice.split(",")]
                        valid_indices = [i for i in indices if 0 <= i < len(account_analysis['problematic_accounts'])]
                        
                        if valid_indices:
                            print(f"\n🔑 Changing passwords for {len(valid_indices)} accounts...")
                            for i in valid_indices:
                                account = account_analysis['problematic_accounts'][i]
                                domain = account['domain']
                                print(f"🔄 Changing password for {domain}...")
                                
                                # البحث عن الحساب في السيرفر
                                accounts = list_accounts(server)
                                acct = None
                                for account in accounts:
                                    if account.get('domain', '').lower() == domain.lower():
                                        acct = account
                                        break
                                
                                if acct:
                                    new_password = generate_strong_password(16)
                                    params = {"user": acct["user"], "password": new_password}
                                    result = whm_api_call(server, "passwd", params)
                                    if "error" not in result:
                                        print("=" * 50)
                                        print(f"🌐 Domain: {domain}")
                                        print(f"👤 cPanel User: {acct['user']}")
                                        print(f"🔑 New Password: {new_password}")
                                        print(f"💻 cPanel URL: https://{domain}:2083")
                                        print(f"📧 Webmail URL: https://webmail.{domain}")
                                        print("=" * 50)
                                    else:
                                        print(f"❌ Failed to change password for {domain}: {result['error']}")
                                else:
                                    print(f"❌ Account {domain} not found")
                        else:
                            print("❌ No valid account numbers provided")
                    except ValueError:
                        print("❌ Invalid input format")
        
        elif action_choice == "4":
            # عرض جميع الحسابات
            print(f"\n📋 ALL ACCOUNTS ON SERVER {server_name}")
            print("=" * 60)
            accounts = list_accounts(server)
            if accounts:
                print(f"{'#':<3} {'Domain':<25} {'User':<15} {'Status':<12} {'Disk (MB)':<12}")
                print("-" * 60)
                for i, acct in enumerate(accounts[:20], 1):  # أول 20 حساب
                    status = "🔴 Suspended" if acct.get('suspended', 0) == 1 else "🟢 Active"
                    print(f"{i:<3} {acct['domain']:<25} {acct['user']:<15} {status:<12} {acct.get('diskused', 'N/A'):<12}")
                
                if len(accounts) > 20:
                    print(f"... and {len(accounts) - 20} more accounts")
        
        elif action_choice == "5":
            # التركيز على الحسابات النشطة المشبوهة
            if active_risky_accounts:
                print(f"\n🔍 FOCUS ON ACTIVE PROBLEMATIC ACCOUNTS")
                print("=" * 60)
                print(f"These {len(active_risky_accounts)} accounts are still active and causing problems:")
                print()
                
                for i, account in enumerate(active_risky_accounts, 1):
                    print(f"{i}. 🌐 {account['domain']}")
                    print(f"   👤 User: {account['user']}")
                    print(f"   🚨 Risk Score: {account['risk_score']}/25")
                    print(f"   📧 Email Accounts: {account['email_accounts']}")
                    print(f"   📊 Estimated Failures: {account['estimated_failures']}")
                    print(f"   💾 Disk Used: {account['disk_used']}MB")
                    print(f"   ⚠️  Risk Factors:")
                    for factor in account['risk_factors']:
                        print(f"      • {factor}")
                    print("-" * 40)
                
                print(f"\n🚨 IMMEDIATE ACTION REQUIRED:")
                print("These accounts are still sending emails and causing failures!")
                
                if confirm_action("Suspend all active high-risk accounts now?"):
                    print(f"\n🔄 Suspending {len(active_risky_accounts)} accounts...")
                    successful_suspensions = 0
                    
                    for account in active_risky_accounts:
                        domain = account['domain']
                        print(f"🔄 Suspending {domain}...")
                        
                        # البحث عن الحساب في السيرفر
                        accounts = list_accounts(server)
                        acct = None
                        for account in accounts:
                            if account.get('domain', '').lower() == domain.lower():
                                acct = account
                                break
                        
                        if acct:
                            params = {"user": acct["user"], "reason": "Active high-risk account - immediate suspension required"}
                            result = whm_api_call(server, "suspendacct", params)
                            if "error" not in result:
                                print(f"✅ {domain} suspended successfully")
                                successful_suspensions += 1
                            else:
                                print(f"❌ Failed to suspend {domain}: {result['error']}")
                        else:
                            print(f"❌ Account {domain} not found")
                    
                    print(f"\n📊 Suspension Results:")
                    print(f"✅ Successfully suspended: {successful_suspensions}")
                    print(f"❌ Failed: {len(active_risky_accounts) - successful_suspensions}")
                    
                    if successful_suspensions > 0:
                        print(f"\n🎉 {successful_suspensions} high-risk accounts have been stopped!")
                        print("This should significantly reduce email failures.")
            else:
                print(f"\n✅ No active high-risk accounts found!")
                print("All problematic accounts are already suspended.")
        
        elif action_choice == "6":
            # فحص دومين محدد بالتفصيل
            print(f"\n🔍 DETAILED DOMAIN ANALYSIS")
            print("=" * 60)
            print("Enter a specific domain to analyze in detail:")
            print("This will show:")
            print("• Email accounts and their status")
            print("• Estimated email failures")
            print("• Disk usage details")
            print("• Risk factors")
            print("• Mail queue information")
            
            target_domain = input("\n🌐 Enter domain to analyze: ").strip()
            if target_domain:
                detailed_domain_analysis(server, target_domain, server_name)
            else:
                print("❌ No domain entered")
        
        elif action_choice == "7":
            # فحص طابور البريد لدومين محدد
            print(f"\n📮 MAIL QUEUE ANALYSIS FOR SPECIFIC DOMAIN")
            print("=" * 60)
            print("This will show:")
            print("• Mail queue status for the domain")
            print("• Email delivery issues")
            print("• Bounce and rejection patterns")
            print("• Recommendations")
            
            target_domain = input("\n🌐 Enter domain to check mail queue: ").strip()
            if target_domain:
                domain_mail_queue_analysis(server, target_domain, server_name)
            else:
                print("❌ No domain entered")
        
        elif action_choice == "8":
            # مقارنة طرق التحليل
            print(f"\n🔍 COMPARING ANALYSIS METHODS")
            print("=" * 60)
            print("This will show the difference between:")
            print("• Account-based analysis (detailed)")
            print("• Quick estimation method (overview)")
            print("• Why results might differ")
            
            compare_analysis_methods(server, server_name)
    
    else:
        print(f"❌ Error analyzing accounts: {account_analysis.get('error')}")
        print(f"\n💡 Trying alternative method...")
        
        # محاولة الطريقة القديمة كبديل
        failed_report = get_failed_emails_report(server, days)
        if failed_report.get("success"):
            print(f"\n📊 Basic Failed Emails Report - {server_name}")
            print("=" * 60)
            print(f"Period: Last {days} days")
            print(f"Total Failures: {failed_report['total_failures']}")
            print(f"Bounced Emails: {failed_report['bounces']}")
            print(f"Rejected Emails: {failed_report['rejects']}")
            print(f"Daily Average: {failed_report['total_failures']/days:.1f} failures/day")
            
            if failed_report.get('source') == 'estimation':
                print(f"⚠️  Note: {failed_report.get('note', '')}")
        else:
            print(f"❌ All analysis methods failed")

def spam_analysis_menu(servers):
    """قائمة تحليل السبام"""
    print("\n⚠️  Spam Analysis")
    print("=" * 50)
    
    online_servers = get_online_servers(servers)
    if not online_servers:
        return
    
    print("1. 🔍 Scan specific server")
    print("2. 🚨 Quick scan all servers")
    print("0. 🚪 Back")
    
    choice = input("\nChoose option: ").strip()
    
    if choice == "1":
        server_choice = input(f"\nChoose server ({'/'.join(online_servers.keys())}): ").strip()
        if server_choice in online_servers:
            spam_analysis_for_server(online_servers[server_choice], server_choice)
        else:
            print("❌ Invalid server choice!")
    
    elif choice == "2":
        quick_spam_scan_all_servers(online_servers)

def spam_analysis_for_server(server, server_name):
    """تحليل السبام لسيرفر محدد"""
    print(f"\n⚠️  Spam Analysis - {server_name}")
    print("=" * 50)
    
    print(f"🔍 Analyzing accounts for potential spam activity...")
    analysis = analyze_potential_spam_accounts(server)
    
    if analysis.get("success"):
        print(f"\n📊 Spam Risk Analysis Results")
        print("=" * 50)
        print(f"Total Accounts: {analysis['total_accounts']}")
        print(f"Suspicious Accounts: {analysis['suspicious_count']}")
        print(f"Risk Percentage: {(analysis['suspicious_count']/analysis['total_accounts'])*100:.1f}%")
        
        if analysis['suspicious_count'] > 0:
            # تصنيف المخاطر
            high_risk = len([acc for acc in analysis['suspicious_accounts'] if acc['risk_score'] >= 8])
            medium_risk = len([acc for acc in analysis['suspicious_accounts'] if 5 <= acc['risk_score'] < 8])
            
            print(f"\n🎯 Risk Classification:")
            print(f"   🔴 High Risk (8+): {high_risk} accounts")
            print(f"   🟡 Medium Risk (5-7): {medium_risk} accounts")
            
            # عرض أعلى المخاطر
            print(f"\n🚨 Top Risk Accounts:")
            print("-" * 80)
            print(f"{'Domain':<25} {'User':<15} {'Risk':<6} {'Emails':<8} {'Factors'}")
            print("-" * 80)
            
            for account in analysis['suspicious_accounts'][:10]:
                factors_str = ', '.join(account['risk_factors'][:2])  # أول عاملين فقط
                if len(account['risk_factors']) > 2:
                    factors_str += f" +{len(account['risk_factors'])-2}"
                
                print(f"{account['domain']:<25} {account['user']:<15} "
                      f"{account['risk_score']:<6} {account['email_accounts']:<8} {factors_str}")
        
        if confirm_action("\nExport spam analysis report?"):
            headers = ["Server", "Domain", "User", "Risk Score", "Email Accounts", "Suspended", "Disk Used", "Risk Factors", "Report Date"]
            data_rows = []
            
            for account in analysis['suspicious_accounts']:
                data_rows.append([
                    server_name,
                    account['domain'],
                    account['user'],
                    account['risk_score'],
                    account['email_accounts'],
                    "Yes" if account['suspended'] else "No",
                    account['disk_used'],
                    "; ".join(account['risk_factors']),
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ])
            
            export_to_excel(data_rows, headers, f"spam_analysis_{server_name}", "Spam Analysis")
            export_to_csv(data_rows, headers, f"spam_analysis_{server_name}")
    
    else:
        print(f"❌ Error analyzing accounts: {analysis.get('error')}")

def quick_spam_scan_all_servers(online_servers):
    """فحص سريع للسبام لجميع السيرفرات"""
    print("\n⚠️  Quick Spam Scan - All Servers")
    print("=" * 50)
    
    total_suspicious = 0
    
    for server_name, server in online_servers.items():
        print(f"\n🖥️  {server_name}:", end=" ")
        analysis = analyze_potential_spam_accounts(server)
        
        if analysis.get("success"):
            suspicious = analysis['suspicious_count']
            total_accounts = analysis['total_accounts']
            
            if suspicious == 0:
                print("✅ Clean")
            else:
                risk_percent = (suspicious / total_accounts) * 100
                if suspicious < 3:
                    print(f"🟡 {suspicious} suspicious ({risk_percent:.1f}%)")
                elif suspicious < 5:
                    print(f"🟠 {suspicious} suspicious ({risk_percent:.1f}%)")
                else:
                    print(f"🔴 {suspicious} suspicious ({risk_percent:.1f}%)")
                
                total_suspicious += suspicious
                
                # عرض أعلى مخاطر
                if analysis['suspicious_accounts']:
                    top_risk = analysis['suspicious_accounts'][0]
                    print(f"      ⚠️  Top risk: {top_risk['domain']} (score: {top_risk['risk_score']})")
        else:
            print("❌ Error")
    
    print(f"\n📊 SCAN SUMMARY:")
    print(f"Servers Scanned: {len(online_servers)}")
    print(f"Total Suspicious Accounts: {total_suspicious}")
    
    if total_suspicious > 0:
        print(f"\n💡 Recommendations:")
        print("1. Run detailed analysis on high-risk servers")
        print("2. Monitor suspicious accounts closely")
        print("3. Consider implementing email rate limits")

def blacklist_check_menu(servers):
    """قائمة فحص القوائم السوداء"""
    print("\n🔍 Blacklist Status Check")
    print("=" * 50)
    
    print("1. 🖥️  Check specific server")
    print("2. 🌐 Check all servers")
    print("3. 📝 Check custom IP address")
    print("0. 🚪 Back")
    
    choice = input("\nChoose option: ").strip()
    
    if choice == "1":
        online_servers = get_online_servers(servers)
        if online_servers:
            server_choice = input(f"\nChoose server ({'/'.join(online_servers.keys())}): ").strip()
            if server_choice in online_servers:
                blacklist_check_for_server(online_servers[server_choice], server_choice)
            else:
                print("❌ Invalid server choice!")
    
    elif choice == "2":
        blacklist_check_all_servers(servers)
    
    elif choice == "3":
        ip_to_check = input("Enter IP address to check: ").strip()
        if ip_to_check:
            blacklist_check_custom_ip(ip_to_check)
        else:
            print("❌ IP address cannot be empty!")

def blacklist_check_for_server(server, server_name):
    """فحص القوائم السوداء لسيرفر محدد"""
    print(f"\n🔍 Blacklist Check - {server_name}")
    print("=" * 50)
    
    print(f"📡 Checking IP {server['ip']} against blacklists...")
    
    blacklist_results = check_blacklist_status(server)
    
    if blacklist_results.get("success"):
        results = blacklist_results["results"]
        listed_count = len([r for r in results.values() if "🔴 LISTED" in r])
        clean_count = len([r for r in results.values() if "🟢 CLEAN" in r])
        
        print(f"\n📊 Blacklist Results for {server['ip']}:")
        print("=" * 60)
        
        for blacklist, status in results.items():
            print(f"{blacklist:<25} {status}")
        
        print("-" * 60)
        print(f"Summary: {clean_count} Clean, {listed_count} Listed")
        
        if listed_count > 0:
            print(f"\n🚨 WARNING: Server is listed on {listed_count} blacklist(s)!")
            print("Immediate action required:")
            print("1. Stop all email sending")
            print("2. Investigate spam sources")
            print("3. Clean compromised accounts")
            print("4. Submit delisting requests")
        else:
            print(f"\n✅ Great! Server is clean on all blacklists")
    
    else:
        print(f"❌ Error checking blacklists: {blacklist_results.get('error')}")

def blacklist_check_all_servers(servers):
    """فحص القوائم السوداء لجميع السيرفرات"""
    print("\n🔍 Blacklist Check - All Servers")
    print("=" * 50)
    
    online_servers = get_online_servers(servers)
    if not online_servers:
        return
    
    total_listings = 0
    servers_blacklisted = 0
    
    for server_name, server in online_servers.items():
        print(f"\n🖥️  {server_name} ({server['ip']}):", end=" ")
        
        blacklist_results = check_blacklist_status(server)
        
        if blacklist_results.get("success"):
            results = blacklist_results["results"]
            listed_count = len([r for r in results.values() if "🔴 LISTED" in r])
            clean_count = len([r for r in results.values() if "🟢 CLEAN" in r])
            
            if listed_count == 0:
                print(f"✅ Clean on all {clean_count} blacklists")
            else:
                print(f"🔴 LISTED on {listed_count} blacklist(s)")
                total_listings += listed_count
                servers_blacklisted += 1
                
                # عرض القوائم المدرج فيها
                listed_bls = [bl for bl, status in results.items() if "🔴 LISTED" in status]
                print(f"      📋 Listed on: {', '.join(listed_bls[:3])}")
                if len(listed_bls) > 3:
                    print(f"         + {len(listed_bls)-3} more")
        else:
            print("❌ Error checking")
    
    print(f"\n📊 BLACKLIST SUMMARY:")
    print(f"Servers Checked: {len(online_servers)}")
    print(f"Servers Blacklisted: {servers_blacklisted}")
    print(f"Total Listings: {total_listings}")
    
    if servers_blacklisted > 0:
        print(f"\n🚨 CRITICAL ALERT!")
        print("Immediate action required for blacklisted servers:")
        print("1. Stop all email sending")
        print("2. Investigate spam sources") 
        print("3. Clean compromised accounts")
        print("4. Submit delisting requests")

def blacklist_check_custom_ip(ip_address):
    """فحص IP مخصص ضد القوائم السوداء"""
    print(f"\n🔍 Custom IP Blacklist Check")
    print("=" * 50)
    
    print(f"📡 Checking IP {ip_address} against blacklists...")
    
    # إنشاء كائن سيرفر مؤقت للفحص
    temp_server = {"ip": ip_address}
    blacklist_results = check_blacklist_status(temp_server, ip_address)
    
    if blacklist_results.get("success"):
        results = blacklist_results["results"]
        listed_count = len([r for r in results.values() if "🔴 LISTED" in r])
        clean_count = len([r for r in results.values() if "🟢 CLEAN" in r])
        
        print(f"\n📊 Blacklist Results for {ip_address}:")
        print("=" * 60)
        
        for blacklist, status in results.items():
            print(f"{blacklist:<25} {status}")
        
        print("-" * 60)
        print(f"Summary: {clean_count} Clean, {listed_count} Listed")
        
        if listed_count > 0:
            print(f"\n🚨 WARNING: IP is listed on {listed_count} blacklist(s)!")
        else:
            print(f"\n✅ Great! IP is clean on all blacklists")
    
    else:
        print(f"❌ Error checking blacklists: {blacklist_results.get('error')}")

def mail_queue_status_menu(servers):
    """قائمة حالة طابور البريد"""
    print("\n📮 Mail Queue Status")
    print("=" * 50)
    
    online_servers = get_online_servers(servers)
    if not online_servers:
        return
    
    print("1. 📊 Check specific server queue")
    print("2. 🌐 Check all servers queues")
    print("0. 🚪 Back")
    
    choice = input("\nChoose option: ").strip()
    
    if choice == "1":
        server_choice = input(f"\nChoose server ({'/'.join(online_servers.keys())}): ").strip()
        if server_choice in online_servers:
            mail_queue_detailed_check(online_servers[server_choice], server_choice)
        else:
            print("❌ Invalid server choice!")
    
    elif choice == "2":
        mail_queue_check_all_servers(online_servers)

def mail_queue_detailed_check(server, server_name):
    """فحص مفصل لطابور البريد"""
    print(f"\n📮 Detailed Mail Queue Check - {server_name}")
    print("=" * 50)
    
    # استخدام الطريقة المتقدمة أولاً
    queue_status = get_mail_queue_status_advanced(server)
    
    if not queue_status.get("success"):
        # إذا فشلت الطريقة المتقدمة، استخدم الطريقة العادية
        queue_status = get_mail_queue_status(server)
    
    if queue_status.get("success"):
        queue_count = queue_status['queue_count']
        messages = queue_status.get('messages', [])
        method = queue_status.get('method', 'unknown')
        note = queue_status.get('note', '')
        
        print(f"📊 Queue Overview:")
        print(f"   Total Messages: {queue_count}")
        print(f"   Method Used: {method}")
        
        if note:
            print(f"   Note: {note}")
        
        print(f"   Queue Status: ", end="")
        
        if queue_count == 0:
            print("✅ Empty (Optimal)")
        elif queue_count < 20:
            print("🟢 Low (Normal)")
        elif queue_count < 100:
            print("🟡 Moderate (Monitor)")
        elif queue_count < 500:
            print("🟠 High (Attention needed)")
        else:
            print("🔴 Critical (Immediate action required)")
        
        # عرض العوامل المستخدمة في الحساب إذا كانت متاحة
        if 'factors' in queue_status:
            factors = queue_status['factors']
            print(f"\n📈 Calculation Factors:")
            print(f"   Total Accounts: {factors.get('total_accounts', 'N/A')}")
            print(f"   Active Accounts: {factors.get('active_accounts', 'N/A')}")
            print(f"   Suspended Accounts: {factors.get('suspended_accounts', 'N/A')}")
        
        if queue_count > 0 and messages:
            print(f"\n📋 Sample Queue Messages (first 5):")
            print("-" * 60)
            for i, msg in enumerate(messages[:5], 1):
                if isinstance(msg, dict):
                    msg_id = msg.get('id', f'msg_{i}')
                    status = msg.get('status', 'Unknown')
                    print(f"{i}. {msg_id}: {status}")
        
        # توصيات بناءً على حجم الطابور
        if queue_count > 100:
            print(f"\n💡 Recommendations:")
            print("1. Check for DNS resolution issues")
            print("2. Verify recipient domains are valid")
            print("3. Review Exim configuration")
            
            if queue_count > 500:
                print("4. ⚠️  URGENT: Consider clearing frozen messages")
                print("5. ⚠️  URGENT: Check server resources")
        
        # ملاحظات خاصة بالطريقة المستخدمة
        if 'estimated' in method.lower():
            print(f"\n⚠️  Note: This is an estimated value based on account analysis.")
            print("   For exact mail queue information, check server logs directly.")
    
    else:
        print(f"❌ Error checking mail queue: {queue_status.get('error', 'Unknown error')}")
        print(f"\n💡 Alternative Solutions:")
        print("1. Check server logs directly via SSH")
        print("2. Use 'mailq' command on server")
        print("3. Check Exim logs in /var/log/exim_mainlog")

def mail_queue_check_all_servers(online_servers):
    """فحص طابور البريد لجميع السيرفرات"""
    print("\n📮 Mail Queue Check - All Servers")
    print("=" * 50)
    
    total_queued = 0
    servers_with_issues = 0
    
    for server_name, server in online_servers.items():
        print(f"\n🖥️  {server_name}:", end=" ")
        
        queue_status = get_mail_queue_status(server)
        
        if queue_status.get("success"):
            queue_count = queue_status['queue_count']
            total_queued += queue_count
            
            if queue_count == 0:
                print("✅ Empty queue")
            elif queue_count < 50:
                print(f"🟡 {queue_count} messages (normal)")
            elif queue_count < 200:
                print(f"🟠 {queue_count} messages (elevated)")
                servers_with_issues += 1
            else:
                print(f"🔴 {queue_count} messages (CRITICAL)")
                servers_with_issues += 1
        else:
            print("❌ Error checking queue")
            servers_with_issues += 1
    
    print(f"\n📊 QUEUE SUMMARY:")
    print(f"Servers Checked: {len(online_servers)}")
    print(f"Servers with Issues: {servers_with_issues}")
    print(f"Total Queued Messages: {total_queued}")

def quick_email_health_check_all_servers(servers):
    """فحص سريع لصحة الإيميل لجميع السيرفرات"""
    print("\n🎯 Quick Email Health Check - All Servers")
    print("=" * 60)
    
    online_servers = get_online_servers(servers)
    if not online_servers:
        return
    
    total_issues = 0
    servers_with_issues = 0
    
    for server_name, server in online_servers.items():
        print(f"\n🖥️  {server_name} ({server['ip']})")
        print("-" * 40)
        
        server_issues = 0
        
        # 1. فحص طابور البريد
        print("📮 Mail Queue:", end=" ")
        queue_status = get_mail_queue_status(server)
        if queue_status.get("success"):
            queue_count = queue_status['queue_count']
            if queue_count == 0:
                print("✅ Empty")
            elif queue_count < 50:
                print(f"🟡 {queue_count} messages (normal)")
            elif queue_count < 200:
                print(f"🟠 {queue_count} messages (elevated)")
                server_issues += 1
            else:
                print(f"🔴 {queue_count} messages (CRITICAL)")
                server_issues += 2
        else:
            print("❌ Error checking")
            server_issues += 1
        
        # 2. فحص الإيميلات الفاشلة
        print("🚫 Failed Emails:", end=" ")
        failed_report = get_failed_emails_report(server, 1)  # آخر 24 ساعة
        if failed_report.get("success"):
            failures = failed_report['total_failures']
            if failures == 0:
                print("✅ None")
            elif failures < 10:
                print(f"🟡 {failures} (low)")
            elif failures < 50:
                print(f"🟠 {failures} (moderate)")
                server_issues += 1
            else:
                print(f"🔴 {failures} (high)")
                server_issues += 2
        else:
            print("❌ Error checking")
            server_issues += 1
        
        # 3. فحص الحسابات المشبوهة
        print("⚠️  Suspicious Accounts:", end=" ")
        spam_analysis = analyze_potential_spam_accounts(server)
        if spam_analysis.get("success"):
            suspicious = spam_analysis['suspicious_count']
            if suspicious == 0:
                print("✅ None")
            elif suspicious < 3:
                print(f"🟡 {suspicious} (low risk)")
            elif suspicious < 5:
                print(f"🟠 {suspicious} (moderate risk)")
                server_issues += 1
            else:
                print(f"🔴 {suspicious} (high risk)")
                server_issues += 2
        else:
            print("❌ Error checking")
            server_issues += 1
        
        # 4. فحص القوائم السوداء
        print("🔍 Blacklist Status:", end=" ")
        blacklist_check = check_blacklist_status(server)
        if blacklist_check.get("success"):
            results = blacklist_check["results"]
            listed_count = len([r for r in results.values() if "🔴 LISTED" in r])
            if listed_count == 0:
                print("✅ Clean")
            else:
                print(f"🔴 Listed on {listed_count} blacklist(s)")
                server_issues += 3  # Critical issue
        else:
            print("❌ Error checking")
            server_issues += 1
        
        # تقييم حالة السيرفر
        if server_issues == 0:
            print("🟢 Overall Status: HEALTHY")
        elif server_issues < 3:
            print("🟡 Overall Status: MONITOR")
            servers_with_issues += 1
        elif server_issues < 5:
            print("🟠 Overall Status: ATTENTION NEEDED")
            servers_with_issues += 1
        else:
            print("🔴 Overall Status: CRITICAL")
            servers_with_issues += 1
        
        total_issues += server_issues
    
    # ملخص عام
    print(f"\n📊 SUMMARY REPORT")
    print("=" * 40)
    print(f"Total Servers Checked: {len(online_servers)}")
    print(f"Servers with Issues: {servers_with_issues}")
    print(f"Total Issues Found: {total_issues}")
    
    if total_issues == 0:
        print(f"\n✅ ALL SYSTEMS HEALTHY")
        print("🛡️  All email systems are operating optimally")
    elif total_issues < 5:
        print(f"\n🟡 MINOR ISSUES DETECTED")
        print("👀 Regular monitoring recommended")
    elif total_issues < 10:
        print(f"\n🟠 MODERATE ISSUES DETECTED")
        print("🔧 Address issues within 24 hours")
    else:
        print(f"\n🔴 CRITICAL ISSUES DETECTED")
        print("🚨 IMMEDIATE ACTION REQUIRED")

def complete_email_audit_menu(servers):
    """قائمة المراجعة الشاملة للإيميل"""
    print("\n📋 Complete Email Audit")
    print("=" * 50)
    
    online_servers = get_online_servers(servers)
    if not online_servers:
        return
    
    print("1. 🔍 Audit specific server")
    print("2. 🌐 Audit all servers")
    print("0. 🚪 Back")
    
    choice = input("\nChoose option: ").strip()
    
    if choice == "1":
        server_choice = input(f"\nChoose server ({'/'.join(online_servers.keys())}): ").strip()
        if server_choice in online_servers:
            complete_audit_for_server(online_servers[server_choice], server_choice)
        else:
            print("❌ Invalid server choice!")
    
    elif choice == "2":
        complete_audit_all_servers(online_servers)

def complete_audit_for_server(server, server_name):
    """مراجعة شاملة لسيرفر محدد"""
    print(f"\n📋 Complete Email Audit - {server_name}")
    print("=" * 60)
    
    audit_results = {
        'server_name': server_name,
        'ip': server['ip'],
        'audit_date': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        'tests': {}
    }
    
    print("🔍 Running comprehensive email audit...")
    
    # 1. فحص طابور البريد
    print("\n1. 📮 Mail Queue Analysis...")
    queue_status = get_mail_queue_status(server)
    audit_results['tests']['mail_queue'] = queue_status
    
    # 2. فحص الإيميلات الفاشلة
    print("2. 🚫 Failed Emails Analysis...")
    failed_report = get_failed_emails_report(server, 7)
    audit_results['tests']['failed_emails'] = failed_report
    
    # 3. تحليل السبام
    print("3. ⚠️  Spam Risk Analysis...")
    spam_analysis = analyze_potential_spam_accounts(server)
    audit_results['tests']['spam_analysis'] = spam_analysis
    
    # 4. فحص القوائم السوداء
    print("4. 🔍 Blacklist Check...")
    blacklist_check = check_blacklist_status(server)
    audit_results['tests']['blacklist'] = blacklist_check
    
    # عرض النتائج
    print(f"\n📊 AUDIT RESULTS SUMMARY")
    print("=" * 50)
    
    # طابور البريد
    if queue_status.get('success'):
        queue_count = queue_status['queue_count']
        if queue_count == 0:
            print("📮 Mail Queue: ✅ Empty (Excellent)")
        elif queue_count < 50:
            print(f"📮 Mail Queue: 🟡 {queue_count} messages (Normal)")
        else:
            print(f"📮 Mail Queue: 🔴 {queue_count} messages (Needs attention)")
    else:
        print("📮 Mail Queue: ❌ Error checking")
    
    # الإيميلات الفاشلة
    if failed_report.get('success'):
        failures = failed_report['total_failures']
        if failures == 0:
            print("🚫 Failed Emails: ✅ None (Excellent)")
        elif failures < 50:
            print(f"🚫 Failed Emails: 🟡 {failures} in 7 days (Acceptable)")
        else:
            print(f"🚫 Failed Emails: 🔴 {failures} in 7 days (High)")
    else:
        print("🚫 Failed Emails: ❌ Error checking")
    
    # تحليل السبام
    if spam_analysis.get('success'):
        suspicious = spam_analysis['suspicious_count']
        if suspicious == 0:
            print("⚠️  Spam Risk: ✅ No suspicious accounts")
        elif suspicious < 5:
            print(f"⚠️  Spam Risk: 🟡 {suspicious} suspicious accounts")
        else:
            print(f"⚠️  Spam Risk: 🔴 {suspicious} suspicious accounts")
    else:
        print("⚠️  Spam Risk: ❌ Error checking")
    
    # القوائم السوداء
    if blacklist_check.get('success'):
        results = blacklist_check["results"]
        listed_count = len([r for r in results.values() if "🔴 LISTED" in r])
        if listed_count == 0:
            print("🔍 Blacklist: ✅ Clean on all lists")
        else:
            print(f"🔍 Blacklist: 🔴 Listed on {listed_count} blacklists")
    else:
        print("🔍 Blacklist: ❌ Error checking")
    
    # التوصيات
    print(f"\n💡 RECOMMENDATIONS:")
    recommendations = []
    
    if queue_status.get('success') and queue_status['queue_count'] > 100:
        recommendations.append("Clear mail queue backlog")
    
    if failed_report.get('success') and failed_report['total_failures'] > 100:
        recommendations.append("Investigate high email failure rate")
    
    if spam_analysis.get('success') and spam_analysis['suspicious_count'] > 5:
        recommendations.append("Review suspicious accounts")
    
    if blacklist_check.get('success'):
        results = blacklist_check["results"]
        listed_count = len([r for r in results.values() if "🔴 LISTED" in r])
        if listed_count > 0:
            recommendations.append("URGENT: Address blacklist issues")
    
    if recommendations:
        for i, rec in enumerate(recommendations, 1):
            print(f"{i}. {rec}")
    else:
        print("✅ No critical issues found - system is healthy")
    
    if confirm_action("\nExport complete audit report?"):
        # تصدير التقرير
        headers = ["Check", "Status", "Details", "Recommendation"]
        data_rows = [
            ["Mail Queue", 
             f"{queue_status['queue_count']} messages" if queue_status.get('success') else "Error",
             "Queue status check", 
             "Monitor if > 100 messages"],
            ["Failed Emails", 
             f"{failed_report['total_failures']} failures" if failed_report.get('success') else "Error",
             "7-day failure analysis", 
             "Investigate if > 100 failures"],
            ["Spam Risk", 
             f"{spam_analysis['suspicious_count']} suspicious" if spam_analysis.get('success') else "Error",
             "Account risk analysis", 
             "Review if > 5 accounts"],
            ["Blacklist", 
             f"{listed_count} listings" if blacklist_check.get('success') else "Error",
             "Reputation check", 
             "Address immediately if listed"]
        ]
        
        export_to_excel(data_rows, headers, f"email_audit_{server_name}", "Email Audit Report")

def complete_audit_all_servers(online_servers):
    """مراجعة شاملة لجميع السيرفرات"""
    print("\n📋 Complete Email Audit - All Servers")
    print("=" * 60)
    
    print("🔍 Running audit on all servers...")
    
    all_results = []
    
    for server_name, server in online_servers.items():
        print(f"\n🖥️  Auditing {server_name}...")
        
        # فحص سريع لكل سيرفر
        queue_status = get_mail_queue_status(server)
        failed_report = get_failed_emails_report(server, 1)  # يوم واحد فقط للسرعة
        spam_analysis = analyze_potential_spam_accounts(server)
        blacklist_check = check_blacklist_status(server)
        
        # حساب نقاط الصحة
        health_score = 100
        issues = []
        
        if queue_status.get('success') and queue_status['queue_count'] > 100:
            health_score -= 25
            issues.append("High queue")
        
        if failed_report.get('success') and failed_report['total_failures'] > 50:
            health_score -= 20
            issues.append("High failures")
        
        if spam_analysis.get('success') and spam_analysis['suspicious_count'] > 3:
            health_score -= 15
            issues.append("Suspicious accounts")
        
        if blacklist_check.get('success'):
            results = blacklist_check["results"]
            listed_count = len([r for r in results.values() if "🔴 LISTED" in r])
            if listed_count > 0:
                health_score -= 40
                issues.append("Blacklisted")
        
        # تحديد الحالة
        if health_score >= 90:
            status = "🟢 EXCELLENT"
        elif health_score >= 75:
            status = "🟡 GOOD"
        elif health_score >= 60:
            status = "🟠 FAIR"
        else:
            status = "🔴 POOR"
        
        print(f"   Status: {status} (Score: {health_score}/100)")
        if issues:
            print(f"   Issues: {', '.join(issues)}")
        
        all_results.append({
            'server': server_name,
            'ip': server['ip'],
            'health_score': health_score,
            'status': status,
            'issues': issues,
            'queue_count': queue_status.get('queue_count', 'Error') if queue_status.get('success') else 'Error',
            'failed_count': failed_report.get('total_failures', 'Error') if failed_report.get('success') else 'Error',
            'suspicious_count': spam_analysis.get('suspicious_count', 'Error') if spam_analysis.get('success') else 'Error',
            'blacklist_count': len([r for r in blacklist_check["results"].values() if "🔴 LISTED" in r]) if blacklist_check.get('success') else 'Error'
        })
    
    # ملخص عام
    print(f"\n📊 OVERALL AUDIT SUMMARY")
    print("=" * 60)
    
    avg_score = sum(r['health_score'] for r in all_results) / len(all_results)
    excellent_count = sum(1 for r in all_results if r['health_score'] >= 90)
    poor_count = sum(1 for r in all_results if r['health_score'] < 60)
    
    print(f"Servers Audited: {len(all_results)}")
    print(f"Average Health Score: {avg_score:.1f}/100")
    print(f"Excellent Servers: {excellent_count}")
    print(f"Poor Servers: {poor_count}")
    
    if confirm_action("\nExport complete audit report?"):
        headers = ["Server", "IP", "Health Score", "Status", "Queue", "Failed Emails", "Suspicious", "Blacklisted", "Issues", "Audit Date"]
        data_rows = []
        
        for result in all_results:
            data_rows.append([
                result['server'],
                result['ip'],
                result['health_score'],
                result['status'],
                result['queue_count'],
                result['failed_count'],
                result['suspicious_count'],
                result['blacklist_count'],
                '; '.join(result['issues']) if result['issues'] else 'None',
                datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ])
        
        export_to_excel(data_rows, headers, "complete_email_audit", "Complete Email Audit")

def detailed_domain_analysis(server, domain, server_name):
    """تحليل مفصل لدومين محدد"""
    print(f"\n🔍 DETAILED ANALYSIS FOR: {domain}")
    print("=" * 80)
    
    try:
        # البحث عن الحساب في السيرفر المحدد
        accounts = list_accounts(server)
        if not accounts:
            print(f"❌ No accounts found on server {server_name}")
            return
        
        # البحث عن الحساب المحدد
        acct = None
        for account in accounts:
            if account.get('domain', '').lower() == domain.lower():
                acct = account
                break
        
        if not acct:
            print(f"❌ Domain {domain} not found on server {server_name}")
            return
        
        print(f"📋 Account Information:")
        print(f"   🌐 Domain: {domain}")
        print(f"   👤 User: {acct['user']}")
        print(f"   📧 Email: {acct.get('email', 'N/A')}")
        print(f"   📦 Package: {acct.get('plan', 'N/A')}")
        print(f"   💾 Disk Used: {acct.get('diskused', 'N/A')}MB")
        print(f"   📅 Created: {datetime.fromtimestamp(int(acct.get('unix_startdate', 0))).strftime('%Y-%m-%d') if acct.get('unix_startdate') else 'Unknown'}")
        print(f"   🔴 Suspended: {'Yes' if acct.get('suspended', 0) == 1 else 'No'}")
        print(f"   🖥️  Server: {server_name} ({server['ip']})")
        
        # فحص حسابات الإيميل
        print(f"\n📧 Email Accounts Analysis:")
        print("-" * 60)
        
        emails = list_email_accounts(server, acct['user'], domain)
        if emails:
            print(f"📊 Total Email Accounts: {len(emails)}")
            
            # تصنيف الإيميلات
            active_emails = [e for e in emails if not e.get('suspended', False)]
            suspended_emails = [e for e in emails if e.get('suspended', False)]
            
            print(f"   🟢 Active: {len(active_emails)}")
            print(f"   🔴 Suspended: {len(suspended_emails)}")
            
            # حساب استخدام القرص
            total_disk_used = sum(e.get('diskused', 0) for e in emails)
            total_quota = sum(e.get('diskquota', 0) for e in emails)
            
            print(f"   💾 Total Disk Used: {total_disk_used:.2f}MB")
            print(f"   💾 Total Quota: {total_quota:.2f}MB")
            
            if total_quota > 0:
                usage_percent = (total_disk_used / total_quota) * 100
                print(f"   📊 Usage: {usage_percent:.1f}%")
            
            # عرض تفاصيل الإيميلات
            print(f"\n📋 Email Accounts Details:")
            print("-" * 80)
            print(f"{'#':<3} {'Email':<30} {'Status':<12} {'Used (MB)':<12} {'Quota (MB)':<12} {'Usage %':<10}")
            print("-" * 80)
            
            for i, email in enumerate(emails[:20], 1):  # أول 20 إيميل
                status = "🔴 Suspended" if email.get('suspended', False) else "🟢 Active"
                used = email.get('diskused', 0)
                quota = email.get('diskquota', 0)
                usage = (used / quota * 100) if quota > 0 else 0
                
                print(f"{i:<3} {email['email']:<30} {status:<12} {used:<12.2f} {quota:<12.2f} {usage:<10.1f}%")
            
            if len(emails) > 20:
                print(f"... and {len(emails) - 20} more email accounts")
        else:
            print("❌ No email accounts found or error loading accounts")
        
        # تحليل المخاطر
        print(f"\n⚠️  Risk Analysis:")
        print("-" * 60)
        
        risk_score = 0
        risk_factors = []
        
        # 1. الحساب معلق
        if acct.get('suspended', 0) == 1:
            risk_score += 5
            risk_factors.append("Account suspended")
        
        # 2. استخدام قرص عالي
        try:
            disk_used = float(acct.get('diskused', 0))
            if disk_used > 1000:  # أكثر من 1GB
                risk_score += 3
                risk_factors.append(f"High disk usage ({disk_used:.0f}MB)")
            elif disk_used > 500:
                risk_score += 2
                risk_factors.append(f"Moderate disk usage ({disk_used:.0f}MB)")
        except:
            pass
        
        # 3. عدد كبير من حسابات الإيميل
        if emails:
            email_count = len(emails)
            if email_count > 50:
                risk_score += 5
                risk_factors.append(f"High email count ({email_count})")
            elif email_count > 20:
                risk_score += 3
                risk_factors.append(f"Moderate email count ({email_count})")
            elif email_count > 10:
                risk_score += 1
                risk_factors.append(f"Elevated email count ({email_count})")
        
        # 4. أنماط الدومين المشبوهة
        domain_lower = domain.lower()
        suspicious_patterns = ['temp', 'test', 'spam', 'bulk', 'mail', 'send', 'newsletter']
        if any(pattern in domain_lower for pattern in suspicious_patterns):
            risk_score += 4
            risk_factors.append(f"Suspicious domain pattern: {domain}")
        
        # عرض نتائج تحليل المخاطر
        print(f"🚨 Risk Score: {risk_score}/25")
        
        if risk_factors:
            print(f"⚠️  Risk Factors:")
            for factor in risk_factors:
                print(f"   • {factor}")
        else:
            print("✅ No risk factors detected")
        
        # تقدير الإيميلات الفاشلة
        estimated_failures = 0
        if risk_score >= 8:
            estimated_failures = int(risk_score * 1.5)
        elif risk_score >= 5:
            estimated_failures = int(risk_score * 1.2)
        elif risk_score >= 3:
            estimated_failures = int(risk_score * 0.8)
        
        print(f"\n📊 Estimated Email Failures:")
        print(f"   📈 Based on risk analysis: {estimated_failures} failures")
        
        if emails:
            # تقدير إضافي بناءً على عدد الإيميلات
            email_based_failures = len(emails) * 0.3  # 30% من الإيميلات قد تفشل
            print(f"   📧 Based on email count: {email_based_failures:.1f} failures")
            
            # التقدير الأعلى
            final_estimate = max(estimated_failures, email_based_failures)
            print(f"   🎯 Final Estimate: {final_estimate:.1f} failures")
        
        # توصيات
        print(f"\n💡 Recommendations:")
        if risk_score >= 15:
            print("   🚨 CRITICAL: Immediate action required!")
            print("   • Suspend account immediately")
            print("   • Review all email accounts")
            print("   • Check for spam activity")
        elif risk_score >= 10:
            print("   ⚠️  HIGH RISK: Action needed within 24 hours")
            print("   • Consider suspending account")
            print("   • Monitor email activity closely")
            print("   • Review email accounts")
        elif risk_score >= 5:
            print("   🟡 MODERATE RISK: Monitor closely")
            print("   • Regular monitoring recommended")
            print("   • Check email activity weekly")
        else:
            print("   ✅ LOW RISK: Regular monitoring sufficient")
            print("   • Standard monitoring procedures")
            print("   • Quarterly review recommended")
        
        # خيارات إضافية
        print(f"\n🔧 Available Actions:")
        print("1. 🚫 Suspend this account")
        print("2. 🔑 Change account password")
        print("3. 📧 Manage email accounts")
        print("4. 📊 Export detailed report")
        print("0. 🚪 Back to main menu")
        
        action_choice = input("\nChoose action: ").strip()
        
        if action_choice == "1":
            if confirm_action(f"Suspend account {domain}?"):
                params = {"user": acct["user"], "reason": "High risk account - detailed analysis"}
                result = whm_api_call(server, "suspendacct", params)
                if "error" not in result:
                    print(f"✅ Account {domain} suspended successfully")
                else:
                    print(f"❌ Failed to suspend account: {result['error']}")
        
        elif action_choice == "2":
            if confirm_action(f"Change password for {domain}?"):
                new_password = generate_strong_password(16)
                params = {"user": acct["user"], "password": new_password}
                result = whm_api_call(server, "passwd", params)
                if "error" not in result:
                    print("=" * 50)
                    print(f"🌐 Domain: {domain}")
                    print(f"👤 cPanel User: {acct['user']}")
                    print(f"🔑 New Password: {new_password}")
                    print(f"💻 cPanel URL: https://{domain}:2083")
                    print(f"📧 Webmail URL: https://webmail.{domain}")
                    print("=" * 50)
                else:
                    print(f"❌ Failed to change password: {result['error']}")
        
        elif action_choice == "3":
            print(f"\n📧 Email Management for {domain}")
            print("=" * 50)
            if emails:
                print(f"Total emails: {len(emails)}")
                print("Use the main email management menu for detailed operations")
            else:
                print("No email accounts found")
        
        elif action_choice == "4":
            # تصدير التقرير المفصل
            headers = ["Domain", "User", "Risk Score", "Email Count", "Disk Used", "Suspended", "Risk Factors", "Estimated Failures", "Analysis Date"]
            data_rows = [[
                domain,
                acct['user'],
                risk_score,
                len(emails) if emails else 0,
                acct.get('diskused', 'N/A'),
                "Yes" if acct.get('suspended', 0) == 1 else "No",
                "; ".join(risk_factors) if risk_factors else "None",
                estimated_failures,
                datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ]]
            
            export_to_excel(data_rows, headers, f"domain_analysis_{domain.replace('.', '_')}", f"Domain Analysis - {domain}")
            export_to_csv(data_rows, headers, f"domain_analysis_{domain.replace('.', '_')}")
            print(f"✅ Detailed report exported for {domain}")
        
    except Exception as e:
        print(f"❌ Error in detailed domain analysis: {str(e)}")
        logging.error(f"Error analyzing domain {domain}: {str(e)}")

def domain_mail_queue_analysis(server, domain, server_name):
    """تحليل طابور البريد لدومين محدد"""
    print(f"\n📮 MAIL QUEUE ANALYSIS FOR: {domain}")
    print("=" * 80)
    
    try:
        # البحث عن الحساب في السيرفر المحدد
        accounts = list_accounts(server)
        if not accounts:
            print(f"❌ No accounts found on server {server_name}")
            return
        
        # البحث عن الحساب المحدد
        acct = None
        for account in accounts:
            if account.get('domain', '').lower() == domain.lower():
                acct = account
                break
        
        if not acct:
            print(f"❌ Domain {domain} not found on server {server_name}")
            return
        
        print(f"📋 Domain Information:")
        print(f"   🌐 Domain: {domain}")
        print(f"   👤 User: {acct['user']}")
        print(f"   🖥️  Server: {server_name} ({server['ip']})")
        print(f"   🔴 Suspended: {'Yes' if acct.get('suspended', 0) == 1 else 'No'}")
        
        # فحص طابور البريد للسيرفر
        print(f"\n📮 Server Mail Queue Status:")
        print("-" * 60)
        
        queue_status = get_mail_queue_status_advanced(server)
        if queue_status.get("success"):
            queue_count = queue_status['queue_count']
            method = queue_status.get('method', 'unknown')
            note = queue_status.get('note', '')
            
            print(f"📊 Total Queue Messages: {queue_count}")
            print(f"🔧 Method Used: {method}")
            if note:
                print(f"📝 Note: {note}")
            
            # تصنيف حالة الطابور
            if queue_count == 0:
                print("✅ Queue Status: Empty (Healthy)")
            elif queue_count < 10:
                print("🟡 Queue Status: Normal")
            elif queue_count < 50:
                print("🟠 Queue Status: Moderate")
            elif queue_count < 100:
                print("🔴 Queue Status: High")
            else:
                print("🚨 Queue Status: Critical")
        else:
            print(f"❌ Error checking mail queue: {queue_status.get('error')}")
        
        # فحص حسابات الإيميل للدومين
        print(f"\n📧 Email Accounts for {domain}:")
        print("-" * 60)
        
        emails = list_email_accounts(server, acct['user'], domain)
        if emails:
            print(f"📊 Total Email Accounts: {len(emails)}")
            
            # تصنيف الإيميلات
            active_emails = [e for e in emails if not e.get('suspended', False)]
            suspended_emails = [e for e in emails if e.get('suspended', False)]
            
            print(f"   🟢 Active: {len(active_emails)}")
            print(f"   🔴 Suspended: {len(suspended_emails)}")
            
            # حساب استخدام القرص
            total_disk_used = sum(e.get('diskused', 0) for e in emails)
            total_quota = sum(e.get('diskquota', 0) for e in emails)
            
            print(f"   💾 Total Disk Used: {total_disk_used:.2f}MB")
            print(f"   💾 Total Quota: {total_quota:.2f}MB")
            
            if total_quota > 0:
                usage_percent = (total_disk_used / total_quota) * 100
                print(f"   📊 Usage: {usage_percent:.1f}%")
                
                # تحذير إذا كان الاستخدام عالي
                if usage_percent > 90:
                    print("   ⚠️  WARNING: High disk usage!")
                elif usage_percent > 80:
                    print("   🟡 WARNING: Moderate disk usage")
        else:
            print("❌ No email accounts found")
        
        # تحليل مشاكل الإيميلات
        print(f"\n🚫 Email Delivery Issues Analysis:")
        print("-" * 60)
        
        # تقدير الإيميلات الفاشلة بناءً على عوامل متعددة
        estimated_failures = 0
        failure_factors = []
        
        # 1. الحساب معلق
        if acct.get('suspended', 0) == 1:
            estimated_failures += 15
            failure_factors.append("Account suspended (major impact)")
        
        # 2. استخدام قرص عالي
        try:
            disk_used = float(acct.get('diskused', 0))
            if disk_used > 1000:
                estimated_failures += 10
                failure_factors.append(f"High disk usage ({disk_used:.0f}MB)")
            elif disk_used > 500:
                estimated_failures += 5
                failure_factors.append(f"Moderate disk usage ({disk_used:.0f}MB)")
        except:
            pass
        
        # 3. عدد كبير من حسابات الإيميل
        if emails:
            email_count = len(emails)
            if email_count > 50:
                estimated_failures += 20
                failure_factors.append(f"High email count ({email_count})")
            elif email_count > 20:
                estimated_failures += 10
                failure_factors.append(f"Moderate email count ({email_count})")
            elif email_count > 10:
                estimated_failures += 5
                failure_factors.append(f"Elevated email count ({email_count})")
        
        # 4. أنماط الدومين المشبوهة
        domain_lower = domain.lower()
        suspicious_patterns = ['temp', 'test', 'spam', 'bulk', 'mail', 'send', 'newsletter']
        if any(pattern in domain_lower for pattern in suspicious_patterns):
            estimated_failures += 15
            failure_factors.append(f"Suspicious domain pattern: {domain}")
        
        # 5. حالة طابور البريد
        if queue_status.get("success"):
            queue_count = queue_status['queue_count']
            if queue_count > 100:
                estimated_failures += 25
                failure_factors.append(f"Critical mail queue ({queue_count} messages)")
            elif queue_count > 50:
                estimated_failures += 15
                failure_factors.append(f"High mail queue ({queue_count} messages)")
            elif queue_count > 20:
                estimated_failures += 10
                failure_factors.append(f"Moderate mail queue ({queue_count} messages)")
        
        print(f"📊 Estimated Email Failures: {estimated_failures}")
        
        if failure_factors:
            print(f"🚫 Failure Factors:")
            for factor in failure_factors:
                print(f"   • {factor}")
        else:
            print("✅ No failure factors detected")
        
        # توصيات لحل المشاكل
        print(f"\n💡 Recommendations to Fix Email Issues:")
        
        if estimated_failures >= 50:
            print("   🚨 CRITICAL ISSUES - IMMEDIATE ACTION REQUIRED:")
            print("   • Suspend account immediately")
            print("   • Clear mail queue")
            print("   • Review all email accounts")
            print("   • Check for spam activity")
        elif estimated_failures >= 30:
            print("   🔴 HIGH PRIORITY - ACTION NEEDED WITHIN 1 HOUR:")
            print("   • Consider suspending account")
            print("   • Monitor mail queue closely")
            print("   • Review email accounts")
            print("   • Check disk usage")
        elif estimated_failures >= 15:
            print("   🟠 MODERATE PRIORITY - ACTION NEEDED WITHIN 24 HOURS:")
            print("   • Monitor account activity")
            print("   • Check mail queue regularly")
            print("   • Review email accounts")
            print("   • Consider disk cleanup")
        elif estimated_failures >= 5:
            print("   🟡 LOW PRIORITY - MONITOR CLOSELY:")
            print("   • Regular monitoring recommended")
            print("   • Check weekly")
            print("   • Monitor disk usage")
        else:
            print("   ✅ NO IMMEDIATE ACTION REQUIRED:")
            print("   • Standard monitoring procedures")
            print("   • Quarterly review recommended")
        
        # خيارات إضافية
        print(f"\n🔧 Available Actions:")
        print("1. 🚫 Suspend this account")
        print("2. 🔑 Change account password")
        print("3. 📧 Manage email accounts")
        print("4. 📊 Export mail queue report")
        print("5. 🔍 Check other domains on this server")
        print("0. 🚪 Back to main menu")
        
        action_choice = input("\nChoose action: ").strip()
        
        if action_choice == "1":
            if confirm_action(f"Suspend account {domain}?"):
                params = {"user": acct["user"], "reason": "Email delivery issues - mail queue analysis"}
                result = whm_api_call(server, "suspendacct", params)
                if "error" not in result:
                    print(f"✅ Account {domain} suspended successfully")
                else:
                    print(f"❌ Failed to suspend account: {result['error']}")
        
        elif action_choice == "2":
            if confirm_action(f"Change password for {domain}?"):
                new_password = generate_strong_password(16)
                params = {"user": acct["user"], "password": new_password}
                result = whm_api_call(server, "passwd", params)
                if "error" not in result:
                    print("=" * 50)
                    print(f"🌐 Domain: {domain}")
                    print(f"👤 cPanel User: {acct['user']}")
                    print(f"🔑 New Password: {new_password}")
                    print(f"💻 cPanel URL: https://{domain}:2083")
                    print(f"📧 Webmail URL: https://webmail.{domain}")
                    print("=" * 50)
                else:
                    print(f"❌ Failed to change password: {result['error']}")
        
        elif action_choice == "3":
            print(f"\n📧 Email Management for {domain}")
            print("=" * 50)
            if emails:
                print(f"Total emails: {len(emails)}")
                print("Use the main email management menu for detailed operations")
            else:
                print("No email accounts found")
        
        elif action_choice == "4":
            # تصدير تقرير طابور البريد
            headers = ["Domain", "User", "Server", "Mail Queue Count", "Email Count", "Estimated Failures", "Failure Factors", "Analysis Date"]
            data_rows = [[
                domain,
                acct['user'],
                server_name,
                queue_status.get('queue_count', 0) if queue_status.get('success') else 'Error',
                len(emails) if emails else 0,
                estimated_failures,
                "; ".join(failure_factors) if failure_factors else "None",
                datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ]]
            
            export_to_excel(data_rows, headers, f"mail_queue_analysis_{domain.replace('.', '_')}", f"Mail Queue Analysis - {domain}")
            export_to_csv(data_rows, headers, f"mail_queue_analysis_{domain.replace('.', '_')}")
            print(f"✅ Mail queue report exported for {domain}")
        
        elif action_choice == "5":
            print(f"\n🔍 Checking other domains on server {server_name}...")
            accounts = list_accounts(server_name)
            if accounts:
                print(f"📊 Total accounts on server: {len(accounts)}")
                print("Use the main menu to analyze other domains")
            else:
                print("No accounts found on this server")
        
    except Exception as e:
        print(f"❌ Error in mail queue analysis: {str(e)}")
        logging.error(f"Error analyzing mail queue for domain {domain}: {str(e)}")

def compare_analysis_methods(server, server_name):
    """مقارنة طرق التحليل المختلفة"""
    print(f"\n🔍 ANALYSIS METHODS COMPARISON - {server_name}")
    print("=" * 80)
    
    try:
        print("📊 Running both analysis methods...")
        print()
        
        # الطريقة الأولى: تحليل الحسابات المفصل
        print("🔍 METHOD 1: Detailed Account Analysis")
        print("-" * 50)
        account_analysis = analyze_failed_emails_by_accounts(server, 7)
        
        if account_analysis.get("success"):
            print(f"✅ Total Accounts: {account_analysis['total_accounts']}")
            print(f"✅ Problematic Accounts: {account_analysis['problematic_count']}")
            print(f"✅ Estimated Failures: {account_analysis['total_failures']}")
            print(f"✅ Method: {account_analysis['method']}")
            print(f"✅ Note: {account_analysis['note']}")
        else:
            print(f"❌ Error: {account_analysis.get('error')}")
        
        print()
        
        # الطريقة الثانية: التقدير السريع
        print("📊 METHOD 2: Quick Estimation Method")
        print("-" * 50)
        quick_report = get_failed_emails_report(server, 7)
        
        if quick_report.get("success"):
            print(f"✅ Total Failures: {quick_report['total_failures']}")
            print(f"✅ Bounces: {quick_report['bounces']}")
            print(f"✅ Rejects: {quick_report['rejects']}")
            print(f"✅ Source: {quick_report['source']}")
            print(f"✅ Note: {quick_report['note']}")
            
            if 'details' in quick_report:
                details = quick_report['details']
                print(f"✅ Details:")
                print(f"   • Suspended Accounts: {details.get('suspended_accounts', 'N/A')}")
                print(f"   • High Disk Accounts: {details.get('high_disk_accounts', 'N/A')}")
                print(f"   • High Email Accounts: {details.get('high_email_accounts', 'N/A')}")
                print(f"   • Total Accounts: {details.get('total_accounts', 'N/A')}")
        else:
            print(f"❌ Error: {quick_report.get('error')}")
        
        print()
        
        # مقارنة النتائج
        print("📊 RESULTS COMPARISON")
        print("=" * 50)
        
        if account_analysis.get("success") and quick_report.get("success"):
            account_failures = account_analysis['total_failures']
            quick_failures = quick_report['total_failures']
            
            print(f"🔍 Account Analysis Method: {account_failures} failures")
            print(f"🔍 Quick Estimation Method: {quick_failures} failures")
            print(f"📊 Difference: {abs(account_failures - quick_failures)} failures")
            
            if account_failures == quick_failures:
                print("✅ Both methods agree - results are consistent!")
            elif abs(account_failures - quick_failures) < 10:
                print("🟡 Minor difference - methods are mostly consistent")
            elif abs(account_failures - quick_failures) < 50:
                print("🟠 Moderate difference - some inconsistency detected")
            else:
                print("🔴 Significant difference - methods show different results")
            
            print()
            print("💡 Why Results Might Differ:")
            print("• Account Analysis: More detailed, checks each account individually")
            print("• Quick Estimation: Faster, uses statistical estimates")
            print("• Different risk thresholds and calculation methods")
            print("• Account analysis requires risk_score > 0 or account_failures > 0")
            print("• Quick estimation includes base failures for all accounts")
            
        else:
            print("❌ Cannot compare - one or both methods failed")
        
        print()
        print("🎯 RECOMMENDATIONS:")
        print("• Use Account Analysis for detailed investigation")
        print("• Use Quick Estimation for overview and monitoring")
        print("• Both methods are estimates - actual logs may differ")
        print("• Consider server-specific factors and patterns")
        
        # خيارات إضافية
        print(f"\n🔧 Available Actions:")
        print("1. 📊 Export comparison report")
        print("2. 🔍 Re-run account analysis with lower thresholds")
        print("3. 📮 Check mail queue status")
        print("0. 🚪 Back to main menu")
        
        action_choice = input("\nChoose action: ").strip()
        
        if action_choice == "1":
            # تصدير تقرير المقارنة
            headers = ["Method", "Total Failures", "Bounces", "Rejects", "Source", "Note", "Analysis Date"]
            data_rows = []
            
            if account_analysis.get("success"):
                data_rows.append([
                    "Account Analysis",
                    account_analysis['total_failures'],
                    "N/A",
                    "N/A",
                    account_analysis['method'],
                    account_analysis['note'],
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ])
            
            if quick_report.get("success"):
                data_rows.append([
                    "Quick Estimation",
                    quick_report['total_failures'],
                    quick_report['bounces'],
                    quick_report['rejects'],
                    quick_report['source'],
                    quick_report['note'],
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ])
            
            if data_rows:
                export_to_excel(data_rows, headers, f"analysis_comparison_{server_name}", f"Analysis Methods Comparison - {server_name}")
                export_to_csv(data_rows, headers, f"analysis_comparison_{server_name}")
                print(f"✅ Comparison report exported for {server_name}")
        
        elif action_choice == "2":
            # إعادة تشغيل التحليل مع عتبات أقل
            print(f"\n🔍 Re-running analysis with lower thresholds...")
            # يمكن إضافة منطق هنا لتشغيل تحليل أكثر تفصيلاً
        
        elif action_choice == "3":
            # فحص حالة طابور البريد
            print(f"\n📮 Checking mail queue status...")
            queue_status = get_mail_queue_status_advanced(server)
            if queue_status.get("success"):
                print(f"📊 Queue Messages: {queue_status['queue_count']}")
                print(f"🔧 Method: {queue_status.get('method', 'unknown')}")
                if 'note' in queue_status:
                    print(f"📝 Note: {queue_status['note']}")
            else:
                print(f"❌ Error: {queue_status.get('error')}")
        
    except Exception as e:
        print(f"❌ Error in comparison: {str(e)}")
        logging.error(f"Error comparing analysis methods: {str(e)}")

# === تحسينات لدوال تحليل الإيميل ===

def get_accurate_email_usage(server, cpanel_user, email):
    """جلب معلومات استخدام الإيميل بطريقة دقيقة ومحسنة"""
    try:
        # طرق متعددة للحصول على البيانات الدقيقة
        methods_to_try = [
            # الطريقة الأولى: Email API مع disk usage
            {
                "module": "Email",
                "function": "list_pops",
                "params": {"include_disk_usage": 1, "domain": email.split('@')[1]}
            },
            # الطريقة الثانية: Quota API
            {
                "module": "Quota",
                "function": "getquotas",
                "params": {}
            },
            # الطريقة الثالثة: Stats API
            {
                "module": "StatsBar",
                "function": "get_stats",
                "params": {"display": "email"}
            }
        ]
        
        for method in methods_to_try:
            try:
                result = cpanel_api_call(server, cpanel_user, method["module"], 
                                       method["function"], method["params"])
                
                if "error" not in result and "result" in result:
                    if method["function"] == "list_pops":
                        # البحث عن الإيميل المحدد
                        if "data" in result["result"]:
                            for email_data in result["result"]["data"]:
                                if email_data.get("email") == email:
                                    # تحويل صحيح للبايت إلى ميجابايت
                                    used_bytes = float(email_data.get("diskused", 0))
                                    quota_bytes = float(email_data.get("diskquota", 0))
                                    
                                    used_mb = used_bytes / (1024 * 1024) if used_bytes > 0 else 0
                                    
                                    # معالجة أفضل للكوتا
                                    if quota_bytes == 0 or quota_bytes > (10 * 1024 * 1024 * 1024):  # أكثر من 10GB
                                        quota_display = "Unlimited"
                                        usage_percent = "N/A"
                                    else:
                                        quota_mb = quota_bytes / (1024 * 1024)
                                        usage_percent = f"{(used_mb / quota_mb) * 100:.2f}%" if quota_mb > 0 else "0%"
                                        quota_display = f"{quota_mb:.2f}MB"
                                    
                                    return {
                                        "success": True,
                                        "quota": quota_display,
                                        "used": f"{used_mb:.2f}MB",
                                        "usage_percent": usage_percent,
                                        "used_bytes": used_bytes,
                                        "quota_bytes": quota_bytes,
                                        "method": f"{method['module']}::{method['function']}"
                                    }
                    
                    # معالجة APIs أخرى...
                    
            except Exception as method_error:
                logging.warning(f"Method {method['module']}::{method['function']} failed: {str(method_error)}")
                continue
        
        return {
            "success": False,
            "error": "All methods failed to get accurate data"
        }
        
    except Exception as e:
        logging.error(f"Error getting accurate email usage: {str(e)}")
        return {
            "success": False,
            "error": str(e)
        }

def get_detailed_domain_info(server, domain, server_name):
    """جلب معلومات مفصلة وشاملة عن الدومين"""
    try:
        # البحث عن الحساب
        accounts = list_accounts(server)
        acct = None
        for account in accounts:
            if account.get('domain', '').lower() == domain.lower():
                acct = account
                break
        
        if not acct:
            return {"success": False, "error": f"Domain {domain} not found"}
        
        # معلومات أساسية محسنة
        domain_info = {
            "domain": domain,
            "user": acct['user'],
            "server_name": server_name,
            "server_ip": server['ip'],
            "suspended": acct.get('suspended', 0) == 1,
            "creation_date": datetime.fromtimestamp(int(acct.get('unix_startdate', 0))).strftime('%Y-%m-%d %H:%M:%S') if acct.get('unix_startdate') else 'Unknown',
            "package": acct.get('plan', 'Unknown'),
            "disk_used_mb": float(acct.get('diskused', 0)),
            "disk_limit_mb": float(acct.get('disklimit', 0)),
            "email_address": acct.get('email', 'Unknown')
        }
        
        # حساب عمر الحساب
        if acct.get('unix_startdate'):
            creation_date = datetime.fromtimestamp(int(acct['unix_startdate']))
            account_age_days = (datetime.now() - creation_date).days
            domain_info["account_age_days"] = account_age_days
        else:
            domain_info["account_age_days"] = 0
        
        # معلومات الإيميل المفصلة
        emails = list_email_accounts(server, acct['user'], domain)
        if emails:
            email_details = []
            total_used_accurate = 0
            total_quota_accurate = 0
            
            for email in emails:
                email_address = email.get("email", "")
                
                # جلب معلومات دقيقة لكل إيميل
                usage_info = get_accurate_email_usage(server, acct['user'], email_address)
                
                if usage_info.get("success"):
                    email_detail = {
                        "address": email_address,
                        "user": email_address.split('@')[0],
                        "suspended": email.get("suspended", False),
                        "used_mb": float(usage_info["used"].replace("MB", "")),
                        "quota_display": usage_info["quota"],
                        "usage_percent": usage_info["usage_percent"],
                        "used_bytes": usage_info.get("used_bytes", 0),
                        "quota_bytes": usage_info.get("quota_bytes", 0)
                    }
                    
                    # جمع الإحصائيات
                    total_used_accurate += email_detail["used_mb"]
                    if usage_info["quota"] != "Unlimited":
                        quota_mb = float(usage_info["quota"].replace("MB", ""))
                        total_quota_accurate += quota_mb
                else:
                    # بيانات احتياطية
                    email_detail = {
                        "address": email_address,
                        "user": email_address.split('@')[0],
                        "suspended": email.get("suspended", False),
                        "used_mb": 0,
                        "quota_display": "Unknown",
                        "usage_percent": "N/A",
                        "used_bytes": 0,
                        "quota_bytes": 0
                    }
                
                email_details.append(email_detail)
            
            domain_info["email_accounts"] = email_details
            domain_info["email_count"] = len(emails)
            domain_info["active_emails"] = len([e for e in email_details if not e["suspended"]])
            domain_info["suspended_emails"] = len([e for e in email_details if e["suspended"]])
            domain_info["total_email_usage_mb"] = total_used_accurate
            domain_info["total_email_quota_mb"] = total_quota_accurate
            
        else:
            domain_info["email_accounts"] = []
            domain_info["email_count"] = 0
            domain_info["active_emails"] = 0
            domain_info["suspended_emails"] = 0
            domain_info["total_email_usage_mb"] = 0
            domain_info["total_email_quota_mb"] = 0
        
        # فحص DNS وMX records
        try:
            import socket
            mx_records = []
            try:
                mx_records = socket.getaddrinfo(domain, None)
                domain_info["dns_resolves"] = True
            except:
                domain_info["dns_resolves"] = False
            
            domain_info["mx_records"] = len(mx_records)
            
        except:
            domain_info["dns_resolves"] = "Unknown"
            domain_info["mx_records"] = "Unknown"
        
        # معلومات إضافية من WHM
        try:
            # جلب إحصائيات الاستخدام من WHM
            whm_stats = whm_api_call(server, "accountsummary", {"user": acct['user']})
            if "error" not in whm_stats and "data" in whm_stats:
                stats = whm_stats["data"]
                domain_info["bandwidth_used"] = stats.get("bandwidth_used", "Unknown")
                domain_info["inodes_used"] = stats.get("inodes_used", "Unknown")
                
        except Exception as whm_error:
            logging.warning(f"Could not get WHM stats: {str(whm_error)}")
        
        return {
            "success": True,
            "data": domain_info
        }
        
    except Exception as e:
        logging.error(f"Error getting detailed domain info: {str(e)}")
        return {
            "success": False,
            "error": str(e)
        }

def get_real_mail_queue_info(server):
    """محاولة جلب معلومات حقيقية عن طابور البريد"""
    try:
        # محاولات متعددة للحصول على معلومات طابور البريد الحقيقي
        queue_methods = [
            # محاولة استخدام Exim API
            {
                "function": "get_exim_configuration",
                "description": "Exim configuration check"
            },
            # محاولة استخدام Mail API
            {
                "function": "get_mail_settings", 
                "description": "Mail server settings"
            },
            # محاولة جلب Log files
            {
                "function": "get_log_files",
                "description": "Mail log analysis"
            }
        ]
        
        queue_info = {
            "method": "multiple_attempts",
            "queue_count": 0,
            "messages": [],
            "queue_size_bytes": 0,
            "oldest_message_age": 0,
            "success": False,
            "attempts": []
        }
        
        for method in queue_methods:
            try:
                # محاولة الحصول على معلومات من كل API
                result = whm_api_call(server, method["function"])
                
                queue_info["attempts"].append({
                    "method": method["function"],
                    "status": "success" if "error" not in result else "failed",
                    "description": method["description"]
                })
                
                if "error" not in result:
                    # معالجة البيانات حسب نوع API
                    if method["function"] == "get_exim_configuration":
                        # تحليل إعدادات Exim
                        if "queue" in result:
                            queue_info["queue_count"] = result.get("queue_size", 0)
                            queue_info["success"] = True
                    
                    # يمكن إضافة معالجة لـ APIs أخرى
                
            except Exception as method_error:
                queue_info["attempts"].append({
                    "method": method["function"],
                    "status": "error",
                    "error": str(method_error),
                    "description": method["description"]
                })
        
        # إذا فشلت جميع المحاولات، استخدم تقدير محسن
        if not queue_info["success"]:
            enhanced_estimate = get_enhanced_queue_estimate(server)
            queue_info.update(enhanced_estimate)
            queue_info["method"] = "enhanced_estimation"
        
        return queue_info
        
    except Exception as e:
        logging.error(f"Error getting real mail queue info: {str(e)}")
        return {
            "success": False,
            "error": str(e),
            "method": "error"
        }

def get_enhanced_queue_estimate(server):
    """تقدير محسن لطابور البريد بناءً على عوامل متعددة"""
    try:
        accounts = list_accounts(server)
        if not accounts:
            return {"queue_count": 0, "confidence": "low"}
        
        # عوامل التقدير المحسنة
        total_accounts = len(accounts)
        suspended_accounts = sum(1 for acct in accounts if acct.get('suspended', 0) == 1)
        active_accounts = total_accounts - suspended_accounts
        
        # حساب الحسابات عالية النشاط
        high_activity_accounts = 0
        total_email_accounts = 0
        
        for acct in accounts:
            try:
                # عد حسابات الإيميل لكل حساب
                emails = list_email_accounts(server, acct.get('user', ''))
                email_count = len(emails) if emails else 0
                total_email_accounts += email_count
                
                # تصنيف الحسابات عالية النشاط
                if email_count > 20 or float(acct.get('diskused', 0)) > 500:
                    high_activity_accounts += 1
                    
            except Exception as acct_error:
                logging.warning(f"Error processing account {acct.get('user', '')}: {str(acct_error)}")
        
        # حساب تقدير الطابور المحسن
        base_queue = active_accounts * 0.3  # أساس للحسابات النشطة
        suspended_queue = suspended_accounts * 5  # الحسابات المعلقة تحتفظ برسائل أكثر
        activity_queue = high_activity_accounts * 2  # الحسابات عالية النشاط
        email_based_queue = (total_email_accounts * 0.1)  # بناءً على عدد الإيميلات
        
        # التقدير النهائي
        estimated_queue = int(base_queue + suspended_queue + activity_queue + email_based_queue)
        
        # تحديد مستوى الثقة
        if total_accounts < 10:
            confidence = "high"
        elif total_accounts < 50:
            confidence = "medium"
        else:
            confidence = "low"
        
        return {
            "queue_count": estimated_queue,
            "confidence": confidence,
            "factors": {
                "total_accounts": total_accounts,
                "active_accounts": active_accounts,
                "suspended_accounts": suspended_accounts,
                "high_activity_accounts": high_activity_accounts,
                "total_email_accounts": total_email_accounts,
                "base_queue": base_queue,
                "suspended_queue": suspended_queue,
                "activity_queue": activity_queue,
                "email_based_queue": email_based_queue
            }
        }
        
    except Exception as e:
        logging.error(f"Error in enhanced queue estimate: {str(e)}")
        return {
            "queue_count": 0,
            "confidence": "error",
            "error": str(e)
        }

def generate_comprehensive_domain_report(server, domain, server_name):
    """إنشاء تقرير شامل ودقيق للدومين"""
    print(f"\n📊 COMPREHENSIVE DOMAIN ANALYSIS: {domain}")
    print("=" * 100)
    
    # جلب المعلومات المفصلة
    domain_data = get_detailed_domain_info(server, domain, server_name)
    
    if not domain_data.get("success"):
        print(f"❌ Error: {domain_data.get('error')}")
        return
    
    info = domain_data["data"]
    
    # عرض المعلومات الأساسية
    print(f"📋 BASIC INFORMATION")
    print("-" * 50)
    print(f"🌐 Domain: {info['domain']}")
    print(f"👤 cPanel User: {info['user']}")
    print(f"📧 Contact Email: {info['email_address']}")
    print(f"📦 Package: {info['package']}")
    print(f"🖥️  Server: {info['server_name']} ({info['server_ip']})")
    print(f"📅 Created: {info['creation_date']}")
    print(f"🕒 Account Age: {info['account_age_days']} days")
    print(f"🔴 Suspended: {'Yes' if info['suspended'] else 'No'}")
    
    # معلومات القرص والموارد
    print(f"\n💾 DISK & RESOURCE USAGE")
    print("-" * 50)
    print(f"💾 Disk Used: {info['disk_used_mb']:.2f}MB")
    if info['disk_limit_mb'] > 0:
        disk_usage_percent = (info['disk_used_mb'] / info['disk_limit_mb']) * 100
        print(f"📊 Disk Limit: {info['disk_limit_mb']:.2f}MB")
        print(f"📈 Disk Usage: {disk_usage_percent:.2f}%")
    else:
        print(f"📊 Disk Limit: Unlimited")
        print(f"📈 Disk Usage: N/A")
    
    if info.get('bandwidth_used') != "Unknown":
        print(f"📡 Bandwidth Used: {info['bandwidth_used']}")
    if info.get('inodes_used') != "Unknown":
        print(f"📁 Inodes Used: {info['inodes_used']}")
    
    # معلومات DNS
    print(f"\n🌐 DNS & CONNECTIVITY")
    print("-" * 50)
    if info['dns_resolves'] == True:
        print(f"✅ DNS Resolution: Working")
    elif info['dns_resolves'] == False:
        print(f"❌ DNS Resolution: Failed")
    else:
        print(f"❓ DNS Resolution: Unknown")
    
    if info['mx_records'] != "Unknown":
        print(f"📮 MX Records: {info['mx_records']} found")
    else:
        print(f"📮 MX Records: Unable to check")
    
    # تحليل الإيميلات المفصل
    print(f"\n📧 EMAIL ACCOUNTS ANALYSIS")
    print("-" * 50)
    print(f"📊 Total Email Accounts: {info['email_count']}")
    print(f"🟢 Active: {info['active_emails']}")
    print(f"🔴 Suspended: {info['suspended_emails']}")
    print(f"💾 Total Email Usage: {info['total_email_usage_mb']:.2f}MB")
    if info['total_email_quota_mb'] > 0:
        print(f"📊 Total Email Quota: {info['total_email_quota_mb']:.2f}MB")
        email_usage_percent = (info['total_email_usage_mb'] / info['total_email_quota_mb']) * 100
        print(f"📈 Email Usage: {email_usage_percent:.2f}%")
    else:
        print(f"📊 Total Email Quota: Unlimited")
    
    # تفاصيل كل إيميل
    if info['email_accounts']:
        print(f"\n📋 EMAIL ACCOUNTS DETAILS")
        print("-" * 100)
        print(f"{'#':<3} {'Email Address':<35} {'Status':<12} {'Used (MB)':<12} {'Quota':<15} {'Usage %':<10}")
        print("-" * 100)
        
        for i, email in enumerate(info['email_accounts'], 1):
            status = "🔴 Suspended" if email['suspended'] else "🟢 Active"
            print(f"{i:<3} {email['address']:<35} {status:<12} {email['used_mb']:<12.2f} "
                  f"{email['quota_display']:<15} {email['usage_percent']:<10}")
    
    # تحليل طابور البريد الحقيقي
    print(f"\n📮 MAIL QUEUE ANALYSIS")
    print("-" * 50)
    
    queue_info = get_real_mail_queue_info(server)
    
    print(f"🔧 Analysis Method: {queue_info.get('method', 'unknown')}")
    print(f"📊 Queue Messages: {queue_info.get('queue_count', 0)}")
    
    if queue_info.get('confidence'):
        print(f"🎯 Confidence Level: {queue_info['confidence']}")
    
    if queue_info.get('attempts'):
        print(f"\n🔍 Analysis Attempts:")
        for attempt in queue_info['attempts']:
            status_icon = "✅" if attempt['status'] == 'success' else "❌" if attempt['status'] == 'failed' else "⚠️"
            print(f"   {status_icon} {attempt['description']}: {attempt['status']}")
    
    # عوامل التقدير إذا كان متاحاً
    if queue_info.get('factors'):
        factors = queue_info['factors']
        print(f"\n📊 Queue Estimation Factors:")
        print(f"   📈 Base calculation: {factors.get('base_queue', 0):.1f}")
        print(f"   🔴 Suspended impact: {factors.get('suspended_queue', 0):.1f}")
        print(f"   ⚡ Activity impact: {factors.get('activity_queue', 0):.1f}")
        print(f"   📧 Email-based impact: {factors.get('email_based_queue', 0):.1f}")
    
    # تحليل المخاطر المحسن
    risk_analysis = analyze_domain_risk_factors(info)
    
    print(f"\n⚠️  RISK ASSESSMENT")
    print("-" * 50)
    print(f"🚨 Risk Score: {risk_analysis['risk_score']}/30")
    print(f"📊 Risk Level: {risk_analysis['risk_level']}")
    
    if risk_analysis['risk_factors']:
        print(f"⚠️  Risk Factors:")
        for factor in risk_analysis['risk_factors']:
            print(f"   • {factor}")
    
    print(f"📊 Estimated Failures: {risk_analysis['estimated_failures']}")
    
    # توصيات مفصلة
    print(f"\n💡 DETAILED RECOMMENDATIONS")
    print("-" * 50)
    
    recommendations = generate_detailed_recommendations(info, queue_info, risk_analysis)
    for i, rec in enumerate(recommendations, 1):
        print(f"{i}. {rec}")
    
    # ملخص التقرير
    print(f"\n📊 REPORT SUMMARY")
    print("=" * 50)
    print(f"Domain: {domain}")
    print(f"Overall Health: {risk_analysis['health_status']}")
    print(f"Action Priority: {risk_analysis['priority_level']}")
    print(f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    return {
        "domain_info": info,
        "queue_info": queue_info,
        "risk_analysis": risk_analysis,
        "recommendations": recommendations
    }

def analyze_domain_risk_factors(domain_info):
    """تحليل عوامل المخاطر للدومين"""
    risk_score = 0
    risk_factors = []
    
    # 1. حالة التعليق
    if domain_info['suspended']:
        risk_score += 10
        risk_factors.append("Domain is suspended (critical)")
    
    # 2. عمر الحساب
    if domain_info['account_age_days'] < 7:
        risk_score += 5
        risk_factors.append(f"Recently created ({domain_info['account_age_days']} days)")
    elif domain_info['account_age_days'] < 30:
        risk_score += 3
        risk_factors.append(f"New account ({domain_info['account_age_days']} days)")
    
    # 3. استخدام القرص
    if domain_info['disk_used_mb'] > 2000:  # أكثر من 2GB
        risk_score += 4
        risk_factors.append(f"High disk usage ({domain_info['disk_used_mb']:.0f}MB)")
    elif domain_info['disk_used_mb'] > 1000:  # أكثر من 1GB
        risk_score += 2
        risk_factors.append(f"Moderate disk usage ({domain_info['disk_used_mb']:.0f}MB)")
    
    # 4. عدد حسابات الإيميل
    if domain_info['email_count'] > 100:
        risk_score += 6
        risk_factors.append(f"Very high email count ({domain_info['email_count']})")
    elif domain_info['email_count'] > 50:
        risk_score += 4
        risk_factors.append(f"High email count ({domain_info['email_count']})")
    elif domain_info['email_count'] > 20:
        risk_score += 2
        risk_factors.append(f"Moderate email count ({domain_info['email_count']})")
    
    # 5. استخدام الإيميل
    if domain_info['total_email_usage_mb'] > 1000:
        risk_score += 3
        risk_factors.append(f"High email storage usage ({domain_info['total_email_usage_mb']:.0f}MB)")
    
    # 6. أنماط الدومين المشبوهة
    domain_lower = domain_info['domain'].lower()
    suspicious_patterns = ['temp', 'test', 'spam', 'bulk', 'mail', 'send', 'newsletter', 'promo', 'offer']
    suspicious_found = [pattern for pattern in suspicious_patterns if pattern in domain_lower]
    if suspicious_found:
        risk_score += 5
        risk_factors.append(f"Suspicious domain patterns: {', '.join(suspicious_found)}")
    
    # 7. مشاكل DNS
    if domain_info['dns_resolves'] == False:
        risk_score += 3
        risk_factors.append("DNS resolution issues")
    
    # تحديد مستوى المخاطر
    if risk_score >= 20:
        risk_level = "🔴 CRITICAL"
        health_status = "Poor"
        priority_level = "IMMEDIATE"
    elif risk_score >= 15:
        risk_level = "🟠 HIGH"
        health_status = "Fair"
        priority_level = "URGENT"
    elif risk_score >= 10:
        risk_level = "🟡 MODERATE"
        health_status = "Good"
        priority_level = "MONITOR"
    elif risk_score >= 5:
        risk_level = "🟢 LOW"
        health_status = "Very Good"
        priority_level = "ROUTINE"
    else:
        risk_level = "✅ MINIMAL"
        health_status = "Excellent"
        priority_level = "STANDARD"
    
    # حساب الإيميلات الفاشلة المقدرة
    estimated_failures = int(risk_score * 1.5) if risk_score > 0 else 0
    
    return {
        "risk_score": risk_score,
        "risk_level": risk_level,
        "health_status": health_status,
        "priority_level": priority_level,
        "risk_factors": risk_factors,
        "estimated_failures": estimated_failures
    }

def show_failed_login_logs(server, lines=50):
    """عرض لوجات فشل تسجيل الدخول من Exim"""
    try:
        print(f"\n🚨 FAILED LOGIN ATTEMPTS ANALYSIS")
        print("=" * 80)
        print(f"📊 Showing last {lines} failed login attempts")
        print(f"🖥️  Server: {server['ip']}")
        print("=" * 80)
        
        # محاولة جلب اللوجات من السيرفر
        try:
            # محاولة استخدام طرق متعددة لجلب اللوجات
            logs = []
            
            # الطريقة الأولى: محاولة استخدام exec API
            try:
                log_command = f"grep 'dovecot_login authenticator failed' /var/log/exim_mainlog | tail -n {lines}"
                result = whm_api_call(server, "exec", {"command": log_command})
                
                if "error" not in result and "data" in result:
                    logs = result["data"].strip().split('\n')
                    print(f"✅ Retrieved logs using exec API")
                    
            except Exception as exec_error:
                print(f"⚠️  Exec API failed: {str(exec_error)}")
                
                # الطريقة الثانية: محاولة استخدام shell API
                try:
                    result = whm_api_call(server, "shell", {"command": log_command})
                    
                    if "error" not in result and "data" in result:
                        logs = result["data"].strip().split('\n')
                        print(f"✅ Retrieved logs using shell API")
                        
                except Exception as shell_error:
                    print(f"⚠️  Shell API failed: {str(shell_error)}")
                    
                    # الطريقة الثالثة: محاولة استخدام ps API مع grep
                    try:
                        # استخدام ps API للتحقق من وجود exim_mainlog
                        ps_result = whm_api_call(server, "ps", {"pattern": "exim"})
                        if "error" not in ps_result:
                            print(f"✅ Exim process found, but cannot access logs directly")
                            
                    except Exception as ps_error:
                        print(f"⚠️  PS API failed: {str(ps_error)}")
            
            # إذا فشلت جميع الطرق، اعرض رسالة واضحة مع تعليمات
            if not logs:
                print(f"⚠️  Cannot access server logs directly due to WHM API v1 limitations")
                print(f"💡 To view real failed login logs, run these commands on the server:")
                print(f"   🔍 Check Dovecot failed logins:")
                print(f"      grep 'dovecot_login authenticator failed' /var/log/exim_mainlog | tail -n {lines}")
                print(f"   🔍 Check general authentication failures:")
                print(f"      tail -n {lines} /var/log/exim_mainlog | grep 'authentication failed'")
                print(f"   🔍 Check mail logs:")
                print(f"      tail -n {lines} /var/log/maillog | grep -i 'failed\\|denied\\|rejected'")
                print(f"   🔍 Check system logs:")
                print(f"      journalctl -u dovecot --since '1 hour ago' | grep -i 'failed\\|denied'")
                print(f"\n📊 Alternative: Use SSH to connect to server and run commands directly")
                print(f"   ssh root@{server['ip']}")
                print(f"   # Then run the grep commands above")
                
                return {"success": False, "error": "Real logs not accessible via API - use SSH commands above"}
                
        except Exception as log_error:
            print(f"❌ Error accessing logs: {str(log_error)}")
            print("💡 Try running manually on server:")
            print(f"   grep 'dovecot_login authenticator failed' /var/log/exim_mainlog | tail -n {lines}")
            return {"success": False, "error": str(log_error)}
            
    except Exception as e:
        print(f"❌ Error in failed login analysis: {str(e)}")
        logging.error(f"Error in failed login analysis: {str(e)}")
        return {"success": False, "error": str(e)}

def parse_failed_login_log(log_line, failed_attempts_count=0):
    """تحليل سطر لوج فشل تسجيل الدخول"""
    try:
        # مثال: 2025-08-23 20:35:50 [81.30.107.33]:10404: 535 Incorrect authentication data (set_id=jms)
        
        # استخراج IP
        ip_match = re.search(r'\[(\d+\.\d+\.\d+\.\d+)\]', log_line)
        ip = ip_match.group(1) if ip_match else "Unknown"
        
        # استخراج البورت
        port_match = re.search(r'\]:(\d+):', log_line)
        port = port_match.group(1) if port_match else "Unknown"
        
        # استخراج الإيميل/اليوزر
        email_match = re.search(r'set_id=([^)]+)', log_line)
        email = email_match.group(1) if email_match else "Unknown"
        
        # استخراج الدومين
        domain = email.split('@')[1] if '@' in email else "Unknown"
        
        # استخراج الوقت
        timestamp_match = re.search(r'^(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})', log_line)
        timestamp = timestamp_match.group(1) if timestamp_match else "Unknown"
        
        # تحديد مستوى المخاطر بناءً على عدد المحاولات المرسل
        risk_level = "LOW"
        if ip.startswith(('81.30.107.', '78.153.140.')):  # IPs متكررة
            risk_level = "MEDIUM"
        if failed_attempts_count > 10:
            risk_level = "HIGH"
        if failed_attempts_count > 20:
            risk_level = "CRITICAL"
        
        # تحديد الموقع (تقريبي)
        location = get_ip_location(ip)
        
        return {
            "ip": ip,
            "port": port,
            "email": email,
            "domain": domain,
            "timestamp": timestamp,
            "risk_level": risk_level,
            "location": location
        }
        
    except Exception as e:
        logging.warning(f"Error parsing log line: {str(e)}")
        return None

def failed_login_analysis_menu(servers):
    """قائمة تحليل محاولات تسجيل الدخول الفاشلة"""
    try:
        print(f"\n🚨 FAILED LOGIN ANALYSIS MENU")
        print("=" * 60)
        
        # اختيار السيرفر
        online_servers = get_online_servers(servers)
        if not online_servers:
            return
        
        server_choice = input(f"\nChoose server ({'/'.join(online_servers.keys())}): ").strip()
        if server_choice not in online_servers:
            print("❌ Invalid server choice!")
            return
        
        server = online_servers[server_choice]
        print(f"\n🖥️  Selected Server: {server_choice} ({server['ip']})")
        
        while True:
            print(f"\n{'='*20} FAILED LOGIN ANALYSIS {'='*20}")
            print("1. 📊 Show last 50 failed attempts")
            print("2. 📊 Show last 100 failed attempts")
            print("3. 📊 Show last 200 failed attempts")
            print("4. 🎯 Custom number of lines")
            print("5. 📈 Export analysis to file")
            print("0. 🔙 Back to main menu")
            print("=" * 60)
            
            choice = input("Choose option: ").strip()
            
            if choice == "1":
                show_failed_login_logs(server, 50)
            elif choice == "2":
                show_failed_login_logs(server, 100)
            elif choice == "3":
                show_failed_login_logs(server, 200)
            elif choice == "4":
                try:
                    lines = input("📊 Enter number of lines to show: ").strip()
                    lines = int(lines) if lines.isdigit() else 50
                    if lines < 1 or lines > 1000:
                        print("❌ Number must be between 1 and 1000")
                        lines = 50
                    show_failed_login_logs(server, lines)
                except ValueError:
                    print("❌ Invalid number")
            elif choice == "5":
                export_failed_login_analysis(server)
            elif choice == "0":
                break
            else:
                print("❌ Invalid option")
                
    except Exception as e:
        print(f"❌ Error in failed login analysis menu: {str(e)}")
        logging.error(f"Error in failed login analysis menu: {str(e)}")

def export_failed_login_analysis(server):
    """تصدير تحليل محاولات تسجيل الدخول الفاشلة إلى ملف"""
    try:
        print(f"\n📁 EXPORTING FAILED LOGIN ANALYSIS")
        print("=" * 50)
        
        # جلب اللوجات
        result = show_failed_login_logs(server, 200)  # آخر 200 محاولة
        
        if result.get("success"):
            # إنشاء اسم الملف
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"failed_login_analysis_{server['ip']}_{timestamp}.txt"
            filepath = os.path.join("reports", filename)
            
            # إنشاء مجلد التقارير إذا لم يكن موجوداً
            os.makedirs("reports", exist_ok=True)
            
            # كتابة التقرير
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write("FAILED LOGIN ATTEMPTS ANALYSIS REPORT\n")
                f.write("=" * 50 + "\n")
                f.write(f"Server: {server['ip']}\n")
                f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Total Attempts: {result['total_attempts']}\n")
                f.write(f"Unique IPs: {result['unique_ips']}\n")
                f.write(f"Unique Domains: {result['unique_domains']}\n\n")
                
                if result.get('suspicious_ips'):
                    f.write("SUSPICIOUS IPs (>5 attempts):\n")
                    f.write("-" * 30 + "\n")
                    for ip, count in result['suspicious_ips'].items():
                        f.write(f"{ip}: {count} attempts\n")
                    f.write("\n")
                
                f.write("DETAILED LOG ENTRIES:\n")
                f.write("-" * 30 + "\n")
                for log_line in result['logs']:
                    f.write(f"{log_line}\n")
            
            print(f"✅ Report exported to: {filepath}")
            
            # عرض إحصائيات سريعة
            print(f"\n📊 QUICK STATS:")
            print(f"   📁 File: {filename}")
            print(f"   📊 Size: {os.path.getsize(filepath)} bytes")
            print(f"   📅 Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            
        else:
            print(f"❌ Cannot export: {result.get('error', 'Unknown error')}")
            
    except Exception as e:
        print(f"❌ Error exporting report: {str(e)}")
        logging.error(f"Error exporting report: {str(e)}")



def get_ip_location(ip):
    """الحصول على موقع IP (تقريبي)"""
    try:
        # يمكن استخدام خدمة خارجية مثل ipinfo.io
        # هنا نستخدم تصنيف بسيط
        if ip.startswith('81.30.107.'):
            return "Egypt (Local)"
        elif ip.startswith('78.153.140.'):
            return "Egypt (Local)"
        elif ip.startswith('182.109.'):
            return "China"
        elif ip.startswith('121.26.'):
            return "China"
        elif ip.startswith('116.114.'):
            return "China"
        elif ip.startswith('76.88.'):
            return "USA"
        elif ip.startswith('88.87.'):
            return "Russia"
        elif ip.startswith('221.12.'):
            return "China"
        else:
            return "Unknown"
    except:
        return "Unknown"

def generate_detailed_recommendations(domain_info, queue_info, risk_analysis):
    """توليد توصيات مفصلة بناءً على التحليل"""
    recommendations = []
    
    # توصيات بناءً على مستوى المخاطر
    if risk_analysis['priority_level'] == "IMMEDIATE":
        recommendations.append("🚨 SUSPEND account immediately to prevent further damage")
        recommendations.append("🔍 Investigate all email accounts for malicious activity")
        recommendations.append("🧹 Clean up any spam or malicious content")
    elif risk_analysis['priority_level'] == "URGENT":
        recommendations.append("⚠️ Review account within 2 hours")
        recommendations.append("🔒 Consider temporary restrictions on email sending")
        recommendations.append("📊 Monitor closely for next 24-48 hours")
    
    # توصيات للإيميل
    if domain_info['email_count'] > 50:
        recommendations.append(f"📧 Review the necessity of {domain_info['email_count']} email accounts")
        recommendations.append("🔍 Check for unused or dormant email accounts")
    
    # توصيات للقرص
    if domain_info['disk_used_mb'] > 1000:
        recommendations.append("💾 Investigate high disk usage and clean up unnecessary files")
        recommendations.append("📁 Check for log files or temporary files taking up space")
    
    # توصيات لطابور البريد
    queue_count = queue_info.get('queue_count', 0)
    if queue_count > 100:
        recommendations.append("📮 Clear mail queue immediately - investigate stuck messages")
        recommendations.append("🔧 Check Exim configuration and DNS settings")
    elif queue_count > 50:
        recommendations.append("📮 Monitor mail queue closely - may indicate delivery issues")
    
    # توصيات DNS
    if domain_info['dns_resolves'] == False:
        recommendations.append("🌐 Fix DNS configuration - this will cause email delivery failures")
        recommendations.append("📮 Update MX records if necessary")
    
    # توصيات عامة
    if domain_info['account_age_days'] < 30:
        recommendations.append("👀 New account - monitor closely for first month")
    
    if not recommendations:
        recommendations.append("✅ Account appears healthy - continue standard monitoring")
        recommendations.append("📊 Perform quarterly review of email usage and patterns")
    
    return recommendations

def get_email_settings(email_details):
    """توليد إعدادات البريد الإلكتروني"""
    domain = email_details['domain']
    
    settings = {
        'outlook': {
            'incoming': {
                'server': f'mail.{domain}',
                'port': '993',
                'ssl': 'Yes',
                'auth': 'Yes',
                'protocol': 'IMAP'
            },
            'outgoing': {
                'server': f'mail.{domain}',
                'port': '465',
                'ssl': 'Yes',
                'auth': 'Yes',
                'protocol': 'SMTP'
            }
        },
        'iphone': {
            'incoming': {
                'server': f'mail.{domain}',
                'port': '993',
                'ssl': 'Yes',
                'auth': 'Yes',
                'protocol': 'IMAP'
            },
            'outgoing': {
                'server': f'mail.{domain}',
                'port': '465',
                'ssl': 'Yes',
                'auth': 'Yes',
                'protocol': 'SMTP'
            }
        },
        'webmail': {
            'url': f'https://webmail.{domain}',
            'ssl': 'Yes'
        }
    }
    
    return settings

def format_email_settings(email_address, settings):
    """تنسيق إعدادات البريد الإلكتروني للعرض"""
    output = []
    output.append(f"\n📧 Email Settings for: {email_address}")
    output.append("=" * 50)
    
    # Outlook Settings
    output.append("\n📨 Microsoft Outlook Settings:")
    output.append("-" * 30)
    output.append("Incoming Mail (IMAP):")
    output.append(f"   Server: {settings['outlook']['incoming']['server']}")
    output.append(f"   Port: {settings['outlook']['incoming']['port']}")
    output.append(f"   Security: {settings['outlook']['incoming']['ssl']}")
    output.append("\nOutgoing Mail (SMTP):")
    output.append(f"   Server: {settings['outlook']['outgoing']['server']}")
    output.append(f"   Port: {settings['outlook']['outgoing']['port']}")
    output.append(f"   Security: {settings['outlook']['outgoing']['ssl']}")
    output.append("\nAuthentication: Required for both incoming and outgoing mail")
    output.append(f"Username: {email_address}")
    output.append("Password: Your email account password")
    
    # iPhone/iPad Settings
    output.append("\n📱 iPhone/iPad Settings:")
    output.append("-" * 30)
    output.append("Incoming Mail Server:")
    output.append(f"   Host Name: {settings['iphone']['incoming']['server']}")
    output.append(f"   Port: {settings['iphone']['incoming']['port']}")
    output.append(f"   Security: {settings['iphone']['incoming']['ssl']}")
    output.append("\nOutgoing Mail Server:")
    output.append(f"   Host Name: {settings['iphone']['outgoing']['server']}")
    output.append(f"   Port: {settings['iphone']['outgoing']['port']}")
    output.append(f"   Security: {settings['iphone']['outgoing']['ssl']}")
    output.append("\nAuthentication: Required for both incoming and outgoing mail")
    output.append(f"Username: {email_address}")
    output.append("Password: Your email account password")
    
    # Webmail Access
    output.append("\n🌐 Webmail Access:")
    output.append("-" * 30)
    output.append(f"URL: {settings['webmail']['url']}")
    output.append(f"Username: {email_address}")
    output.append("Password: Your email account password")
    
    # Additional Notes
    output.append("\n📝 Important Notes:")
    output.append("-" * 30)
    output.append("1. Make sure to use SSL/TLS encryption")
    output.append("2. Username should always be your full email address")
    output.append("3. Authentication is required for both incoming and outgoing mail")
    output.append("4. If you can't connect, check your password and security settings")
    
    return "\n".join(output)

def show_email_settings(servers):
    """عرض إعدادات البريد الإلكتروني (Outlook/iPhone)"""
    print("\n⚙️ View Email Settings (Outlook/iPhone)")
    print("=" * 60)
    
    domain = input("🌐 Enter domain: ").strip()
    if not domain:
        print("❌ Domain cannot be empty")
        return
        
    # خيارات البحث
    print("\n🔍 Search Options:")
    print("1. 🚀 Fast search (main domains only)")
    print("2. 🧠 Smart search (main domains first, then subdomains if needed)")
    print("3. 🔍 Full search (all domains + subdomains)")
    
    search_choice = input("Choose search type (1-3, default 2): ").strip()
    
    if search_choice == "1":
        search_mode = "fast"
    elif search_choice == "3":
        search_mode = "full"
    else:
        search_mode = "smart"
    
    print(f"\n🔍 Searching for domain: {domain}...")
    
    # اختيار دالة البحث المناسبة بناءً على نوع البحث
    if search_mode == "fast":
        print("🚀 Using fast search (main domains only)...")
        server, acct, server_name = find_server_by_domain_fast(domain, servers)
    elif search_mode == "full":
        print("🧠 Using full search (all domains + subdomains)...")
        server, acct, server_name = find_server_by_domain_full(domain, servers)
    else:  # smart mode
        print("🧠 Using smart search (main domains first, then subdomains if needed)...")
        server, acct, server_name = find_server_by_domain_smart(domain, servers)
    
    if not server:
        print("❌ Domain not found on any server!")
        return
        
    cpanel_user = acct["user"]
    print(f"\n✅ Domain found on Server {server_name}")
    
    # جلب قائمة الإيميلات
    print(f"📧 Loading email accounts...")
    emails = list_email_accounts(server, cpanel_user, domain)
    
    if not emails:
        print("❌ No email accounts found")
        return
    
    print(f"\n📋 Available Email Accounts ({len(emails)} found):")
    for i, email in enumerate(emails, 1):
        email_address = email.get("email", "Unknown")
        print(f"{i}. {email_address}")
    
    email_index = input(f"\nEnter email number (1-{len(emails)}): ").strip()
    try:
        index = int(email_index) - 1
        if 0 <= index < len(emails):
            email = emails[index]
            settings = get_email_settings(email)
            formatted_settings = format_email_settings(email['email'], settings)
            
            print(formatted_settings)
            
            if confirm_action("\nDo you want to export these settings to a text file?"):
                filename = f"reports/email_settings_{email['email'].replace('@', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
                try:
                    # إنشاء مجلد reports إذا لم يكن موجوداً
                    os.makedirs('reports', exist_ok=True)
                    
                    with open(filename, 'w', encoding='utf-8') as f:
                        f.write(formatted_settings)
                    print(f"✅ Settings exported to: {filename}")
                except Exception as e:
                    print(f"❌ Error exporting settings: {str(e)}")
        else:
            print("❌ Invalid email number")
    except ValueError:
        print("❌ Invalid input")

# === دوال إدارة Forward Rules ===
def manage_email_forward_rules(servers):
    """إدارة قواعد إعادة توجيه الإيميلات"""
    print("\n📤 Manage Email Forward Rules")
    print("=" * 50)
    
    domain = input("🌐 Enter domain: ").strip()
    if not domain:
        print("❌ Domain cannot be empty")
        return
        
    # خيارات البحث
    print("\n🔍 Search Options:")
    print("1. 🚀 Fast search (main domains only)")
    print("2. 🧠 Smart search (main domains first, then subdomains if needed)")
    print("3. 🔍 Full search (all domains + subdomains)")
    
    search_choice = input("Choose search type (1-3, default 2): ").strip()
    
    if search_choice == "1":
        search_mode = "fast"
    elif search_choice == "3":
        search_mode = "full"
    else:
        search_mode = "smart"
    
    print(f"\n🔍 Searching for domain: {domain}...")
    
    # اختيار دالة البحث المناسبة بناءً على نوع البحث
    if search_mode == "fast":
        print("🚀 Using fast search (main domains only)...")
        server, acct, server_name = find_server_by_domain_fast(domain, servers)
    elif search_mode == "full":
        print("🧠 Using full search (all domains + subdomains)...")
        server, acct, server_name = find_server_by_domain_full(domain, servers)
    else:  # smart mode
        print("🧠 Using smart search (main domains first, then subdomains if needed)...")
        server, acct, server_name = find_server_by_domain_smart(domain, servers)
    
    if not server:
        print("❌ Domain not found on any server!")
        return
        
    cpanel_user = acct["user"]
    print(f"\n✅ Domain found on Server {server_name}")
    
    while True:
        print(f"\n📤 Forward Rules Management for {domain}")
        print("=" * 50)
        print("1. 📋 List all forward rules")
        print("2. ➕ Add new forward rule")
        print("3. ✏️  Edit forward rule")
        print("4. 🗑️  Delete forward rule")
        print("5. 📊 Export forward rules report")
        print("6. 🔍 Search forward rules")
        print("7. 🧪 Test forward rules loading")
        print("0. 🔙 Back")
        
        choice = input("Choose option: ").strip()
        
        if choice == "1":
            list_forward_rules(server, cpanel_user, domain)
        elif choice == "2":
            add_forward_rule(server, cpanel_user, domain)
        elif choice == "3":
            edit_forward_rule(server, cpanel_user, domain)
        elif choice == "4":
            delete_forward_rule(server, cpanel_user, domain)
        elif choice == "5":
            export_forward_rules_report(server, cpanel_user, domain)
        elif choice == "6":
            search_forward_rules(server, cpanel_user, domain)
        elif choice == "7":
            test_forward_rules_loading(server, cpanel_user, domain)
        elif choice == "0":
            break
        else:
            print("❌ Invalid option")

def list_forward_rules(server, cpanel_user, domain):
    """عرض قائمة قواعد إعادة التوجيه"""
    print(f"\n📋 Forward Rules for {domain}")
    print("=" * 50)
    
    try:
        # جلب قواعد إعادة التوجيه
        result = cpanel_api_call(server, cpanel_user, "Email", "list_forwards", {"domain": domain})
        
        if "error" in result:
            print(f"❌ Error fetching forward rules: {result['error']}")
            return
            
        if "result" in result and "data" in result["result"]:
            forwards = result["result"]["data"]
            
            if not forwards:
                print("ℹ️  No forward rules found for this domain")
                return
                
            print(f"Found {len(forwards)} forward rule(s):")
            print("-" * 50)
            
            for i, forward in enumerate(forwards, 1):
                email = forward.get("email", "Unknown")
                forward_to = forward.get("forward", "Unknown")
                status = "✅ Active" if forward.get("active", 1) else "❌ Inactive"
                
                print(f"{i}. {email} → {forward_to} ({status})")
                
        else:
            print("ℹ️  No forward rules found")
            
    except Exception as e:
        print(f"❌ Error: {str(e)}")

def add_forward_rule(server, cpanel_user, domain):
    """إضافة قاعدة إعادة توجيه جديدة"""
    print(f"\n➕ Add New Forward Rule for {domain}")
    print("=" * 50)
    
    email = input("📧 Enter email address (e.g., info@domain.com): ").strip()
    if not email:
        print("❌ Email address cannot be empty")
        return
        
    # التحقق من صحة الإيميل
    if "@" not in email or not email.endswith(f"@{domain}"):
        print(f"❌ Email must be a valid address from {domain}")
        return
        
    forward_to = input("📤 Forward to (email address): ").strip()
    if not forward_to or "@" not in forward_to:
        print("❌ Forward address must be a valid email")
        return
        
    # خيارات إضافية
    print("\n🔧 Additional Options:")
    print("1. Keep copy in original mailbox")
    print("2. Delete from original mailbox")
    
    copy_choice = input("Choose option (1-2): ").strip()
    keep_copy = copy_choice == "1"
    
    try:
        params = {
            "email": email,
            "forward": forward_to,
            "domain": domain
        }
        
        if not keep_copy:
            params["delete"] = 1
        
        result = cpanel_api_call(server, cpanel_user, "Email", "add_forward", params)
        
        if "error" in result:
            print(f"❌ Error adding forward rule: {result['error']}")
        else:
            print("✅ Forward rule added successfully!")
            print(f"📧 {email} → {forward_to}")
            if keep_copy:
                print("📁 Copy will be kept in original mailbox")
            else:
                print("🗑️  Emails will be deleted from original mailbox")
                
    except Exception as e:
        print(f"❌ Error: {str(e)}")

def edit_forward_rule(server, cpanel_user, domain):
    """تعديل قاعدة إعادة توجيه موجودة"""
    print(f"\n✏️  Edit Forward Rule for {domain}")
    print("=" * 50)
    
    # عرض القواعد الموجودة أولاً
    list_forward_rules(server, cpanel_user, domain)
    
    try:
        result = cpanel_api_call(server, cpanel_user, "Email", "list_forwards", {"domain": domain})
        
        if "error" in result or "result" not in result or "data" not in result["result"]:
            print("❌ No forward rules to edit")
            return
            
        forwards = result["result"]["data"]
        if not forwards:
            print("❌ No forward rules to edit")
            return
            
        rule_index = input(f"\nEnter rule number to edit (1-{len(forwards)}): ").strip()
        try:
            index = int(rule_index) - 1
            if 0 <= index < len(forwards):
                forward = forwards[index]
                email = forward.get("email", "")
                current_forward = forward.get("forward", "")
                
                print(f"\n📧 Editing: {email} → {current_forward}")
                print("=" * 40)
                
                new_forward = input(f"New forward address (current: {current_forward}): ").strip()
                if not new_forward:
                    new_forward = current_forward
                    
                if "@" not in new_forward:
                    print("❌ Invalid email address")
                    return
                    
                # حذف القاعدة القديمة وإضافة الجديدة
                delete_result = cpanel_api_call(server, cpanel_user, "Email", "delete_forward", {
                    "email": email,
                    "domain": domain
                })
                
                if "error" in delete_result:
                    print(f"❌ Error deleting old rule: {delete_result['error']}")
                    return
                    
                # إضافة القاعدة الجديدة
                add_result = cpanel_api_call(server, cpanel_user, "Email", "add_forward", {
                    "email": email,
                    "forward": new_forward,
                    "domain": domain
                })
                
                if "error" in add_result:
                    print(f"❌ Error adding new rule: {add_result['error']}")
                else:
                    print("✅ Forward rule updated successfully!")
                    print(f"📧 {email} → {new_forward}")
            else:
                print("❌ Invalid rule number")
        except ValueError:
            print("❌ Invalid input")
            
    except Exception as e:
        print(f"❌ Error: {str(e)}")

def delete_forward_rule(server, cpanel_user, domain):
    """حذف قاعدة إعادة توجيه"""
    print(f"\n🗑️  Delete Forward Rule for {domain}")
    print("=" * 50)
    
    # عرض القواعد الموجودة أولاً
    list_forward_rules(server, cpanel_user, domain)
    
    try:
        result = cpanel_api_call(server, cpanel_user, "Email", "list_forwards", {"domain": domain})
        
        if "error" in result or "result" not in result or "data" not in result["result"]:
            print("❌ No forward rules to delete")
            return
            
        forwards = result["result"]["data"]
        if not forwards:
            print("❌ No forward rules to delete")
            return
            
        rule_index = input(f"\nEnter rule number to delete (1-{len(forwards)}): ").strip()
        try:
            index = int(rule_index) - 1
            if 0 <= index < len(forwards):
                forward = forwards[index]
                email = forward.get("email", "")
                forward_to = forward.get("forward", "")
                
                if confirm_action(f"Delete forward rule: {email} → {forward_to}?"):
                    delete_result = cpanel_api_call(server, cpanel_user, "Email", "delete_forward", {
                        "email": email,
                        "domain": domain
                    })
                    
                    if "error" in delete_result:
                        print(f"❌ Error deleting forward rule: {delete_result['error']}")
                    else:
                        print("✅ Forward rule deleted successfully!")
                        print(f"📧 {email} → {forward_to}")
            else:
                print("❌ Invalid rule number")
        except ValueError:
            print("❌ Invalid input")
            
    except Exception as e:
        print(f"❌ Error: {str(e)}")

def search_forward_rules(server, cpanel_user, domain):
    """البحث في قواعد إعادة التوجيه"""
    print(f"\n🔍 Search Forward Rules for {domain}")
    print("=" * 50)
    
    search_term = input("Enter search term (email or forward address): ").strip().lower()
    if not search_term:
        print("❌ Search term cannot be empty")
        return
        
    try:
        result = cpanel_api_call(server, cpanel_user, "Email", "list_forwards", {"domain": domain})
        
        if "error" in result:
            print(f"❌ Error fetching forward rules: {result['error']}")
            return
            
        if "result" in result and "data" in result["result"]:
            forwards = result["result"]["data"]
            
            if not forwards:
                print("ℹ️  No forward rules found")
                return
                
            # البحث في القواعد
            matching_forwards = []
            for forward in forwards:
                email = forward.get("email", "").lower()
                forward_to = forward.get("forward", "").lower()
                
                if search_term in email or search_term in forward_to:
                    matching_forwards.append(forward)
            
            if not matching_forwards:
                print(f"❌ No forward rules found matching '{search_term}'")
                return
                
            print(f"\n✅ Found {len(matching_forwards)} matching forward rule(s):")
            print("-" * 50)
            
            for i, forward in enumerate(matching_forwards, 1):
                email = forward.get("email", "Unknown")
                forward_to = forward.get("forward", "Unknown")
                status = "✅ Active" if forward.get("active", 1) else "❌ Inactive"
                
                print(f"{i}. {email} → {forward_to} ({status})")
                
        else:
            print("ℹ️  No forward rules found")
            
    except Exception as e:
        print(f"❌ Error: {str(e)}")

def export_forward_rules_report(server, cpanel_user, domain):
    """تصدير تقرير قواعد إعادة التوجيه"""
    print(f"\n📊 Export Forward Rules Report for {domain}")
    print("=" * 50)
    
    try:
        # جلب قواعد إعادة التوجيه
        result = cpanel_api_call(server, cpanel_user, "Email", "list_forwards", {"domain": domain})
        
        if "error" in result:
            print(f"❌ Error fetching forward rules: {result['error']}")
            return
            
        if "result" in result and "data" in result["result"]:
            forwards = result["result"]["data"]
            
            if not forwards:
                print("ℹ️  No forward rules found to export")
                return
                
            # عرض Forward Rules المحملة للتأكيد
            print(f"📋 Found {len(forwards)} forward rule(s):")
            for i, forward in enumerate(forwards, 1):
                email = forward.get("email", "Unknown")
                forward_to = forward.get("forward", "Unknown")
                status = "✅ Active" if forward.get("active", 1) else "❌ Inactive"
                print(f"{i}. {email} → {forward_to} ({status})")
            
            # إنشاء اسم الملف
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"forward_rules_{domain}_{timestamp}.txt"
            
            with open(filename, "w", encoding="utf-8") as f:
                f.write(f"Forward Rules Report for {domain}\n")
                f.write("=" * 50 + "\n")
                f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Server: {server['ip']}\n")
                f.write(f"cPanel User: {cpanel_user}\n")
                f.write(f"Total Forward Rules: {len(forwards)}\n")
                f.write("=" * 50 + "\n\n")
                
                f.write(f"{'#':<3} {'Email Address':<35} {'Forward To':<35} {'Status':<10}\n")
                f.write("-" * 85 + "\n")
                
                for i, forward in enumerate(forwards, 1):
                    email = forward.get("email", "Unknown")
                    forward_to = forward.get("forward", "Unknown")
                    status = "Active" if forward.get("active", 1) else "Inactive"
                    
                    f.write(f"{i:<3} {email:<35} {forward_to:<35} {status:<10}\n")
                
                f.write("\n" + "=" * 50 + "\n")
                f.write("Report generated by WHM Email Management Script\n")
            
            print(f"\n✅ Report exported successfully!")
            print(f"📁 File: {filename}")
            print(f"📊 Total forward rules: {len(forwards)}")
            
        else:
            print("ℹ️  No forward rules found to export")
            
    except Exception as e:
        print(f"❌ Error: {str(e)}")

def test_forward_rules_loading(server, cpanel_user, domain):
    """اختبار تحميل Forward Rules للتشخيص"""
    print(f"\n🧪 Testing Forward Rules Loading for {domain}")
    print("=" * 50)
    
    # اختبار جميع الطرق المتاحة
    print("🔍 Testing all available methods...")
    
    # 1. اختبار cPanel API methods
    cpanel_methods = [
        ("Email", "list_forwards", {"domain": domain}),
        ("Email", "list_forwards", {}),
        ("Email", "list_forwards", {"user": cpanel_user}),
        ("Email", "list_forwards", {"cpanel_jsonapi_user": cpanel_user}),
        ("Email", "list_forwards", {"email": "*"}),
        ("Email", "list_forwards", {"email": f"*@{domain}"}),
        ("Email", "list_forwards", {"forward": "*"}),
        ("Email", "list_forwards", {"type": "forward"}),
        ("Email", "list_forwards", {"type": "all"}),
        ("Email", "list_forwards", {"include": "forward"}),
        ("Email", "list_forwards", {"include": "all"}),
        ("Email", "list_forwards", {"format": "json"}),
        ("Email", "list_forwards", {"format": "xml"}),
        ("Email", "list_forwards", {"version": "2"}),
        ("Email", "list_forwards", {"version": "1"}),
    ]
    
    print("\n📋 Testing cPanel API Methods:")
    for i, (module, function, params) in enumerate(cpanel_methods, 1):
        print(f"\n{i}. Testing {module}::{function} with params: {params}")
        try:
            result = cpanel_api_call(server, cpanel_user, module, function, params)
            
            if result is None:
                print(f"   ❌ Returned None")
            elif isinstance(result, dict):
                if "error" in result:
                    print(f"   ❌ API Error: {result['error']}")
                elif "result" in result:
                    result_data = result["result"]
                    if isinstance(result_data, dict) and "data" in result_data:
                        forwards_data = result_data["data"]
                        if isinstance(forwards_data, list):
                            print(f"   ✅ Success - Found {len(forwards_data)} forward(s)")
                            if forwards_data:
                                print(f"   📋 Sample: {forwards_data[0]}")
                        else:
                            print(f"   ⚠️  Data is not a list: {type(forwards_data)}")
                    else:
                        print(f"   ⚠️  No data key or result not dict")
                else:
                    print(f"   ⚠️  No result key")
            else:
                print(f"   ⚠️  Unexpected type: {type(result)}")
                
        except Exception as e:
            print(f"   ❌ Exception: {str(e)}")
    
    # 2. اختبار APIs بديلة
    print(f"\n📋 Testing Alternative cPanel APIs:")
    alternative_apis = [
        ("Email", "list_pops_with_disk", {"domain": domain}),
        ("Email", "list_pops", {"domain": domain}),
        ("Email", "list_pops", {"user": cpanel_user}),
        ("Email", "list_pops", {"include": "forward"}),
        ("Email", "list_pops", {"include": "all"}),
        ("Email", "list_pops", {"type": "all"}),
        ("Email", "list_pops", {"type": "forward"}),
        ("Email", "list_pops", {"format": "json"}),
        ("Email", "list_pops", {"format": "xml"}),
        ("Email", "list_pops", {"version": "2"}),
        ("Email", "list_pops", {"version": "1"}),
    ]
    
    for i, (module, function, params) in enumerate(alternative_apis, 1):
        print(f"\n{i}. Testing {module}::{function} with params: {params}")
        try:
            result = cpanel_api_call(server, cpanel_user, module, function, params)
            
            if result is None:
                print(f"   ❌ Returned None")
            elif isinstance(result, dict):
                if "error" in result:
                    print(f"   ❌ API Error: {result['error']}")
                elif "result" in result:
                    result_data = result["result"]
                    if isinstance(result_data, dict) and "data" in result_data:
                        alt_data = result_data["data"]
                        if isinstance(alt_data, list):
                            print(f"   ✅ Success - Found {len(alt_data)} item(s)")
                            if alt_data:
                                print(f"   📋 Sample: {alt_data[0]}")
                                # فحص إذا كان يحتوي على Forward Rules
                                forward_found = False
                                for item in alt_data[:3]:  # فحص أول 3 عناصر
                                    if isinstance(item, dict):
                                        if "forward" in item or "forward_to" in item or "forwarder" in item:
                                            forward_found = True
                                            print(f"   🔍 Found forward info in: {item}")
                                            break
                                if forward_found:
                                    print(f"   ✅ Forward Rules found in this API!")
                                else:
                                    print(f"   ℹ️  No forward rules found in this API")
                        else:
                            print(f"   ⚠️  Data is not a list: {type(alt_data)}")
                    else:
                        print(f"   ⚠️  No data key or result not dict")
                else:
                    print(f"   ⚠️  No result key")
            else:
                print(f"   ⚠️  Unexpected type: {type(result)}")
                
        except Exception as e:
            print(f"   ❌ Exception: {str(e)}")
    
    # 3. اختبار WHM API
    print(f"\n📋 Testing WHM API:")
    whm_methods = [
        ("get_domain_forwarders", {"domain": domain}),
        ("get_domain_forwarders", {}),
        ("get_domain_forwarders", {"user": cpanel_user}),
        ("get_domain_forwarders", {"type": "forward"}),
        ("get_domain_forwarders", {"include": "forward"}),
        ("get_domain_info", {"domain": domain}),
        ("get_domain_info", {"user": cpanel_user}),
        ("get_domain_info", {}),
        ("get_domain_stats", {"domain": domain}),
        ("get_domain_stats", {"user": cpanel_user}),
    ]
    
    for i, (whm_function, whm_params) in enumerate(whm_methods, 1):
        print(f"\n{i}. Testing WHM API: {whm_function} with params: {whm_params}")
        try:
            whm_result = whm_api_call(server, whm_function, whm_params)
            if whm_result and "error" not in whm_result:
                if "data" in whm_result:
                    whm_data = whm_result["data"]
                    if isinstance(whm_data, list):
                        print(f"   ✅ WHM API Success - Found {len(whm_data)} item(s)")
                        if whm_data:
                            print(f"   📋 Sample: {whm_data[0]}")
                            # فحص إذا كان يحتوي على Forward Rules
                            forward_found = False
                            for item in whm_data[:3]:  # فحص أول 3 عناصر
                                if isinstance(item, dict):
                                    if "forward" in item or "forward_to" in item or "forwarder" in item:
                                        forward_found = True
                                        print(f"   🔍 Found forward info in: {item}")
                                        break
                            if forward_found:
                                print(f"   ✅ Forward Rules found in this WHM API!")
                            else:
                                print(f"   ℹ️  No forward rules found in this WHM API")
                    else:
                        print(f"   ⚠️  WHM data is not a list: {type(whm_data)}")
                else:
                    print(f"   ⚠️  WHM API no data key")
            else:
                print(f"   ❌ WHM API Error: {whm_result.get('error', 'Unknown error')}")
        except Exception as e:
            print(f"   ❌ WHM API Exception: {str(e)}")
    
    # 4. اختبار اتصال عام
    print(f"\n📋 Testing General Connection:")
    try:
        # اختبار API بسيط
        test_result = cpanel_api_call(server, cpanel_user, "Email", "list_pops", {"domain": domain})
        if test_result:
            print(f"   ✅ General cPanel API connection: Success")
            if isinstance(test_result, dict) and "result" in test_result:
                print(f"   📋 Available modules: Email")
        else:
            print(f"   ❌ General cPanel API connection: Failed")
    except Exception as e:
        print(f"   ❌ General connection exception: {str(e)}")
    
    print(f"\n🔍 Testing completed. Check the results above.")
    print(f"📊 Total methods tested: {len(cpanel_methods) + len(alternative_apis) + len(whm_methods)}")

# === دوال التصدير ===
def export_passwords_to_excel(data, filename):
    """تصدير كلمات المرور إلى ملف Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # إنشاء workbook جديد
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Passwords"
        
        # إضافة العناوين
        headers = ["Email", "Domain", "New Password", "Webmail URL", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # إضافة البيانات
        for row, item in enumerate(data, 2):
            ws.cell(row=row, column=1, value=item["Email"])
            ws.cell(row=row, column=2, value=item["Domain"])
            ws.cell(row=row, column=3, value=item["New Password"])
            ws.cell(row=row, column=4, value=item["Webmail URL"])
            
            # تلوين الحالة
            status_cell = ws.cell(row=row, column=5, value=item["Status"])
            if "Success" in item["Status"]:
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # ضبط عرض الأعمدة
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # حفظ الملف
        wb.save(filename)
        print(f"✅ Excel file saved: {filename}")
        
    except ImportError:
        print("❌ openpyxl not installed. Installing...")
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        print("✅ openpyxl installed. Please run the export again.")
    except Exception as e:
        print(f"❌ Error exporting to Excel: {str(e)}")

def export_passwords_to_csv(data, filename):
    """تصدير كلمات المرور إلى ملف CSV"""
    try:
        import csv
        
        with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = ["Email", "Domain", "New Password", "Webmail URL", "Status"]
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            
            writer.writeheader()
            for item in data:
                writer.writerow(item)
        
        print(f"✅ CSV file saved: {filename}")
        
    except Exception as e:
        print(f"❌ Error exporting to CSV: {str(e)}")

# === القائمة الرئيسية ===
def main():
    """الدالة الرئيسية"""
    try:
        servers = initialize_script("WHM Email Management & Monitoring")
        
        while True:
            print(f"\n{'='*20} EMAIL MANAGEMENT & MONITORING {'='*20}")
            print("📧 Basic Email Management:")
            print("1.  ➕ Create single email")
            print("2.  🔢 Bulk create emails")
            print("3.  🔐 Change email password(s)")
            print("4.  🗑️  Delete email")
            print("5.  📋 List & export emails")
            print("6.  📤 Manage email forward rules")
            
            print("\n📊 Email Monitoring & Analysis:")
            print("7.  📈 Email monitoring dashboard")
            print("8.  🚫 Failed emails analysis")
            print("9.  ⚠️  Spam accounts detection")
            print("10. 🔍 Blacklist status checker")
            print("11. 📮 Mail queue monitoring")
            print("12. 🎯 Quick email health check (All servers)")
            print("13. 📋 Complete email security audit")
            print("14. 🚨 Failed login attempts analysis")
            print("15. ⚙️ View email settings (Outlook/iPhone)")
            
            print("\n🔧 System Tools:")
            print("16. 🌐 Check server status")
            print("17. 📜 View operation logs")
            print("18. 🎲 Generate random password")
            
            print("\n0.  🚪 Exit")
            print("=" * 75)
            
            choice = input("Choose option: ").strip()

            if choice == "1":
                create_single_email(servers)
            elif choice == "2":
                bulk_create_emails(servers)
            elif choice == "3":
                change_email_passwords(servers)
            elif choice == "4":
                delete_email_accounts(servers)
            elif choice == "5":
                list_and_export_emails(servers)
            elif choice == "6":
                manage_email_forward_rules(servers)
            elif choice == "7":
                email_monitoring_dashboard(servers)
            elif choice == "8":
                failed_emails_analysis_menu(servers)
            elif choice == "9":
                spam_analysis_menu(servers)
            elif choice == "10":
                blacklist_check_menu(servers)
            elif choice == "11":
                mail_queue_status_menu(servers)
            elif choice == "12":
                quick_email_health_check_all_servers(servers)
            elif choice == "13":
                complete_email_audit_menu(servers)
            elif choice == "14":
                failed_login_analysis_menu(servers)
            elif choice == "15":
                show_email_settings(servers)
            elif choice == "16":
                display_server_status(servers)
            elif choice == "17":
                show_logs()
            elif choice == "18":
                length = input("🔢 Password length (default 12): ").strip() or "12"
                try:
                    length = int(length)
                    if length < 4:
                        print("❌ Password length must be at least 4")
                    else:
                        password = generate_password(length)
                        print(f"🎲 Generated password: {password}")
                except ValueError:
                    print("❌ Invalid length")
            elif choice == "0":
                print("👋 Goodbye!")
                logging.info("Email Management & Monitoring closed")
                break
            else:
                print("❌ Invalid option")

    except KeyboardInterrupt:
        print("\n\n⚠️  Operation cancelled by user")
        logging.info("Program interrupted by user")
        sys.exit(0)
    except Exception as e:
        handle_script_error(e, "Email Management & Monitoring")
        sys.exit(1)

if __name__ == "__main__":
    main()
